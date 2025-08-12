import os
import sys
import pandas as pd
import tkinter as tk
from tkinter import filedialog, ttk, messagebox
import threading
import time
import warnings
import logging
from openpyxl import load_workbook
import gc

import xlwt

# 忽略pandas的警告
warnings.filterwarnings('ignore')


# 配置日志
def setup_logging():
    # 获取当前脚本所在目录或exe所在目录
    if getattr(sys, 'frozen', False):
        # 如果是exe环境，获取exe文件所在目录
        script_dir = os.path.dirname(sys.executable)
    else:
        # 如果是开发环境，获取脚本文件所在目录
        script_dir = os.path.dirname(os.path.abspath(__file__))
    # 创建日志文件路径
    log_file = os.path.join(script_dir, 'excel_tool.log')

    # 创建logger
    logger = logging.getLogger('ExcelMergerSplitter')
    logger.setLevel(logging.DEBUG)

    # 创建文件处理器
    file_handler = logging.FileHandler(log_file, encoding='utf-8')
    file_handler.setLevel(logging.DEBUG)

    # 创建控制台处理器
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)

    # 创建格式器
    formatter = logging.Formatter(
        '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )

    # 设置格式器
    file_handler.setFormatter(formatter)
    console_handler.setFormatter(formatter)

    # 添加处理器到logger
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)

    return logger


# 创建全局logger实例
logger = setup_logging()


class ExcelMergerSplitterApp:
    def __init__(self, root):
        self.root = root
        self.root.title("表格文件合并与拆分工具")
        self.root.geometry("800x750")
        self.root.resizable(True, True)

        logger.info("应用程序启动")

        # 设置中文字体
        self.style = ttk.Style()
        self.style.configure("TLabel", font=("SimHei", 10))
        self.style.configure("TButton", font=("SimHei", 10))
        self.style.configure("TCheckbutton", font=("SimHei", 10))
        self.style.configure("TProgressbar", thickness=20)

        # 存储选中的文件和表头信息
        self.merge_files = []
        self.split_file = ""
        self.output_path = ""
        self.first_header = None  # 存储第一个文件的表头
        self._cached_columns = None  # 保存列名，避免重复加载

        # 拆分功能新增变量
        self.split_by_column = tk.BooleanVar(value=False)  # 是否按列拆分
        self.selected_column = tk.StringVar(value="")  # 选中的拆分列名
        self.split_to_sheets = tk.BooleanVar(value=False)  # 是否拆分为多个页签
        self.header_styles = {}  # 存储表头样式
        # 创建UI
        self.create_widgets()

    def create_widgets(self):
        # 创建主框架
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 标题
        title_label = ttk.Label(main_frame, text="表格文件合并与拆分工具", font=("SimHei", 16, "bold"))
        title_label.pack(pady=(0, 20))

        # 创建选项卡
        tab_control = ttk.Notebook(main_frame)

        # 合并功能选项卡
        merge_tab = ttk.Frame(tab_control)
        tab_control.add(merge_tab, text="文件合并")

        # 拆分功能选项卡
        split_tab = ttk.Frame(tab_control)
        tab_control.add(split_tab, text="文件拆分")

        tab_control.pack(expand=1, fill="both")

        # ==================== 合并功能UI ====================
        # 文件选择区域
        select_frame = ttk.LabelFrame(merge_tab, text="选择文件", padding="10")
        select_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        # 按钮区域
        btn_frame = ttk.Frame(select_frame)
        btn_frame.pack(fill=tk.X, pady=(0, 10))

        self.add_btn = ttk.Button(btn_frame, text="添加文件", command=self.add_merge_files)
        self.add_btn.pack(side=tk.LEFT, padx=(0, 10))

        self.remove_btn = ttk.Button(btn_frame, text="移除选中", command=self.remove_merge_files)
        self.remove_btn.pack(side=tk.LEFT, padx=(0, 10))

        self.clear_btn = ttk.Button(btn_frame, text="清空列表", command=self.clear_merge_files)
        self.clear_btn.pack(side=tk.LEFT)

        # 文件列表
        list_frame = ttk.Frame(select_frame)
        list_frame.pack(fill=tk.BOTH, expand=True)

        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.file_listbox = tk.Listbox(
            list_frame,
            yscrollcommand=scrollbar.set,
            selectmode=tk.EXTENDED,
            font=("SimHei", 10),
            height=10
        )
        self.file_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.file_listbox.yview)

        # 选项区域
        merge_options_frame = ttk.LabelFrame(merge_tab, text="合并选项", padding="10")
        merge_options_frame.pack(fill=tk.X, pady=(0, 10))

        # 表头处理选项（默认勾选且强制生效）
        self.header_var = tk.BooleanVar(value=True)
        header_check = ttk.Checkbutton(
            merge_options_frame,
            text="仅保留第一个文件的表头（强制生效）",
            variable=self.header_var,
            state="disabled"  # 禁用修改，确保表头处理逻辑一致
        )
        header_check.pack(anchor=tk.W, pady=(0, 5))

        # 输出文件路径
        output_frame = ttk.Frame(merge_options_frame)
        output_frame.pack(fill=tk.X, pady=(5, 0))

        ttk.Label(output_frame, text="输出文件:").pack(side=tk.LEFT, padx=(0, 10))

        self.merge_output_entry = ttk.Entry(output_frame)
        self.merge_output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        self.merge_output_entry.insert(0, os.path.join(os.getcwd(), "合并结果.xlsx"))

        self.merge_browse_btn = ttk.Button(output_frame, text="浏览...", command=self.browse_merge_output)
        self.merge_browse_btn.pack(side=tk.LEFT)

        # ==================== 拆分功能UI ====================
        # 选择要拆分的文件
        split_file_frame = ttk.LabelFrame(split_tab, text="选择文件", padding="10")
        split_file_frame.pack(fill=tk.X, pady=(0, 10))

        split_file_path_frame = ttk.Frame(split_file_frame)
        split_file_path_frame.pack(fill=tk.X, pady=(5, 0))

        ttk.Label(split_file_path_frame, text="待拆分文件:").pack(side=tk.LEFT, padx=(0, 10))

        self.split_file_entry = ttk.Entry(split_file_path_frame)
        self.split_file_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))

        self.split_browse_btn = ttk.Button(split_file_path_frame, text="浏览...", command=self.browse_split_file)
        self.split_browse_btn.pack(side=tk.LEFT)

        # 拆分选项
        split_options_frame = ttk.LabelFrame(split_tab, text="拆分选项", padding="10")
        split_options_frame.pack(fill=tk.X, pady=(0, 10))

        # 拆分方式选择
        split_method_frame = ttk.LabelFrame(split_options_frame, text="拆分方式", padding="5")
        split_method_frame.pack(fill=tk.X, pady=(5, 0))

        # 按行拆分单选
        self.row_split_radio = ttk.Radiobutton(
            split_method_frame,
            text="按条目数拆分",
            variable=self.split_by_column,
            value=False,
            command=self.toggle_split_method
        )
        self.row_split_radio.pack(anchor=tk.W, pady=(2, 2))

        # 按列拆分单选
        self.col_split_radio = ttk.Radiobutton(
            split_method_frame,
            text="按列值拆分",
            variable=self.split_by_column,
            value=True,
            command=self.toggle_split_method
        )
        self.col_split_radio.pack(anchor=tk.W, pady=(2, 2))

        # 条目数设置
        rows_frame = ttk.Frame(split_options_frame)
        rows_frame.pack(fill=tk.X, pady=(5, 0))

        ttk.Label(rows_frame, text="每个文件包含条目数:").pack(side=tk.LEFT, padx=(0, 10))

        self.rows_per_file_var = tk.StringVar(value="1000")
        self.rows_per_file_entry = ttk.Entry(rows_frame, textvariable=self.rows_per_file_var, width=10)
        self.rows_per_file_entry.pack(side=tk.LEFT, padx=(0, 10))

        # 列选择下拉框（初始禁用）
        column_frame = ttk.Frame(split_options_frame)
        column_frame.pack(fill=tk.X, pady=(5, 0))

        ttk.Label(column_frame, text="拆分列:").pack(side=tk.LEFT, padx=(0, 10))

        self.column_combobox = ttk.Combobox(
            column_frame,
            textvariable=self.selected_column,
            state="disabled"
        )
        self.column_combobox.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))

        # 输出目录
        split_output_frame = ttk.Frame(split_options_frame)
        split_output_frame.pack(fill=tk.X, pady=(5, 0))

        ttk.Label(split_output_frame, text="输出目录:").pack(side=tk.LEFT, padx=(0, 10))

        self.split_output_entry = ttk.Entry(split_output_frame)
        self.split_output_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))
        self.split_output_entry.insert(0, os.path.join(os.getcwd(), "拆分结果"))

        self.split_output_browse_btn = ttk.Button(split_output_frame, text="浏览...", command=self.browse_split_output)
        self.split_output_browse_btn.pack(side=tk.LEFT)

        # 表头处理选项
        self.split_header_var = tk.BooleanVar(value=True)
        split_header_check = ttk.Checkbutton(
            split_options_frame,
            text="每个拆分文件都包含表头",
            variable=self.split_header_var
        )
        split_header_check.pack(anchor=tk.W, pady=(5, 0))

        # 拆分结果选项
        result_option_frame = ttk.Frame(split_options_frame)
        result_option_frame.pack(fill=tk.X, pady=(5, 0))

        self.split_result_var = tk.BooleanVar(value=False)
        split_result_check = ttk.Checkbutton(
            result_option_frame,
            text="拆分为一个文件的多个页签（否则为多个文件）",
            variable=self.split_to_sheets
        )
        split_result_check.pack(anchor=tk.W, pady=(5, 0))

        # ==================== 公共进度区域 ====================
        progress_frame = ttk.LabelFrame(main_frame, text="处理进度", padding="10")
        progress_frame.pack(fill=tk.X, pady=(10, 10))

        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(
            progress_frame,
            variable=self.progress_var,
            maximum=100
        )
        self.progress_bar.pack(fill=tk.X, pady=(0, 5))

        self.status_label = ttk.Label(progress_frame, text="等待开始...")
        self.status_label.pack(anchor=tk.W)

        # 操作按钮
        action_frame = ttk.Frame(main_frame)
        action_frame.pack(fill=tk.X, pady=(10, 0))

        self.merge_btn = ttk.Button(
            action_frame,
            text="开始合并",
            command=self.start_merge
        )
        self.merge_btn.pack(side=tk.RIGHT, padx=(0, 10))

        self.split_btn = ttk.Button(
            action_frame,
            text="开始拆分",
            command=self.start_split
        )
        self.split_btn.pack(side=tk.RIGHT)

        # 支持格式说明
        format_frame = ttk.Frame(main_frame)
        format_frame.pack(fill=tk.X, pady=(10, 0))

        ttk.Label(
            format_frame,
            text="支持格式: .xlsx, .xls, .csv, .et (WPS表格)",
            foreground="gray"
        ).pack(anchor=tk.W)

    # ==================== 合并功能相关方法 ====================
    def add_merge_files(self):
        file_types = [
            ("所有支持的表格文件", "*.xlsx *.xls *.csv *.et"),
            ("Excel文件", "*.xlsx *.xls"),
            ("CSV文件", "*.csv"),
            ("WPS表格文件", "*.et"),
            ("所有文件", "*")
        ]

        files = filedialog.askopenfilenames(
            title="选择要合并的文件",
            filetypes=file_types
        )

        if files:
            for file in files:
                if file not in self.merge_files:
                    self.merge_files.append(file)
                    self.file_listbox.insert(tk.END, os.path.basename(file))

            logger.info(f"添加了 {len(files)} 个文件到合并列表")

    def remove_merge_files(self):
        selected_indices = self.file_listbox.curselection()
        if not selected_indices:
            messagebox.showwarning("警告", "请先选择要移除的文件")
            logger.warning("用户尝试移除文件但未选择任何文件")
            return

        # 从后往前删除，避免索引问题
        for i in sorted(selected_indices, reverse=True):
            removed_file = self.merge_files[i]
            self.file_listbox.delete(i)
            del self.merge_files[i]

        logger.info(f"从列表中移除了 {len(selected_indices)} 个文件")

    def clear_merge_files(self):
        if self.merge_files:
            confirm = messagebox.askyesno("确认", "确定要清空文件列表吗?")
            if confirm:
                self.file_listbox.delete(0, tk.END)
                self.merge_files = []
                logger.info("清空了合并文件列表")
        else:
            logger.debug("尝试清空空的文件列表")

    def browse_merge_output(self):
        initial_dir = os.path.dirname(self.merge_output_entry.get())
        initial_file = os.path.basename(self.merge_output_entry.get())

        file_path = filedialog.asksaveasfilename(
            title="保存合并结果",
            defaultextension=".xlsx",
            filetypes=[
                ("Excel文件", "*.xlsx"),
                ("Excel 97-2003文件", "*.xls"),
                ("CSV文件", "*.csv"),
                ("WPS表格文件", "*.et")
            ],
            initialdir=initial_dir,
            initialfile=initial_file
        )

        if file_path:
            self.merge_output_entry.delete(0, tk.END)
            self.merge_output_entry.insert(0, file_path)
            logger.info(f"设置合并输出路径: {file_path}")

    # ==================== 拆分功能相关方法 ====================
    def browse_split_file(self):
        file_types = [
            ("所有支持的表格文件", "*.xlsx *.xls *.csv *.et"),
            ("Excel文件", "*.xlsx *.xls"),
            ("CSV文件", "*.csv"),
            ("WPS表格文件", "*.et"),
            ("所有文件", "*")
        ]

        file_path = filedialog.askopenfilename(
            title="选择要拆分的文件",
            filetypes=file_types
        )

        if file_path:
            self.split_file_entry.delete(0, tk.END)
            self.split_file_entry.insert(0, file_path)
            # 自动设置输出目录
            dir_name = os.path.dirname(file_path)
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            output_path = os.path.join(dir_name, f"{base_name}_拆分结果")
            # 标准化路径分隔符
            output_path = os.path.normpath(output_path)
            self.split_output_entry.delete(0, tk.END)
            self.split_output_entry.insert(0, output_path)
            logger.info(f"选择了待拆分文件: {file_path}")
            # 一次性加载列名并缓存
            self._load_and_cache_columns(file_path)
            # # 如果当前是按列拆分模式，加载列名
            # if self.split_by_column.get():
            #     self.load_columns_from_file()

    def browse_split_output(self):
        initial_dir = self.split_output_entry.get() or os.getcwd()
        dir_path = filedialog.askdirectory(
            title="选择拆分结果保存目录",
            initialdir=initial_dir
        )

        if dir_path:
            self.split_output_entry.delete(0, tk.END)
            self.split_output_entry.insert(0, dir_path)
            logger.info(f"设置拆分输出目录: {dir_path}")

    def toggle_split_method(self):
        """切换拆分方式时启用/禁用相关控件"""
        if self.split_by_column.get():
            # 按列拆分：禁用行数输入，启用列选择
            self.rows_per_file_entry.config(state="disabled")
            self.column_combobox.config(state="readonly")

            # 不再重新加载，直接用缓存
            if self._cached_columns:
                self.column_combobox['values'] = self._cached_columns
                self.selected_column.set(self._cached_columns[0] if self._cached_columns else "")
            else:
                self.rows_per_file_entry.config(state="normal")
                self.column_combobox.config(state="disabled")
            # 尝试加载列名（如果已选择文件）
            # self.load_columns_from_file()
        else:
            # 按行拆分：启用行数输入，禁用列选择
            self.rows_per_file_entry.config(state="normal")
            self.column_combobox.config(state="disabled")

    def load_columns_from_file(self):
        """从选中的文件加载列名到下拉框"""
        split_file = self.split_file_entry.get()
        if not split_file or not os.path.exists(split_file):
            return

        try:
            # 读取文件获取表头
            df = self.read_table_file(split_file)
            if len(df) > 0:
                headers = df.iloc[0].tolist()
                # 过滤空表头
                headers = [h for h in headers if str(h).strip()]
                self.column_combobox['values'] = headers
                if headers:
                    self.selected_column.set(headers[0])
                logger.info(f"从文件加载了 {len(headers)} 个列名")
        except Exception as e:
            logger.error(f"加载列名失败: {str(e)}")
            messagebox.showerror("错误", f"加载列名失败: {str(e)}")

    # ==================== 公共方法 ====================
    def update_status(self, message):
        self.status_label.config(text=message)
        self.root.update_idletasks()
        logger.debug(f"状态更新: {message}")

    def update_progress(self, value):
        self.progress_var.set(value)
        self.root.update_idletasks()
        logger.debug(f"进度更新: {value}%")

    def _disable_all_buttons(self):
        """禁用所有操作按钮，防止重复操作"""
        self.add_btn.config(state=tk.DISABLED)
        self.remove_btn.config(state=tk.DISABLED)
        self.clear_btn.config(state=tk.DISABLED)
        self.merge_browse_btn.config(state=tk.DISABLED)
        self.merge_btn.config(state=tk.DISABLED)
        self.split_browse_btn.config(state=tk.DISABLED)
        self.split_output_browse_btn.config(state=tk.DISABLED)
        self.split_btn.config(state=tk.DISABLED)

    def _enable_all_buttons(self):
        """重新启用所有操作按钮"""
        self.add_btn.config(state=tk.NORMAL)
        self.remove_btn.config(state=tk.NORMAL)
        self.clear_btn.config(state=tk.NORMAL)
        self.merge_browse_btn.config(state=tk.NORMAL)
        self.merge_btn.config(state=tk.NORMAL)
        self.split_browse_btn.config(state=tk.NORMAL)
        self.split_output_browse_btn.config(state=tk.NORMAL)
        self.split_btn.config(state=tk.NORMAL)

    def read_table_file(self, file_path):
        """
        公共文件读取方法，支持多种表格格式
        返回包含所有数据的DataFrame（包含表头行）
        """
        file_ext = os.path.splitext(file_path)[1].lower()
        logger.info(f"开始解析文件: {file_path}，格式: {file_ext}")

        try:
            if file_ext == '.csv':
                # CSV文件处理（保持原逻辑）
                encodings = ['utf-8', 'gbk', 'gb2312', 'utf-16']
                df = None
                for encoding in encodings:
                    try:
                        df = pd.read_csv(
                            file_path,
                            encoding=encoding,
                            header=None,  # 不自动解析表头
                            dtype=str,  # 所有列按字符串读取
                            keep_default_na=False
                        )
                        logger.debug(f"成功使用 {encoding} 编码读取CSV文件")
                        break
                    except Exception as e:
                        logger.debug(f"使用 {encoding} 编码读取CSV失败: {str(e)}")
                        continue
                if df is None:
                    raise Exception(f"无法解析CSV文件，尝试了多种编码")
                return df

            elif file_ext in ['.xlsx', '.et']:
                # XLSX和WPS表格处理（包含您的格式处理逻辑）
                wb = load_workbook(file_path, read_only=True)
                ws = wb.active
                data = []
                for row in ws.iter_rows():
                    current_row = []
                    for cell in row:
                        value = cell.value
                        fmt = cell.number_format  # 获取单元格格式

                        if fmt is None:
                            current_row.append(str(value) if value is not None else "")
                        else:
                            if fmt != 'General':

                                if isinstance(value, (float, int)):
                                    if '.' in fmt:
                                        decimal_part = fmt.split('.')[-1]
                                        decimal_places = decimal_part.count('0')
                                        # decimal_places = len(
                                        #     decimal_part.replace('%', '').replace('#', '').replace('0', ''))  # 精确获取位数
                                        formatted_value = f"{value:.{decimal_places}f}"
                                    else:
                                        formatted_value = f"{value:.0f}"
                                    current_row.append(formatted_value)
                                elif hasattr(value, 'strftime'):  # 日期时间类型
                                    if ';' in fmt:
                                        date_fmt = fmt.split(';')[0]
                                    else:
                                        date_fmt = fmt
                                    try:
                                        # 转换Excel格式到strftime格式
                                        fmt_converted = (date_fmt.replace('yyyy', '%Y')
                                                         .replace('yy', '%y')
                                                         .replace('mm', '%m')
                                                         .replace('m', '%m')
                                                         .replace('dd', '%d')
                                                         .replace('d', '%d')
                                                         .replace('hh', '%H')
                                                         .replace('ss', '%S'))
                                        formatted_value = value.strftime(fmt_converted)
                                    except:
                                        formatted_value = str(value)
                                    current_row.append(formatted_value)
                                elif '%' in fmt and isinstance(value, (int, float)):  # 百分比类型
                                    decimal_places = fmt.count('0') - 1  # e.g. 0.00% -> 2
                                    formatted_value = f"{value:.{decimal_places}%}"
                                    current_row.append(formatted_value)
                                else:
                                    current_row.append(str(value) if value is not None else "")
                            else:
                                current_row.append(str(value) if value is not None else "")
                    data.append(current_row)
                wb.close()  # 关闭工作簿释放资源
                df = pd.DataFrame(data)
                logger.debug(f"成功解析XLSX/ET文件，共 {len(df)} 行")
                return df

            elif file_ext == '.xls':
                # XLS文件处理
                try:
                    # 尝试使用xlrd引擎（适用于真正的.xls文件）
                    df = pd.read_excel(
                        file_path,
                        header=None,
                        engine='xlrd',
                        dtype=str,
                        keep_default_na=False
                    )
                    logger.debug(f"成功使用xlrd引擎解析XLS文件，共 {len(df)} 行")
                except Exception as xlrd_error:
                    # 如果xlrd失败，可能是.xlsx文件被错误标记为.xls，尝试openpyxl
                    try:
                        df = pd.read_excel(
                            file_path,
                            header=None,
                            engine='openpyxl',
                            dtype=str,
                            keep_default_na=False
                        )
                        logger.debug(f"成功使用openpyxl引擎解析文件，共 {len(df)} 行")
                    except Exception as openpyxl_error:
                        # 如果两种引擎都失败，抛出更详细的错误信息
                        raise Exception(
                            f"无法解析XLS文件: xlrd错误: {str(xlrd_error)}, openpyxl错误: {str(openpyxl_error)}")

                return df

            else:
                raise Exception(f"不支持的文件格式: {file_ext}")

        except Exception as e:
            logger.error(f"解析文件 {file_path} 失败: {str(e)}")
            raise Exception(f"文件解析错误: {str(e)}")

    def read_table_file_chunked(self, file_path, chunksize=1000):
        """
        分块读取文件的方法，用于处理大文件
        """
        file_ext = os.path.splitext(file_path)[1].lower()
        logger.info(f"开始分块解析文件: {file_path}，格式: {file_ext}")

        if file_ext == '.csv':
            encodings = ['utf-8', 'gbk', 'gb2312', 'utf-16']
            for encoding in encodings:
                try:
                    for chunk in pd.read_csv(
                            file_path,
                            encoding=encoding,
                            header=None,
                            dtype=str,
                            keep_default_na=False,
                            chunksize=chunksize
                    ):
                        yield chunk
                    break
                except Exception as e:
                    logger.debug(f"使用 {encoding} 编码读取CSV失败: {str(e)}")
                    continue
            else:
                raise Exception(f"无法解析CSV文件，尝试了多种编码")

        elif file_ext in ['.xlsx', '.et']:
            df = self.read_table_file(file_path)
            total_rows = len(df)
            for i in range(0, total_rows, chunksize):
                yield df.iloc[i:i + chunksize]

        elif file_ext == '.xls':
            # 对于XLS文件，同样需要一次性读取
            df = self.read_table_file(file_path)
            total_rows = len(df)
            for i in range(0, total_rows, chunksize):
                yield df.iloc[i:i + chunksize]

        else:
            raise Exception(f"不支持的文件格式: {file_ext}")

    # ==================== 合并处理 ====================
    def start_merge(self):
        if not self.merge_files:
            messagebox.showwarning("警告", "请先添加要合并的文件")
            logger.warning("尝试开始合并但未选择任何文件")
            return

        output_path = self.merge_output_entry.get()
        if not output_path:
            messagebox.showwarning("警告", "请指定输出文件路径")
            logger.warning("尝试开始合并但未指定输出路径")
            return

        # 禁用按钮，防止重复操作
        self._disable_all_buttons()

        # 重置表头信息
        self.first_header = None

        # 在新线程中执行合并操作，避免UI卡顿
        merge_thread = threading.Thread(target=self.merge_files_proc, args=(output_path,))
        merge_thread.daemon = True
        merge_thread.start()
        logger.info("启动合并线程")

    def merge_files_proc(self, output_path):
        try:
            self.update_status("准备合并文件...")
            self.update_progress(0)
            logger.info(f"开始合并 {len(self.merge_files)} 个文件到 {output_path}")

            output_ext = os.path.splitext(output_path)[1].lower()
            first_file = True
            total_files = len(self.merge_files)

            # 根据输出文件类型选择合适的写入方式
            if output_ext == '.csv':
                with open(output_path, 'w', encoding='utf-8-sig', newline='') as f:
                    for i, file_path in enumerate(self.merge_files):
                        # 更新进度
                        progress = (i / total_files) * 100
                        self.update_progress(progress)
                        self.update_status(f"正在处理: {os.path.basename(file_path)} ({i + 1}/{total_files})")

                        logger.info(f"正在处理文件 {i + 1}/{total_files}: {file_path}")

                        try:
                            # 分块读取文件
                            for chunk_idx, df_chunk in enumerate(
                                    self.read_table_file_chunked(file_path, chunksize=1000)):
                                if df_chunk.empty:
                                    continue

                                # 处理表头
                                if first_file and chunk_idx == 0:
                                    # 第一个文件的第一块：保存表头并写入
                                    self.first_header = df_chunk.iloc[0].copy()
                                    df_chunk.to_csv(f, index=False, header=False)
                                    first_file = False
                                else:
                                    # 后续文件或块：只写入数据部分
                                    if chunk_idx == 0:
                                        # 第一块需要跳过表头行
                                        if len(df_chunk) > 1:
                                            df_chunk.iloc[1:].to_csv(f, index=False, header=False)
                                    else:
                                        # 后续块直接写入
                                        df_chunk.to_csv(f, index=False, header=False)

                                # 释放内存
                                del df_chunk
                                gc.collect()

                        except Exception as e:
                            self.update_status(f"处理 {os.path.basename(file_path)} 时出错: {str(e)}")
                            logger.error(f"处理文件 {file_path} 时出错: {str(e)}")
                            time.sleep(2)
                            continue

            elif output_ext in ['.xlsx', '.et']:
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    row_offset = 0

                    for i, file_path in enumerate(self.merge_files):
                        # 更新进度
                        progress = (i / total_files) * 100
                        self.update_progress(progress)
                        self.update_status(f"正在处理: {os.path.basename(file_path)} ({i + 1}/{total_files})")

                        logger.info(f"正在处理文件 {i + 1}/{total_files}: {file_path}")

                        try:
                            # 分块读取文件
                            first_chunk_of_file = True
                            for chunk_idx, df_chunk in enumerate(
                                    self.read_table_file_chunked(file_path, chunksize=1000)):
                                if df_chunk.empty:
                                    continue

                                # 处理表头
                                if first_file and first_chunk_of_file:
                                    # 第一个文件的第一块：保存表头并写入
                                    self.first_header = df_chunk.iloc[0].copy()
                                    df_chunk.to_excel(writer, index=False, sheet_name='Sheet1', startrow=row_offset,
                                                      header=False)
                                    row_offset += len(df_chunk)
                                    first_file = False
                                else:
                                    # 后续文件或块：只写入数据部分
                                    if first_chunk_of_file:
                                        # 第一块需要跳过表头行
                                        if len(df_chunk) > 1:
                                            data_chunk = df_chunk.iloc[1:]
                                            data_chunk.to_excel(writer, index=False, sheet_name='Sheet1',
                                                                startrow=row_offset, header=False)
                                            row_offset += len(data_chunk)
                                    else:
                                        # 后续块直接写入
                                        df_chunk.to_excel(writer, index=False, sheet_name='Sheet1', startrow=row_offset,
                                                          header=False)
                                        row_offset += len(df_chunk)

                                first_chunk_of_file = False

                                # 释放内存
                                del df_chunk
                                gc.collect()

                        except Exception as e:
                            self.update_status(f"处理 {os.path.basename(file_path)} 时出错: {str(e)}")
                            logger.error(f"处理文件 {file_path} 时出错: {str(e)}")
                            time.sleep(2)
                            continue

                    # 设置列名（表头）
                    if self.first_header is not None:
                        # 手动设置第一行作为表头
                        worksheet = writer.sheets['Sheet1']
                        for col_idx, header_val in enumerate(self.first_header):
                            worksheet.cell(row=1, column=col_idx + 1, value=header_val)
                            worksheet.cell(row=1, column=col_idx + 1).number_format = '@'

            elif output_ext == '.xls':
                with pd.ExcelWriter(output_path, engine='xlwt') as writer:
                    for i, file_path in enumerate(self.merge_files):
                        # 更新进度
                        progress = (i / total_files) * 100
                        self.update_progress(progress)
                        self.update_status(f"正在处理: {os.path.basename(file_path)} ({i + 1}/{total_files})")

                        logger.info(f"正在处理文件 {i + 1}/{total_files}: {file_path}")

                        try:
                            all_data = []
                            first_file_chunk = True

                            # 分块读取文件
                            for chunk_idx, df_chunk in enumerate(
                                    self.read_table_file_chunked(file_path, chunksize=1000)):
                                if df_chunk.empty:
                                    continue

                                # 处理表头
                                if first_file and first_file_chunk:
                                    # 第一个文件的第一块：保存表头
                                    self.first_header = df_chunk.iloc[0].copy()
                                    if len(df_chunk) > 1:
                                        all_data.append(df_chunk.iloc[1:].copy())
                                    first_file = False
                                else:
                                    # 后续文件或块：只添加数据部分
                                    if first_file_chunk:
                                        # 第一块需要跳过表头行
                                        if len(df_chunk) > 1:
                                            all_data.append(df_chunk.iloc[1:].copy())
                                    else:
                                        # 后续块直接添加
                                        all_data.append(df_chunk.copy())

                                first_file_chunk = False

                                # 释放内存
                                del df_chunk
                                gc.collect()

                            # 合并当前文件的所有数据块并写入
                            if all_data:
                                combined_chunk = pd.concat(all_data, ignore_index=True)
                                if i == 0:
                                    # 第一个文件写入表头
                                    combined_chunk.columns = self.first_header
                                    combined_chunk.to_excel(writer, index=False, sheet_name='Sheet1')
                                else:
                                    # 后续文件追加写入
                                    startrow = writer.sheets['Sheet1'].nrows
                                    worksheet = writer.sheets['Sheet1']
                                    # 写入数据
                                    for row_idx, row in combined_chunk.iterrows():
                                        for col_idx, value in enumerate(row):
                                            worksheet.write(
                                                startrow + row_idx,
                                                col_idx,
                                                value,
                                                xlwt.easyxf('text: format_string="@"')
                                            )

                                # 释放内存
                                del all_data, combined_chunk
                                gc.collect()

                        except Exception as e:
                            self.update_status(f"处理 {os.path.basename(file_path)} 时出错: {str(e)}")
                            logger.error(f"处理文件 {file_path} 时出错: {str(e)}")
                            time.sleep(2)
                            continue

            self.update_progress(100)
            self.update_status("合并完成!")
            logger.info("文件合并成功完成")
            messagebox.showinfo("成功", f"文件合并完成，已保存到:\n{output_path}")

        except MemoryError:
            self.update_status("合并失败: 内存不足")
            logger.error("合并过程出错: 内存不足", exc_info=True)
            messagebox.showerror("错误", "合并失败: 内存不足，请关闭其他程序或使用64位版本的Python")
        except Exception as e:
            self.update_status(f"合并失败: {str(e)}")
            logger.error(f"合并过程出错: {str(e)}", exc_info=True)
            messagebox.showerror("错误", f"合并失败: {str(e)}")
        finally:
            # 重新启用按钮
            self._enable_all_buttons()

    # ==================== 拆分处理 ====================
    def start_split(self):
        split_file = self.split_file_entry.get()
        if not split_file or not os.path.exists(split_file):
            messagebox.showwarning("警告", "请选择有效的待拆分文件")
            logger.warning("尝试开始拆分但未选择有效文件")
            return

        # 检查拆分方式
        if not self.split_by_column.get():
            # 按行数拆分
            try:
                rows_per_file = int(self.rows_per_file_var.get())
                if rows_per_file <= 0:
                    raise ValueError("条目数必须为正数")
            except ValueError as e:
                messagebox.showwarning("警告", f"无效的条目数: {str(e)}")
                logger.warning(f"无效的条目数: {str(e)}")
                return
        else:
            # 按列拆分
            if not self.selected_column.get():
                messagebox.showwarning("警告", "请选择拆分列")
                logger.warning("尝试按列拆分但未选择列")
                return

        output_dir = self.split_output_entry.get()
        if not output_dir:
            messagebox.showwarning("警告", "请指定输出目录")
            logger.warning("尝试开始拆分但未指定输出目录")
            return

        # 确保输出目录存在
        os.makedirs(output_dir, exist_ok=True)

        # 禁用按钮，防止重复操作
        self._disable_all_buttons()

        # 在新线程中执行拆分操作
        split_thread = threading.Thread(
            target=self.split_file_proc,
            args=(split_file, output_dir)
        )
        split_thread.daemon = True
        split_thread.start()
        logger.info("启动拆分线程")

    def split_file_proc(self, split_file, output_dir):
        try:
            self.update_status("准备拆分文件...")
            self.update_progress(0)
            # 初始化 message 变量
            message = ""
            file_ext = os.path.splitext(split_file)[1].lower()
            file_name = os.path.splitext(os.path.basename(split_file))[0]

            # 读取文件表头
            self.update_status("正在读取文件表头...")
            logger.info(f"读取文件表头: {split_file}")

            try:
                # 只读取第一行获取表头
                header_df = self.read_table_file(split_file)
                header = header_df.iloc[0]
            except Exception as e:
                raise Exception(f"读取文件表头失败: {str(e)}")

            logger.info(f"文件表头读取完成")
            self.update_status(f"文件表头读取完成")

            # 根据拆分方式执行不同逻辑
            if not self.split_by_column.get():
                # 按行数拆分
                rows_per_file = int(self.rows_per_file_var.get())
                logger.info(f"开始按行数拆分: 每个文件 {rows_per_file} 行")

                # 计算总行数
                total_rows = 0
                try:
                    if file_ext == '.csv':
                        # 对于CSV文件，使用pandas计算行数
                        encodings = ['utf-8', 'gbk', 'gb2312', 'utf-16']
                        for encoding in encodings:
                            try:
                                total_rows = sum(1 for row in pd.read_csv(
                                    split_file,
                                    encoding=encoding,
                                    chunksize=10000
                                ))
                                break
                            except:
                                continue
                    elif file_ext in ['.xlsx', '.et']:
                        # 对于Excel文件，使用openpyxl计算行数
                        wb = load_workbook(split_file, read_only=True)
                        ws = wb.active
                        total_rows = ws.max_row
                        wb.close()
                    elif file_ext == '.xls':
                        df = pd.read_excel(
                            split_file,
                            engine='xlrd'
                        )
                        total_rows = len(df)
                        del df  # 立即释放内存
                        gc.collect()  # 强制垃圾回收
                except Exception as e:
                    logger.warning(f"无法准确计算总行数: {e}")
                    total_rows = 100000  # 估计一个较大的值

                if total_rows <= 1:
                    raise Exception("待拆分文件只有表头，没有实际数据")

                # 减去表头行，计算实际数据行数
                data_rows = total_rows - 1
                total_chunks = (data_rows + rows_per_file - 1) // rows_per_file
                logger.info(f"总行数: {total_rows}, 数据行数: {data_rows}, 预计分片数: {total_chunks}")

                # 判断是否拆分为一个文件的多个页签
                if self.split_to_sheets.get():
                    # 多页签模式 - 只生成一个文件
                    output_path = os.path.join(output_dir, f"{file_name}_split_by_rows{file_ext}")

                    # 只支持xlsx和et格式的多页签
                    if file_ext in ['.xlsx', '.et']:
                        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                            chunk_count = 0
                            row_processed = 0

                            # 分块读取并处理
                            for df_chunk in self.read_table_file_chunked(split_file, chunksize=rows_per_file * 2):
                                if df_chunk.empty:
                                    continue

                                # 跳过表头（如果在第一块中）
                                if row_processed == 0 and len(df_chunk) > 0:
                                    df_chunk = df_chunk.iloc[1:].copy()

                                if df_chunk.empty:
                                    continue

                                # 处理当前块
                                for _, row in df_chunk.iterrows():
                                    # 当达到指定行数时，保存一个分片
                                    if row_processed % rows_per_file == 0 and row_processed > 0:
                                        chunk_count += 1

                                    row_processed += 1

                                # 释放内存
                                del df_chunk
                                gc.collect()

                            # 重新读取并写入文件
                            chunk_count = 0
                            row_processed = 0
                            current_data = []

                            for df_chunk in self.read_table_file_chunked(split_file, chunksize=rows_per_file * 2):
                                if df_chunk.empty:
                                    continue

                                # 跳过表头（如果在第一块中）
                                if row_processed == 0 and len(df_chunk) > 0:
                                    df_chunk = df_chunk.iloc[1:].copy()

                                if df_chunk.empty:
                                    continue

                                for _, row in df_chunk.iterrows():
                                    current_data.append(row)
                                    row_processed += 1

                                    # 当达到指定行数时，保存一个分片
                                    if len(current_data) >= rows_per_file:
                                        chunk_df = pd.DataFrame(current_data)
                                        chunk_df.columns = header

                                        # 页签名
                                        sheet_name = f'第{chunk_count + 1}部分'
                                        chunk_df.to_excel(writer, index=False, sheet_name=sheet_name)
                                        # 设置文本格式
                                        worksheet = writer.sheets[sheet_name]
                                        for column in worksheet.columns:
                                            for cell in column:
                                                cell.number_format = '@'

                                        chunk_count += 1
                                        current_data = []

                                        # 更新进度
                                        progress = min((row_processed / data_rows) * 100, 100)
                                        self.update_progress(progress)
                                        self.update_status(f"已完成 {chunk_count}/{total_chunks} 个页签")

                                # 释放内存
                                del df_chunk
                                gc.collect()

                            # 处理最后一个块
                            if current_data:
                                chunk_df = pd.DataFrame(current_data)
                                chunk_df.columns = header

                                sheet_name = f'第{chunk_count + 1}部分'
                                chunk_df.to_excel(writer, index=False, sheet_name=sheet_name)
                                # 设置文本格式
                                worksheet = writer.sheets[sheet_name]
                                for column in worksheet.columns:
                                    for cell in column:
                                        cell.number_format = '@'

                                chunk_count += 1

                                # 更新进度
                                self.update_progress(100)
                                self.update_status(f"已完成 {chunk_count}/{total_chunks} 个页签")

                        message = f"文件拆分完成，生成1个文件包含 {chunk_count} 个页签"
                    else:
                        # 不支持多页签的格式，自动切换为多文件模式
                        logger.warning("不支持多页签的格式，自动切换为多文件模式")
                        self._split_to_multiple_files(split_file, output_dir, file_name, file_ext, header,
                                                      rows_per_file, data_rows, total_chunks)

                else:
                    self._split_to_multiple_files(split_file, output_dir, file_name, file_ext, header, rows_per_file,
                                                  data_rows, total_chunks)

            else:
                # 按列值拆分
                column_name = self.selected_column.get()
                logger.info(f"开始按列值拆分: 列名 '{column_name}'")

                # 检查列是否存在
                if column_name not in header.values:
                    raise Exception(f"文件中不存在列: {column_name}")

                # 使用字典存储每个唯一值的数据
                unique_data = {}
                row_count = 0

                # 分块读取文件
                for df_chunk in self.read_table_file_chunked(split_file, chunksize=1000):
                    if df_chunk.empty:
                        continue

                    # 设置列名
                    df_chunk.columns = header

                    # 跳过表头（如果在第一块中）
                    if row_count == 0 and len(df_chunk) > 0:
                        df_chunk = df_chunk.iloc[1:].copy()

                    if df_chunk.empty:
                        continue

                    # 按列值分组存储数据
                    for _, row in df_chunk.iterrows():
                        value = row[column_name]
                        if pd.notna(value) and str(value).strip():
                            if value not in unique_data:
                                unique_data[value] = []
                            unique_data[value].append(row)

                        row_count += 1

                    # 释放内存
                    del df_chunk
                    gc.collect()

                if not unique_data:
                    raise Exception(f"列 '{column_name}' 没有有效的唯一值")

                total_chunks = len(unique_data)
                logger.info(f"发现 {total_chunks} 个唯一值，将进行拆分")

                # 拆分为多个页签
                if self.split_to_sheets.get():
                    output_path = os.path.join(output_dir, f"{file_name}_split_by_{column_name}{file_ext}")
                    if file_ext in ['.xlsx', '.et']:
                        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                            for i, (value, rows) in enumerate(unique_data.items()):
                                # 创建DataFrame
                                chunk = pd.DataFrame(rows)
                                chunk.columns = header

                                # 写入页签
                                sheet_name = f"{value}"[:31]  # Excel页签最大31字符
                                chunk.to_excel(writer, index=False, sheet_name=sheet_name)

                                # 设置文本格式
                                worksheet = writer.sheets[sheet_name]
                                for column in worksheet.columns:
                                    for cell in column:
                                        cell.number_format = '@'

                                # 更新进度
                                progress = ((i + 1) / total_chunks) * 100
                                self.update_progress(progress)
                                self.update_status(f"已完成 {i + 1}/{total_chunks} 个页签")

                                # 释放内存
                                del chunk, rows
                                gc.collect()

                        message = f"文件拆分完成，生成1个文件包含 {total_chunks} 个页签"
                    else:
                        raise Exception("只有.xlsx和.et格式支持多页签拆分")

                # 拆分为多个文件
                else:
                    for i, (value, rows) in enumerate(unique_data.items()):
                        # 创建DataFrame
                        chunk = pd.DataFrame(rows)
                        chunk.columns = header

                        # 生成文件名（处理特殊字符）
                        safe_value = self.sanitize_filename(str(value))
                        output_filename = f"{file_name}_{column_name}_{safe_value}{file_ext}"
                        output_path = os.path.join(output_dir, output_filename)

                        self.save_split_chunk(chunk, header, output_path, file_ext)

                        # 更新进度
                        progress = ((i + 1) / total_chunks) * 100
                        self.update_progress(progress)
                        self.update_status(f"已完成 {i + 1}/{total_chunks} 个文件")

                        # 释放内存
                        del chunk, rows
                        gc.collect()

                    message = f"文件拆分完成，共生成 {total_chunks} 个文件"

            self.update_progress(100)
            self.update_status("拆分完成!")
            logger.info("文件拆分成功完成")
            messagebox.showinfo("成功", f"{message}，已保存到:\n{output_dir}")

        except MemoryError:
            self.update_status("拆分失败: 内存不足")
            logger.error("拆分过程出错: 内存不足", exc_info=True)
            messagebox.showerror("错误", "拆分失败: 内存不足，请关闭其他程序或使用64位版本的Python")
        except Exception as e:
            self.update_status(f"拆分失败: {str(e)}")
            logger.error(f"拆分过程出错: {str(e)}", exc_info=True)
            messagebox.showerror("错误", f"拆分失败: {str(e)}")
        finally:
            # 重新启用按钮
            self._enable_all_buttons()

    def _split_to_multiple_files(self, split_file, output_dir, file_name, file_ext, header, rows_per_file, data_rows,
                                 total_chunks):
        """
        拆分为多个文件的辅助方法
        """
        chunk_count = 0
        row_processed = 0
        current_data = []

        # 分块读取并处理
        for df_chunk in self.read_table_file_chunked(split_file, chunksize=rows_per_file * 2):
            if df_chunk.empty:
                continue

            # 跳过表头（如果在第一块中）
            if row_processed == 0 and len(df_chunk) > 0:
                df_chunk = df_chunk.iloc[1:].copy()

            if df_chunk.empty:
                continue

            for _, row in df_chunk.iterrows():
                current_data.append(row)
                row_processed += 1

                # 当达到指定行数时，保存一个分片
                if len(current_data) >= rows_per_file:
                    chunk_df = pd.DataFrame(current_data)
                    chunk_df.columns = header

                    output_filename = f"{file_name}_part_{chunk_count + 1}{file_ext}"
                    output_path = os.path.join(output_dir, output_filename)
                    self.save_split_chunk(chunk_df, header, output_path, file_ext)

                    chunk_count += 1
                    current_data = []

                    # 更新进度
                    progress = min((row_processed / data_rows) * 100, 100)
                    self.update_progress(progress)
                    self.update_status(f"已完成 {chunk_count}/{total_chunks} 个分片")

            # 释放内存
            del df_chunk
            gc.collect()

        # 处理最后一个块
        if current_data:
            chunk_df = pd.DataFrame(current_data)
            chunk_df.columns = header

            output_filename = f"{file_name}_part_{chunk_count + 1}{file_ext}"
            output_path = os.path.join(output_dir, output_filename)
            self.save_split_chunk(chunk_df, header, output_path, file_ext)

            chunk_count += 1

            # 更新进度
            self.update_progress(100)
            self.update_status(f"已完成 {chunk_count}/{total_chunks} 个分片")

    def save_split_chunk(self, chunk, header, output_path, file_ext):
        """保存拆分后的块数据"""
        try:
            if file_ext == '.csv':
                chunk.to_csv(
                    output_path,
                    index=False,
                    encoding='utf-8-sig'
                )
            elif file_ext in ['.xlsx', '.et']:
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    chunk.to_excel(writer, index=False, sheet_name='Sheet1')
                    # 设置所有单元格为文本格式
                    worksheet = writer.sheets['Sheet1']
                    for column in worksheet.columns:
                        for cell in column:
                            cell.number_format = '@'  # 文本格式
            elif file_ext == '.xls':
                import xlwt
                workbook = xlwt.Workbook()
                worksheet = workbook.add_sheet('Sheet1')

                # 写入表头
                for col_idx, header_val in enumerate(header):
                    worksheet.write(0, col_idx, header_val)

                # 写入数据
                for row_idx, row in chunk.iterrows():
                    for col_idx, value in enumerate(row):
                        worksheet.write(row_idx + 1, col_idx, value)

                workbook.save(output_path)
            logger.debug(f"保存分片文件成功: {output_path}")
        except Exception as e:
            raise Exception(f"保存分片失败: {str(e)}")

    def sanitize_filename(self, filename):
        # 清理文件名中的非法字符
        # Windows中不允许的字符: \ / : * ? " < > |

        illegal_chars = ['\\', '/', ':', '*', '?', '"', '<', '>', '|']
        sanitized = filename
        for char in illegal_chars:
            sanitized = sanitized.replace(char, '_')
        # 去除首尾空格和点号
        sanitized = sanitized.strip('. ')
        # 如果文件名为空，返回默认名称
        if not sanitized:
            sanitized = 'unnamed'
        # 限制文件名长度（Windows最大255字符）
        if len(sanitized) > 200:  # 留一些空间给文件扩展名
            sanitized = sanitized[:200]
        return sanitized

    def _load_and_cache_columns(self, file_path):
        """读取文件并缓存列名"""
        try:
            file_ext = os.path.splitext(file_path)[1].lower()

            if file_ext in ['.xlsx', '.et']:
                # 对于Excel文件，只读取第一行来提高性能
                wb = load_workbook(file_path, read_only=True)
                ws = wb.active
                # 只读取第一行
                header_row = next(ws.iter_rows(max_row=1))
                headers = [str(cell.value) if cell.value is not None else "" for cell in header_row]
                wb.close()
            elif file_ext == '.xls':
                # 对于XLS文件，只读取前几行
                try:
                    df = pd.read_excel(
                        file_path,
                        header=None,
                        engine='xlrd',
                        nrows=1,  # 只读取第一行
                        dtype=str,
                        keep_default_na=False
                    )
                except Exception:
                    df = pd.read_excel(
                        file_path,
                        header=None,
                        engine='openpyxl',
                        nrows=1,  # 只读取第一行
                        dtype=str,
                        keep_default_na=False
                    )
                headers = df.iloc[0].tolist() if not df.empty else []
            elif file_ext == '.csv':
                # 对于CSV文件，只读取第一行
                encodings = ['utf-8', 'gbk', 'gb2312', 'utf-16']
                headers = []
                for encoding in encodings:
                    try:
                        df = pd.read_csv(
                            file_path,
                            encoding=encoding,
                            header=None,
                            nrows=1,  # 只读取第一行
                            dtype=str,
                            keep_default_na=False
                        )
                        headers = df.iloc[0].tolist() if not df.empty else []
                        break
                    except Exception:
                        continue
            else:
                # 对于其他格式，回退到原来的完整读取方法
                df = self.read_table_file(file_path)
                headers = df.iloc[0].tolist() if not df.empty else []

            # 过滤空表头
            headers = [h for h in headers if str(h).strip()]
            self._cached_columns = headers
            # 填充下拉框
            self.column_combobox['values'] = headers
            if headers:
                self.selected_column.set(headers[0])
            logger.info(f"已缓存列名：{len(headers)} 个")
        except Exception as e:
            logger.error(f"加载列名失败：{e}")
            self._cached_columns = []


if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelMergerSplitterApp(root)
    root.mainloop()
