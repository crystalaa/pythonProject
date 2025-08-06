import sys
import traceback
import logging
import os
import pandas as pd
from concurrent.futures import ThreadPoolExecutor
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QFileDialog, QLabel, QVBoxLayout, QHBoxLayout, \
    QPlainTextEdit, QProgressBar, QTabWidget, QComboBox, QProgressDialog
from PyQt5.QtCore import QThread, pyqtSignal, Qt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import xlrd

# 配置日志记录器
logging.basicConfig(
    filename="../jiangsu/error_log.txt",
    level=logging.ERROR,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

def read_rules(file_path):
    """读取规则文件，返回规则字典"""
    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        ws = wb['比对规则']
        rules = {}

        for row in ws.iter_rows(min_row=2, values_only=True):  # 假设第一行是标题
            table1_field, table2_field, data_type, tail_diff, is_primary, calc_rule = row[:6]
            if table1_field is None or table2_field is None:
                continue  # 跳过空行
            rules[table1_field] = {
                "table2_field": table2_field,
                "data_type": data_type.lower(),
                "tail_diff": tail_diff,
                "is_primary": is_primary == "是",
                "calc_rule": calc_rule  # 新增：存储计算规则
            }
        wb.close()
        return rules
    except Exception as e:
        raise Exception(f"读取规则文件时发生错误: {str(e)}")


def resource_path(relative_path):
    """获取打包后资源的绝对路径"""
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("../work"), relative_path)


def read_excel_columns(file_path, sheet_name):
    """快速读取Excel文件的列名"""
    try:
        if not sheet_name:  # 空字符串、None 都视为未选择
            return

        if file_path.lower().endswith('.xls'):
            # 处理 .xls 文件
            import xlrd
            try:
                workbook = xlrd.open_workbook(file_path)
                worksheet = workbook.sheet_by_name(sheet_name)
                if worksheet.nrows > 0:
                    columns = [cell.value for cell in worksheet.row(0)]
                    cleaned_columns = [col.replace('*', '').strip() if isinstance(col, str) else col for col in columns]
                    return cleaned_columns
            except xlrd.biffh.XLRDError as e:
                # 如果是版本问题，尝试用 openpyxl 处理（可能是实际为 .xlsx 格式的文件）
                if "xlsx file" in str(e).lower():
                    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
                    ws = wb[sheet_name]
                    columns = [cell.value for cell in next(ws.iter_rows())]
                    cleaned_columns = [col.replace('*', '').strip() if isinstance(col, str) else col for col in columns]
                    wb.close()
                    return cleaned_columns
                else:
                    raise e
        else:
            # 处理 .xlsx 文件
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            ws = wb[sheet_name]
            columns = [cell.value for cell in next(ws.iter_rows())]
            cleaned_columns = [col.replace('*', '').strip() if isinstance(col, str) else col for col in columns]
            wb.close()
            return cleaned_columns
    except Exception as e:
        raise Exception(f"读取Excel列名时发生错误: {str(e)}")


def read_excel_fast(file_path, sheet_name):
    """快速读取Excel文件"""
    try:
        if file_path.lower().endswith('.xls'):
            # 处理 .xls 文件

            workbook = xlrd.open_workbook(file_path)
            worksheet = workbook.sheet_by_name(sheet_name)

            data = []
            columns = None
            for i in range(worksheet.nrows):
                row = worksheet.row(i)
                if i == 0:
                    columns = [cell.value for cell in row]
                else:
                    data.append([cell.value for cell in row])

            return pd.DataFrame(data, columns=columns)
        else:
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            ws = wb[sheet_name]
            data = []
            columns = None
            for i, row in enumerate(ws.rows):
                if i == 0:
                    columns = [cell.value for cell in row]
                else:
                    data.append([cell.value for cell in row])
            wb.close()
            return pd.DataFrame(data, columns=columns)
    except Exception as e:
        raise Exception(f"读取Excel文件时发生错误: {str(e)}")


def get_sheet_names(file_path):
    """获取 Excel 文件的所有页签名称"""
    try:

        if file_path.lower().endswith('.xls'):
            # 处理 .xls 文件
            # 处理 .xls 文件
            import xlrd
            try:
                workbook = xlrd.open_workbook(file_path)
                sheetnames = workbook.sheet_names()
                return sheetnames
            except xlrd.biffh.XLRDError as e:
                # 如果是版本问题，尝试用 openpyxl 处理（可能是实际为 .xlsx 格式的文件）
                if "xlsx file" in str(e).lower():
                    wb = load_workbook(filename=file_path, read_only=True)
                    sheetnames = wb.sheetnames
                    wb.close()
                    return sheetnames
                else:
                    raise e
        else:
            # 处理 .xlsx 文件
            wb = load_workbook(filename=file_path, read_only=True)
            sheetnames = wb.sheetnames
            wb.close()
            return sheetnames
    except Exception as e:
        raise Exception(f"读取页签名称时发生错误: {str(e)}")


class LoadColumnWorker(QThread):
    """用于在独立线程中读取列名"""
    columns_loaded = pyqtSignal(str, list)  # 参数为文件路径和列名列表
    error_occurred = pyqtSignal(str)
    sheet_names_loaded = pyqtSignal(str, list)

    def __init__(self, file_path, sheet_name=None):
        super().__init__()
        self.file_path = file_path
        self.sheet_name = sheet_name

    def run(self):
        try:
            # 读取页签名称
            sheet_names = get_sheet_names(self.file_path)
            self.sheet_names_loaded.emit(self.file_path, sheet_names)
            if not self.sheet_name:
                return
            # 读取列名
            columns = read_excel_columns(self.file_path, self.sheet_name)
            if columns is None:
                return
            self.columns_loaded.emit(self.file_path, columns)
        except Exception as e:
            self.error_occurred.emit(str(e))


def read_mapping_table(file_path):
    """读取资产分类映射表，返回 DataFrame"""
    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        if '资产分类映射表' not in wb.sheetnames:
            raise Exception("未找到'资产分类映射表'页签")

        ws = wb['资产分类映射表']

        # 只读取第二行作为表头（第二级表头）
        headers = [cell.value if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=2, max_row=2))]

        # 读取数据行（从第3行开始）
        data = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            data.append(row)

        df = pd.DataFrame(data, columns=headers)
        wb.close()
        return df
    except Exception as e:
        raise Exception(f"读取资产分类映射表时发生错误: {str(e)}")



class CompareWorker(QThread):
    """用于在独立线程中执行比较操作"""
    log_signal = pyqtSignal(str)
    progress_signal = pyqtSignal(int)  # 用于更新进度条

    def __init__(self, file1, file2, rule_file, sheet_name1, sheet_name2, primary_keys=None, rules=None):
        super().__init__()
        self.file1 = file1
        self.file2 = file2
        self.rule_file = rule_file
        self.sheet_name1 = sheet_name1
        self.sheet_name2 = sheet_name2
        self.primary_keys = primary_keys if primary_keys else []
        self.rules = rules if rules else {}
        self.missing_assets = []
        self.diff_records = []
        self.summary = {}
        self.missing_rows = []
        self.extra_in_file2 = []
        self.diff_full_rows = []

    @staticmethod
    def normalize_value(val):
        """统一空值表示"""
        if pd.isna(val) or val is None or (isinstance(val, str) and str(val).strip() == ''):
            return ''
        return str(val).strip()

    def calculate_field(self, df, calc_rule, data_type):
        """
        根据计算规则和数据类型生成表二字段值
        支持：
        - 文本类型：字符串拼接（如"公司代码+资产编码"）
        - 数值类型：数值运算（如"使用年限+使用期间/12"）
        - 字符串截取（如"WBS编码[:12]"）
        - 字段运算（如"累计购置值-累计折旧额"）
        """
        if not calc_rule:
            return None

        try:
            # 处理字符串截取（通用逻辑）
            if '[:' in calc_rule and ']' in calc_rule:
                field, length_str = calc_rule.split('[:')
                field = field.strip()
                length = int(length_str.strip(']').strip())
                if field not in df.columns:
                    raise Exception(f"字段不存在：{field}")
                # 截取字符串前N位（处理空值）
                return df[field].fillna('').astype(str).str[:length]

            # 根据数据类型处理不同运算
            if data_type == "文本":
                # 文本类型：+表示字符串拼接
                fields = [f.strip() for f in calc_rule.split('+')]
                # 检查所有字段是否存在
                missing_fields = [f for f in fields if f not in df.columns]
                if missing_fields:
                    raise Exception(f"表达式中包含不存在的字段：{missing_fields}")
                # 字符串拼接（空值处理为空白字符串）
                result = df[fields[0]].fillna('').astype(str)
                for field in fields[1:]:
                    result += df[field].fillna('').astype(str)
                return result
            elif data_type == "数值":
                # 数值类型：执行数值运算
                import re
                field_pattern = re.compile(r'[a-zA-Z\u4e00-\u9fa5]+')
                fields_in_rule = field_pattern.findall(calc_rule)
                missing_fields = [f for f in fields_in_rule if f not in df.columns]
                if missing_fields:
                    raise Exception(f"表达式中包含不存在的字段：{missing_fields}")

                # 转换为数值类型，空值处理为0
                df_numeric = df.copy()
                for field in fields_in_rule:
                    # 如果字段名包含"折旧"，取绝对值
                    if "折旧" in field:
                        df_numeric[field] = pd.to_numeric(df[field], errors='coerce').fillna(0).abs()
                    else:
                        df_numeric[field] = pd.to_numeric(df[field], errors='coerce').fillna(0)

                # 执行计算
                result = df_numeric.eval(calc_rule)
                return result
            else:
                raise Exception(f"不支持的数据类型：{data_type}，请指定为'文本'或'数值'")

        except Exception as e:
            raise Exception(f"计算规则执行失败（{calc_rule}）：{str(e)}")

    def _get_value(self, df, part):
        """辅助函数：解析运算中的值（可能是字段或数值）"""
        part = part.strip()
        # 检查是否为字段
        if part in df.columns:
            # 尝试转换为数值（处理空值为0）
            return pd.to_numeric(df[part], errors='coerce').fillna(0)
        # 尝试解析为数值（如"12"）
        try:
            return float(part)
        except ValueError:
            raise Exception(f"无法解析值：{part}（不是字段也不是数值）")

    def values_equal_by_rule(self, val1, val2, data_type, tail_diff, field_name=""):
        """
        根据规则判断两个值是否相等
        """
        # 统一空值表示
        val1 = self.normalize_value(val1)
        val2 = self.normalize_value(val2)

        # 如果两个值都是空，则认为相等
        if val1 == "" and val2 == "":
            return True

        if data_type == "数值":
            # 数值型比较
            num1 = pd.to_numeric(val1, errors='coerce')
            num2 = pd.to_numeric(val2, errors='coerce')

            # 检查是否为NaN
            if pd.isna(num1) and pd.isna(num2):
                return True
            elif pd.isna(num1) or pd.isna(num2):
                return False

            # 如果字段名包含"折旧"，取绝对值
            field_contains_depreciation = "折旧" in field_name
            if field_contains_depreciation:
                num1 = abs(num1)
                num2 = abs(num2)

            if tail_diff is None:
                return num1 == num2
            else:
                return abs(num1 - num2) <= float(tail_diff)

        elif data_type == "日期":
            # 日期型比较
            def parse_date(date_str):
                """尝试多种格式解析日期，返回标准化字符串（YYYY-MM-DD）"""
                if pd.isna(date_str) or str(date_str).strip() == "":
                    return ""
                date_str = str(date_str).strip()
                if not date_str:
                    return ""
                # 尝试多种常见日期格式
                formats = ['%Y-%m-%d', '%Y/%m/%d', '%Y年%m月%d日',
                           '%m-%d-%Y', '%m/%d/%Y', '%Y%m%d']
                for fmt in formats:
                    try:
                        return pd.to_datetime(date_str, format=fmt).strftime('%Y-%m-%d')
                    except (ValueError, TypeError):
                        continue
                # 如果所有格式都解析失败，返回原始字符串
                return date_str

            # 统一解析两列日期
            parsed1 = parse_date(val1)
            parsed2 = parse_date(val2)

            # 处理空值情况
            if parsed1 == "" and parsed2 == "":
                return True

            # 根据精度需求截取
            if tail_diff == "月":
                cmp1 = parsed1[:7]  # YYYY-MM
                cmp2 = parsed2[:7]
            elif tail_diff == "年":
                cmp1 = parsed1[:4]  # YYYY
                cmp2 = parsed2[:4]
            else:  # 默认精确到日
                cmp1 = parsed1  # YYYY-MM-DD
                cmp2 = parsed2

            return cmp1 == cmp2

        elif data_type == "文本":
            # 文本型比较
            return val1 == val2

        return val1 == val2

    def convert_asset_category(self, df1, df2, mapping_df):
        """资产分类转换逻辑"""
        # 表一的资产分类字段
        asset_category_col1 = "资产分类"
        # 表二的资产分类字段
        asset_category_col2 = "SAP资产类别描述"

        # 映射表的相关字段
        source_col = "同源目录完整名称"
        target_col = "21年资产目录大类"
        detail_col = "ERP资产明细类描述"
        code_col = "同源目录编码"
        erp_detail_col = "ERP资产明细类别"

        # 创建映射字典，用于将转换后的编码映射回原始值
        self.asset_code_to_original = {}  # 转换后编码 -> 原始值

        # 转换表一的资产分类
        def convert_category(value):
            # 在映射表中查找匹配的记录
            matches = mapping_df[mapping_df[source_col] == value]
            if len(matches) == 0:
                converted_code = None  # 没有匹配项
            elif len(matches) == 1:
                # 唯一匹配项，直接返回同源目录编码前4位
                converted_code = str(matches.iloc[0][code_col])[:4]
            else:
                # 多个匹配项，需要根据表二的值来确定唯一项
                # 拼接21年资产目录大类和ERP资产明细类描述，与表二的SAP资产类别描述比较
                converted_code = None
                for _, row in matches.iterrows():
                    # 构造映射后的值
                    mapped_value = f"{row[target_col]}-{row[detail_col]}"
                    # 检查是否在表二中存在
                    if mapped_value in df2[asset_category_col2].values:
                        # 找到匹配项，返回同源目录编码前4位
                        converted_code = str(row[code_col])[:4]
                        break
                # 如果没有找到匹配项，返回第一条记录的同源目录编码前4位
                if converted_code is None:
                    converted_code = str(matches.iloc[0][code_col])[:4]

            # 保存编码到原始值的映射
            if converted_code is not None:
                self.asset_code_to_original[converted_code] = value

            return converted_code

        # 转换表二的资产分类
        def convert_category_table2(sap_value):
            # 在映射表中查找匹配的记录
            matches = mapping_df[
                mapping_df.apply(lambda row: f"{row[target_col]}-{row[detail_col]}" == sap_value, axis=1)
            ]
            if len(matches) > 0:
                # 找到匹配项，返回ERP资产明细类别前4位
                converted_code = str(matches.iloc[0][erp_detail_col])[:4]
            else:
                # 没有找到匹配项，返回原值前4位
                converted_code = str(sap_value)[:4]

            # 保存编码到原始值的映射
            self.asset_code_to_original[converted_code] = sap_value

            return converted_code

        # 应用转换函数
        df1[asset_category_col1] = df1[asset_category_col1].apply(convert_category)
        df2[asset_category_col2] = df2[asset_category_col2].apply(convert_category_table2)

        return df1, df2

    def run(self):
        try:
            self.log_signal.emit("正在并行读取Excel文件...")

            with ThreadPoolExecutor(max_workers=2) as executor:
                future1 = executor.submit(read_excel_fast, self.file1, self.sheet_name1)
                future2 = executor.submit(read_excel_fast, self.file2, self.sheet_name2)
                try:
                    df1 = future1.result()
                    df2 = future2.result()
                except Exception as e:
                    raise Exception(f"读取文件时发生错误: {str(e)}")

            self.log_signal.emit("✅ Excel文件读取完成，开始比较数据...")
            self.log_signal.emit("开始比较数据...")
            # 读取资产分类映射表
            mapping_df = read_mapping_table(self.rule_file)

            # 转换资产分类
            df1, df2 = self.convert_asset_category(df1, df2, mapping_df)

            # 检查数据行是否存在
            if df1.empty:
                self.log_signal.emit("❌ 错误：表一除了表头外没有数据行，请检查文件内容！")
                return

            if df2.empty:
                self.log_signal.emit("❌ 错误：表二除了表头外没有数据行，请检查文件内容！")
                return

            df1.columns = df1.columns.str.replace('[*\\s]', '', regex=True)
            df2.columns = df2.columns.str.replace('[*\\s]', '', regex=True)

            # 检查规则中的列是否在表一和表二表二中都存在
            table2_columns_to_check = []
            for rule in self.rules.values():
                # 如果有计算规则，则不需要检查表二是否存在该字段
                if not rule.get("calc_rule") and rule["table2_field"]:
                    table2_columns_to_check.append(rule["table2_field"])

            table1_columns_to_compare = list(self.rules.keys())  # 表一字段名
            # table2_columns_to_compare = [rule["table2_field"] for rule in self.rules.values()]  # 表二字段名
            columns_to_compare = list(self.rules.keys())

            missing_in_file1 = [col for col in table1_columns_to_compare if col not in df1.columns]
            missing_in_file2 = [col for col in table2_columns_to_check if col not in df2.columns]

            if missing_in_file1 or missing_in_file2:
                error_msg = ""
                if missing_in_file1:
                    error_msg += f"表一缺失以下规则定义的列：{', '.join(missing_in_file1)}\n"
                if missing_in_file2:
                    error_msg += f"表二缺失以下规则定义的列：{', '.join(missing_in_file2)}\n"
                self.log_signal.emit(f"❌ 比对失败：{error_msg}")
                return

            # 在"检查规则中的列是否存在"之后添加计算字段逻辑
            self.log_signal.emit("✅ 开始处理计算字段...")

            # 处理需要计算的字段
            # 存储计算字段的临时名称映射（原字段名 -> 临时名称）
            calc_temp_fields = {}

            # 处理需要计算的字段（使用临时名称避免冲突）
            for field1, rule in self.rules.items():
                if rule.get("calc_rule") and field1 in df2.columns:
                    df2.drop(columns=[field1], inplace=True)
                    self.log_signal.emit(f"删除表二中原有的 '{rule['table2_field']}' 列，将使用计算规则生成的新列")
                if field1 != rule["table2_field"] and field1 in df2.columns:
                    df2.drop(columns=[field1], inplace=True)
                if rule.get("calc_rule"):
                    self.log_signal.emit(f"正在计算表二字段: {field1} (规则: {rule['calc_rule']})")
                    try:
                        # 生成临时字段名（避免与表二原有字段冲突）
                        temp_field = f"__calc_{field1}__"
                        calc_temp_fields[field1] = temp_field

                        # 计算字段值并存储到临时字段
                        calculated_values = self.calculate_field(
                            df2,
                            rule["calc_rule"],
                            rule["data_type"]
                        )
                        df2[temp_field] = calculated_values
                    except Exception as e:
                        self.log_signal.emit(f"⚠️ 计算字段 {field1} 时出错: {str(e)}")

            # 修改映射逻辑（使用临时字段名进行映射）
            mapped_columns = {}
            for field1, rule in self.rules.items():
                field2 = rule["table2_field"]
                # 如果是计算字段，使用临时字段名映射
                if field1 in calc_temp_fields:
                    mapped_columns[calc_temp_fields[field1]] = field1
                elif field2 in df2.columns:
                    mapped_columns[field2] = field1
            # 映射完成后打印日志
            mapped_log = "\n".join([f"  {k} -> {v}" for k, v in mapped_columns.items()])
            self.log_signal.emit(f"字段映射关系：\n{mapped_log}")
            df2.rename(columns=mapped_columns, inplace=True)
            # 删除临时字段（可选）
            for temp_field in calc_temp_fields.values():
                if temp_field in df2.columns:
                    del df2[temp_field]
            # 保留需要比对的列
            all_needed_columns = list(set(columns_to_compare + self.primary_keys))
            df1 = df1[all_needed_columns]
            df2 = df2[all_needed_columns]

            # 检查主键列是否为空
            if not self.primary_keys:
                self.log_signal.emit("❌ 错误：规则文件中未定义主键字段，请检查规则文件！")
                return

                # 检查主键列在数据中是否存在
            for pk in self.primary_keys:
                if pk not in df1.columns:
                    self.log_signal.emit(f"❌ 错误：表一中不存在主键列 '{pk}'")
                    return
                if pk not in df2.columns:
                    self.log_signal.emit(f"❌ 错误：表二中不存在主键列 '{pk}'")
                    return

                # 检查主键是否有重复值
            df1_duplicates = df1[df1.duplicated(subset=self.primary_keys, keep=False)]
            if not df1_duplicates.empty:
                duplicate_count = df1_duplicates.shape[0]
                self.log_signal.emit(f"❌ 错误：表一中存在 {duplicate_count} 条重复的主键记录")
                # 显示前几个重复的主键示例
                duplicate_examples = df1_duplicates[self.primary_keys].head(5)
                example_lines = []
                for _, row in duplicate_examples.iterrows():
                    keys = [str(row[pk]) for pk in self.primary_keys]
                    example_lines.append(" + ".join(keys))
                examples = "\n".join([f" - {example}" for example in example_lines])
                self.log_signal.emit(f"重复主键示例（前5个）：\n{examples}")
                return

            df2_duplicates = df2[df2.duplicated(subset=self.primary_keys, keep=False)]
            if not df2_duplicates.empty:
                duplicate_count = df2_duplicates.shape[0]
                self.log_signal.emit(f"❌ 错误：表二中存在 {duplicate_count} 条重复的主键记录")
                # 显示前几个重复的主键示例
                duplicate_examples = df2_duplicates[self.primary_keys].head(5)
                example_lines = []
                for _, row in duplicate_examples.iterrows():
                    keys = [str(row[pk]) for pk in self.primary_keys]
                    example_lines.append(" + ".join(keys))
                examples = "\n".join([f" - {example}" for example in example_lines])
                self.log_signal.emit(f"重复主键示例（前5个）：\n{examples}")
                return

            # 检查主键列是否有空值
            for pk in self.primary_keys:
                df1_empty_keys = df1[pd.isna(df1[pk]) | (df1[pk].astype(str).astype(str).str.strip() == '')]
                df2_empty_keys = df2[pd.isna(df2[pk]) | (df2[pk].astype(str).astype(str).str.strip() == '')]

                if len(df1_empty_keys) > 0:
                    self.log_signal.emit(f"⚠️ 警告：表一中主键列 '{pk}' 存在 {len(df1_empty_keys)} 条空值记录")

                if len(df2_empty_keys) > 0:
                    self.log_signal.emit(f"⚠️ 警告：表二中主键列 '{pk}' 存在 {len(df2_empty_keys)} 条空值记录")

            # 保存原始数据帧用于导出（包含主键列）
            df1_original = df1.copy()
            df2_original = df2.copy()

            # 设置主键索引
            df1.set_index(self.primary_keys, inplace=True)
            df2.set_index(self.primary_keys, inplace=True)

            # 规范化索引格式
            df1.index = df1.index.map(lambda x: tuple(str(i) for i in x) if isinstance(x, tuple) else (str(x),))
            df2.index = df2.index.map(lambda x: tuple(str(i) for i in x) if isinstance(x, tuple) else (str(x),))
            # 检查索引中是否有空值
            df1_empty_index = df1.index.map(
                lambda x: any(pd.isna(i) or str(i).strip() == '' for i in (x if isinstance(x, tuple) else (x,))))
            df2_empty_index = df2.index.map(
                lambda x: any(pd.isna(i) or str(i).strip() == '' for i in (x if isinstance(x, tuple) else (x,))))

            df1_empty_count = sum(df1_empty_index)
            df2_empty_count = sum(df2_empty_index)

            if df1_empty_count > 0:
                self.log_signal.emit(f"⚠️ 警告：表一中有 {df1_empty_count} 条记录的主键为空")

            if df2_empty_count > 0:
                self.log_signal.emit(f"⚠️ 警告：表二中有 {df2_empty_count} 条记录的主键为空")

            if len(df1) != len(df2):
                self.log_signal.emit(f"提示：两个文件的行数不一致（表一有 {len(df1)} 行，表二有 {len(df2)} 行）")

            # 查找表二中缺失的主键
            missing_in_file2 = df1.index.difference(df2.index)
            if not missing_in_file2.empty:
                missing_df = df1.loc[missing_in_file2].copy()
                original_codes = missing_in_file2.map(lambda x: ' + '.join(map(str, x)))
                missing_df.reset_index(drop=True, inplace=True)

                for idx, key in enumerate(self.primary_keys):
                    missing_df.insert(1 + idx, key, original_codes.map(lambda x: x.split(' + ')[idx]))

                self.missing_rows = missing_df.to_dict(orient='records')
                missing_list = "\n".join([f" - {code}" for code in missing_in_file2])
                self.log_signal.emit(f"【表二中缺失的主键】（共 {len(missing_in_file2)} 条）：\n{missing_list}")

            # 查找表二中多出的主键
            missing_in_file1 = df2.index.difference(df1.index)
            if not missing_in_file1.empty:
                missing_df_file1 = df2.loc[missing_in_file1].copy()
                original_codes_file1 = missing_in_file1.map(lambda x: ' + '.join(map(str, x)))
                missing_df_file1.reset_index(drop=True, inplace=True)

                for idx, key in enumerate(self.primary_keys):
                    missing_df_file1.insert(1 + idx, key, original_codes_file1.map(lambda x: x.split(' + ')[idx]))

                self.extra_in_file2 = missing_df_file1.to_dict(orient='records')
                missing_list_file1 = "\n".join([f" - {code}" for code in missing_in_file1])
                self.log_signal.emit(
                    f"【表二中多出的主键】（表一中没有，共 {len(missing_in_file1)} 条）：\n{missing_list_file1}")

            # 找出共同的主键
            common_codes = df1.index.intersection(df2.index)
            if common_codes.empty:
                self.log_signal.emit("警告：两个文件中没有共同的主键！")
                return

            # 替换原有的数据比较部分为以下代码：
            try:
                self.log_signal.emit("开始进行向量化数据比较...")
                df1_common = df1.loc[common_codes]
                df2_common = df2.loc[common_codes]

                # 格式化索引
                df1_common.index = df1_common.index.map(lambda x: ' + '.join(x) if isinstance(x, tuple) else str(x))
                df2_common.index = df2_common.index.map(lambda x: ' + '.join(x) if isinstance(x, tuple) else str(x))
                # 使用向量化操作进行批量比较
                diff_dict = {}

                # 预处理索引以便后续使用
                index_mapping = pd.Series(df1_common.index, index=df1_common.index)

                for field1, rule in self.rules.items():
                    # 只比对规则文件中定义的列
                    if field1 not in df1_common.columns or field1 not in df2_common.columns:
                        continue

                    data_type = rule["data_type"]
                    tail_diff = rule.get("tail_diff")

                    # 向量化获取两列数据
                    series1 = df1_common[field1]
                    series2 = df2_common[field1]

                    if data_type == "数值":
                        # 数值型比较
                        series1_num = pd.to_numeric(series1, errors='coerce')
                        series2_num = pd.to_numeric(series2, errors='coerce')
                        # 如果字段名包含"折旧"，取绝对值
                        if "折旧" in field1:
                            series1_num = series1_num.abs()
                            series2_num = series2_num.abs()
                        if tail_diff is None:
                            diff_mask = (series1_num != series2_num) & \
                                        ~(pd.isna(series1_num) & pd.isna(series2_num))
                        else:
                            diff_mask = (abs(series1_num - series2_num) > float(tail_diff)) & \
                                        ~(pd.isna(series1_num) & pd.isna(series2_num))

                    elif data_type == "日期":
                        # 日期型比较
                        # 日期型比较：先统一解析为标准日期格式
                        def parse_date(date_str):
                            """尝试多种格式解析日期，返回标准化字符串（YYYY-MM-DD）"""
                            if pd.isna(date_str):
                                return ""
                            date_str = str(date_str).strip()
                            if not date_str:
                                return ""
                            # 尝试多种常见日期格式
                            formats = ['%Y-%m-%d', '%Y/%m/%d', '%Y年%m月%d日',
                                       '%m-%d-%Y', '%m/%d/%Y', '%Y%m%d']
                            for fmt in formats:
                                try:
                                    return pd.to_datetime(date_str, format=fmt).strftime('%Y-%m-%d')
                                except (ValueError, TypeError):
                                    continue
                            # 如果所有格式都解析失败，返回原始字符串
                            return date_str

                        # 统一解析两列日期
                        series1_parsed = series1.apply(parse_date)
                        series2_parsed = series2.apply(parse_date)

                        # 处理空值情况
                        both_empty = (series1_parsed == "") & (series2_parsed == "")

                        # 根据精度需求截取
                        if tail_diff == "月":
                            series1_cmp = series1_parsed.str[:7]  # YYYY-MM
                            series2_cmp = series2_parsed.str[:7]
                        elif tail_diff == "年":
                            series1_cmp = series1_parsed.str[:4]  # YYYY
                            series2_cmp = series2_parsed.str[:4]
                        else:  # 默认精确到日
                            series1_cmp = series1_parsed  # YYYY-MM-DD
                            series2_cmp = series2_parsed

                        diff_mask = (series1_cmp != series2_cmp) & ~both_empty

                    elif data_type == "文本":
                        # 文本型比较
                        series1_norm = series1.fillna('').astype(str).str.strip()
                        series2_norm = series2.fillna('').astype(str).str.strip()
                        diff_mask = series1_norm != series2_norm

                    # 找出有差异的行索引
                    diff_indices = df1_common[diff_mask].index

                    # 批量添加差异记录
                    for idx in diff_indices:
                        if idx not in diff_dict:
                            diff_dict[idx] = []

                        val1 = CompareWorker.normalize_value(series1.loc[idx])
                        val2 = CompareWorker.normalize_value(series2.loc[idx])
                        diff_dict[idx].append((field1, val1, val2))

                self.log_signal.emit(f"向量化比较完成，共发现 {len(diff_dict)} 条差异记录")

            except Exception as e:
                self.log_signal.emit(f"向量化比较出错，使用传统方法: {str(e)}")
                # 如果向量化方法出错，回退到原来的逐行比较方法

            # 生成日志和汇总信息
            # 生成日志和汇总信息
            diff_log_messages = []
            self.diff_full_rows = []

            # 创建主键到原始值的映射，用于恢复导出数据中的主键列
            pk_mapping = {}
            for code in common_codes:
                code_str = ' + '.join(code) if isinstance(code, tuple) else str(code)
                pk_mapping[code_str] = code

            for code, diffs in diff_dict.items():
                code_str = ' + '.join(code) if isinstance(code, tuple) else str(code)
                diff_details = []

                for col, val1, val2 in diffs:
                    if col == "资产分类" and hasattr(self, 'asset_code_to_original'):
                        # 使用原始中文值显示
                        original_val1 = self.asset_code_to_original.get(val1, val1)
                        original_val2 = self.asset_code_to_original.get(val2, val2)
                        diff_details.append(f" - 列 [{col}] 不一致：表一={original_val1}, 表二={original_val2}")
                    else:
                        diff_details.append(f" - 列 [{col}] 不一致：表一={val1}, 表二={val2}")

                diff_log_messages.append(f"\n主键：{code}")
                diff_log_messages.extend(diff_details)

                # 使用原始数据帧查找完整行数据（包含主键列）
                try:
                    # 获取原始主键值
                    original_pk_values = pk_mapping.get(code_str, code)

                    # 构建筛选条件
                    if isinstance(original_pk_values, tuple):
                        # 多主键情况
                        condition1 = True
                        condition2 = True
                        for i, pk in enumerate(self.primary_keys):
                            condition1 = condition1 & (df1_original[pk].astype(str) == original_pk_values[i])
                            condition2 = condition2 & (df2_original[pk].astype(str) == original_pk_values[i])
                    else:
                        # 单主键情况
                        pk = self.primary_keys[0]
                        condition1 = (df1_original[pk].astype(str) == original_pk_values)
                        condition2 = (df2_original[pk].astype(str) == original_pk_values)

                    # 获取完整行数据
                    source_dict = df1_original[condition1].iloc[0].to_dict()
                    target_dict = df2_original[condition2].iloc[0].to_dict()

                    self.diff_full_rows.append({
                        "source": source_dict,
                        "target": target_dict
                    })
                except (IndexError, KeyError, Exception) as e:
                    # 出现异常时使用原来的方法作为备选
                    source_dict = df1_common.loc[code_str].to_dict()
                    target_dict = df2_common.loc[code_str].to_dict()

                    # 手动添加主键信息
                    original_pk_values = pk_mapping.get(code_str, code)
                    if isinstance(original_pk_values, tuple):
                        for i, pk in enumerate(self.primary_keys):
                            source_dict[pk] = original_pk_values[i]
                            target_dict[pk] = original_pk_values[i]
                    else:
                        if self.primary_keys:
                            source_dict[self.primary_keys[0]] = original_pk_values
                            target_dict[self.primary_keys[0]] = original_pk_values

                    self.diff_full_rows.append({
                        "source": source_dict,
                        "target": target_dict
                    })

                except (IndexError, KeyError, Exception) as e:
                    # 出现异常时使用原来的方法作为备选
                    source_dict = df1_common.loc[code_str].to_dict()
                    target_dict = df2_common.loc[code_str].to_dict()

                    # 手动添加主键信息
                    original_pk_values = pk_mapping.get(code_str, code)
                    if isinstance(original_pk_values, tuple):
                        for i, pk in enumerate(self.primary_keys):
                            source_dict[pk] = original_pk_values[i]
                            target_dict[pk] = original_pk_values[i]
                    else:
                        if self.primary_keys:
                            source_dict[self.primary_keys[0]] = original_pk_values
                            target_dict[self.primary_keys[0]] = original_pk_values

                    self.diff_full_rows.append({
                        "source": source_dict,
                        "target": target_dict
                    })

            diff_count = len(diff_dict)
            equal_count = len(common_codes) - diff_count
            primary_key_str = " + ".join(self.primary_keys)

            self.summary = {
                "primary_key": primary_key_str,
                "total_file1": len(df1),
                "total_file2": len(df2),
                "missing_count": len(missing_in_file2),
                "extra_count": len(missing_in_file1),
                "common_count": len(common_codes),
                "diff_count": diff_count,
                "equal_count": equal_count,
                "diff_ratio": diff_count / len(common_codes) if len(common_codes) > 0 else 0.0,
            }

            if diff_count == 0:
                self.log_signal.emit("【共同主键的数据完全一致】，没有差异。")
            else:
                self.log_signal.emit(f"【存在差异的列】（共 {diff_count} 行）：")
                if diff_log_messages:
                    self.log_signal.emit('\n'.join(diff_log_messages))
                else:
                    self.log_signal.emit("⚠️ 未找到具体差异列，请检查数据是否一致。")

        except Exception as e:
            logging.error(traceback.format_exc())
            self.log_signal.emit(f"发生错误：{str(e)}")
        finally:
            self.quit()
            self.wait()


class ExcelComparer(QWidget):
    """主窗口类"""

    def __init__(self):
        super().__init__()
        self.file1 = ""
        self.file2 = ""
        self.sheet_name1 = ""
        self.sheet_name2 = ""
        self.initUI()
        self.worker = None
        self.summary_data = {}
        self.columns1 = []
        self.columns2 = []
        self.rules = {}  # 存储解析后的规则
        self.rule_file = ""
        # 初始化 worker 变量
        self.worker_sheet1 = None
        self.worker_sheet2 = None
        self.worker_load1 = None
        self.worker_load2 = None
        self.loading_dialog = None
        # 读取规则文件
        self.load_rules_file()

    def load_rules_file(self):
        """加载规则文件"""
        try:
            # 获取exe文件所在目录
            if hasattr(sys, '_MEIPASS'):
                # 打包后的exe环境
                exe_dir = os.path.dirname(sys.executable)
            else:
                # 开发环境
                exe_dir = os.path.dirname(os.path.abspath(__file__))

            rule_file_path = os.path.join(exe_dir, "rule.xlsx")
            self.rule_file = rule_file_path
            if os.path.exists(rule_file_path):
                self.rules = read_rules(rule_file_path)
                self.log(f"✅ 成功加载规则文件: {rule_file_path}")
            else:
                self.log(f"❌ 未找到规则文件: {rule_file_path}")
                # 可以选择是否继续运行或者退出
        except Exception as e:
            self.log(f"❌ 读取规则文件失败: {str(e)}")
    def initUI(self):
        """初始化用户界面"""
        self.setWindowTitle("Excel文件比较工具V2.4")
        self.resize(1000, 700)

        main_layout = QVBoxLayout()

        # 文件选择区域
        file_layout = QHBoxLayout()

        left_layout = QVBoxLayout()
        self.label1 = QLabel("未选择表一")
        self.btn1 = QPushButton("选择表一")
        self.btn1.clicked.connect(self.select_file1)


        self.sheet_label1 = QLabel("选择表一页签：")
        self.sheet_combo1 = QComboBox()
        self.sheet_combo1.currentTextChanged.connect(self.on_sheet_selection_changed)

        left_layout.addWidget(self.label1)
        left_layout.addWidget(self.btn1)
        left_layout.addWidget(self.sheet_label1)
        left_layout.addWidget(self.sheet_combo1)

        right_layout = QVBoxLayout()
        self.label2 = QLabel("未选择表二")
        self.btn2 = QPushButton("选择表二")
        self.btn2.clicked.connect(self.select_file2)

        self.sheet_label2 = QLabel("选择表二页签：")
        self.sheet_combo2 = QComboBox()
        self.sheet_combo2.currentTextChanged.connect(self.on_sheet_selection_changed)

        right_layout.addWidget(self.label2)
        right_layout.addWidget(self.btn2)
        right_layout.addWidget(self.sheet_label2)
        right_layout.addWidget(self.sheet_combo2)
        file_layout.addLayout(left_layout)
        file_layout.addLayout(right_layout)
        # 按钮区域
        button_layout = QHBoxLayout()
        self.compare_btn = QPushButton("比较文件")
        self.compare_btn.setFixedWidth(150)
        self.compare_btn.clicked.connect(self.compare_files)
        self.compare_btn.setEnabled(False)
        self.export_btn = QPushButton("导出报告")
        self.export_btn.setFixedWidth(150)
        self.export_btn.setEnabled(False)
        self.export_btn.clicked.connect(self.export_report)
        button_layout.addStretch()
        button_layout.addWidget(self.compare_btn)
        button_layout.addWidget(self.export_btn)
        # 日志和报告区域
        self.tab_widget = QTabWidget()
        self.log_area = QPlainTextEdit()
        self.log_area.setReadOnly(True)
        self.log_area.setStyleSheet("background-color: #f0f0f0;")
        self.summary_area = QPlainTextEdit()
        self.summary_area.setReadOnly(True)
        self.summary_area.setStyleSheet("background-color: #f0f0f0;")
        self.tab_widget.addTab(self.log_area, "比对日志")
        self.tab_widget.addTab(self.summary_area, "汇总报告")
        # 主布局组合
        main_layout.addLayout(file_layout)
        main_layout.addLayout(button_layout)
        main_layout.addWidget(self.tab_widget)

        self.setLayout(main_layout)

    def closeEvent(self, event):
        """窗口关闭时确保线程安全退出"""
        if hasattr(self, 'worker') and self.worker is not None and self.worker.isRunning():
            self.worker.quit()
            self.worker.wait()
        if hasattr(self, 'worker_load1') and self.worker_load1 is not None and self.worker_load1.isRunning():
            self.worker_load1.quit()
            self.worker_load1.wait()
        if hasattr(self, 'worker_load2') and self.worker_load2 is not None and self.worker_load2.isRunning():
            self.worker_load2.quit()
            self.worker_load2.wait()
        if hasattr(self, 'worker_sheet1') and self.worker_sheet1 is not None and self.worker_sheet1.isRunning():
            self.worker_sheet1.quit()
            self.worker_sheet1.wait()
        if hasattr(self, 'worker_sheet2') and self.worker_sheet2 is not None and self.worker_sheet2.isRunning():
            self.worker_sheet2.quit()
            self.worker_sheet2.wait()
        super().closeEvent(event)

    def reset_file_state(self, is_file1=True, is_file2=False):
        if is_file1:
            self.columns1 = []
            self.sheet_combo1.clear()
            self.sheet_combo1.setEnabled(True)
            self.sheet_label1.setText("选择表一页签：")
            if hasattr(self, 'worker_sheet1'):
                self.worker_sheet1 = None
        if is_file2:
            self.columns2 = []
            self.sheet_combo2.clear()
            self.sheet_combo2.setEnabled(True)
            self.sheet_label2.setText("选择表二页签：")
            if hasattr(self, 'worker_sheet2'):
                self.worker_sheet2 = None
        self.compare_btn.setEnabled(False)
        self.log_area.clear()
        self.summary_area.clear()
        self.export_btn.setEnabled(False)

    def select_file1(self):
        self.reset_file_state(is_file1=True, is_file2=False)
        file, _ = QFileDialog.getOpenFileName(self, "选择 Excel 文件", "", "Excel 文件 (*.xlsx *.xls)")
        if file:
            self.file1 = file
            filename = os.path.basename(file)
            self.label1.setText(f"表一: {filename}")
            # 显示加载对话框
            self.show_loading_dialog("正在加载表一页签...")
            self.load_sheet_and_columns(file, is_file1=True)

    def select_file2(self):
        self.reset_file_state(is_file1=False, is_file2=True)
        file, _ = QFileDialog.getOpenFileName(self, "选择 Excel 文件", "", "Excel 文件 (*.xlsx *.xls)")
        if file:
            self.file2 = file
            filename = os.path.basename(file)

            self.label2.setText(f"表二: {filename}")
            self.show_loading_dialog("正在加载表二页签...")
            self.load_sheet_and_columns(file, is_file2=True)

    def show_loading_dialog(self, message="正在加载，请稍候..."):
        """显示加载对话框"""
        if not self.loading_dialog:
            self.loading_dialog = QProgressDialog(message, None, 0, 0, self)
            self.loading_dialog.setWindowModality(Qt.WindowModal)
            self.loading_dialog.setWindowTitle("加载中")
            self.loading_dialog.setCancelButton(None)
            self.loading_dialog.show()
    def load_sheet_and_columns(self, file_path, is_file1=False, is_file2=False):

        worker = LoadColumnWorker(file_path)
        worker.sheet_names_loaded.connect(self.on_sheet_names_loaded)
        worker.sheet_names_loaded.connect(self.close_loading_dialog)
        # worker.columns_loaded.connect(self.on_columns_loaded)
        # worker.error_occurred.connect(self.on_column_error)
        if is_file1:
            self.worker_load1 = worker
        elif is_file2:
            self.worker_load2 = worker
        worker.start()

    def on_sheet_names_loaded(self, file_path, sheet_names):
        if file_path == self.file1:
            self.sheet_combo1.clear()
            self.sheet_combo1.addItems(sheet_names)
            self.sheet_combo1.setCurrentIndex(0)
        elif file_path == self.file2:
            self.sheet_combo2.clear()
            self.sheet_combo2.addItems(sheet_names)
            self.sheet_combo2.setCurrentIndex(0)

    def on_sheet_selection_changed(self):
        """页签选择变化时的处理函数"""
        # 简单更新比较按钮状态
        self.update_compare_button_state()


    def update_compare_button_state(self):
        sheet_selected = self.sheet_combo1.currentText() and self.sheet_combo2.currentText()
        if not sheet_selected:
            self.compare_btn.setEnabled(False)
            return

        self.compare_btn.setEnabled(True)

    def compare_files(self):
        if not self.file1 or not self.file2:
            self.log("请先选择两个 Excel 文件！")
            return
        sheet_name1 = self.sheet_combo1.currentText()
        sheet_name2 = self.sheet_combo2.currentText()
        if not sheet_name1 or not sheet_name2:
            self.log("请选择两个文件的页签！")
            return

        self.log_area.clear()
        self.summary_area.clear()
        self.export_btn.setEnabled(False)

        # 获取主键字段
        primary_keys = [field for field, rule in self.rules.items() if rule["is_primary"]]
        if not primary_keys:
            self.log("规则文件中未定义主键字段，请检查规则文件！")
            return
        self.loading_dialog = QProgressDialog("正在比较文件，请稍候...", None, 0, 0, self)
        self.loading_dialog.setWindowModality(Qt.WindowModal)
        self.loading_dialog.setWindowTitle("比较中")
        self.loading_dialog.setCancelButton(None)
        self.loading_dialog.show()

        self.worker = CompareWorker(self.file1, self.file2, self.rule_file, sheet_name1, sheet_name2, primary_keys=primary_keys,
                                    rules=self.rules)
        self.worker.log_signal.connect(self.log)
        # 连接信号以在比较完成时关闭对话框
        self.worker.finished.connect(self.close_loading_dialog)
        self.worker.finished.connect(lambda: self.export_btn.setEnabled(True))
        self.worker.finished.connect(self.on_compare_finished)
        self.worker.start()

    def close_loading_dialog(self):
        """关闭加载对话框"""
        if self.loading_dialog:
            self.loading_dialog.close()
            self.loading_dialog = None

    def on_compare_finished(self):
        try:
            if hasattr(self.worker, 'summary'):
                self.summary_data = self.worker.summary
                primary_key = self.summary_data.get("primary_key", "主键")
                total_file1 = self.summary_data['total_file1']
                total_file2 = self.summary_data['total_file2']
                missing_count = self.summary_data['missing_count']
                extra_count = self.summary_data.get('extra_count', 0)
                common_count = self.summary_data['common_count']
                diff_count = self.summary_data['diff_count']
                equal_count = self.summary_data['equal_count']
                diff_ratio = self.summary_data['diff_ratio']
                missing_columns = self.summary_data.get("missing_columns", [])
                missing_columns_str = ", ".join(missing_columns) if missing_columns else "无"

                summary_text = (
                    f"📊 比对汇总报告\n"
                    f"--------------------------------\n"
                    f"• 总{primary_key}数量（表一）：{total_file1}\n"
                    f"• 总{primary_key}数量（表二）：{total_file2}\n"
                    f"• 表二中缺失的{primary_key}：{missing_count}\n"
                    f"• 表二中多出的{primary_key}：{extra_count}\n"
                    f"• 共同{primary_key}数量：{common_count}\n"
                    f"• 列不一致的{primary_key}数量：{diff_count}\n"
                    f"• 列一致的{primary_key}数量：{equal_count}\n"
                    f"• 表二中缺失的列：{missing_columns_str}\n"
                    f"--------------------------------\n"
                    f"• 差异数据占比：{diff_ratio:.2%}\n"
                )
                self.summary_area.setPlainText(summary_text)
                self.export_btn.setEnabled(True)
        except Exception as e:
            self.summary_area.setPlainText(f"❌ 显示汇总报告时发生错误：{str(e)}\n请查看比对日志了解详细信息。")
            self.export_btn.setEnabled(False)

    def export_report(self):
        """复制原始文件并修改副本，添加对比结果和差异详情"""
        if not hasattr(self, 'worker') or not hasattr(self.worker, 'missing_rows') or not hasattr(self.worker,
                                                                                                  'diff_full_rows'):
            self.log("没有可导出的数据，请先执行比对！")
            return

        try:
            # 获取保存路径
            directory = QFileDialog.getExistingDirectory(self, "选择保存路径")
            if not directory:
                self.log("导出已取消。")
                return

            # 复制并修改表一文件
            file1_name = os.path.splitext(os.path.basename(self.file1))[0]
            file1_copy_path = f"{directory}/{file1_name}_比对结果.xlsx"
            import shutil
            shutil.copy2(self.file1, file1_copy_path)
            self._modify_original_file(file1_copy_path, self.sheet_combo1.currentText(), is_first_file=True)

            # 复制并修改表二文件
            file2_name = os.path.splitext(os.path.basename(self.file2))[0]
            file2_copy_path = f"{directory}/{file2_name}_比对结果.xlsx"
            shutil.copy2(self.file2, file2_copy_path)
            self._modify_original_file(file2_copy_path, self.sheet_combo2.currentText(), is_first_file=False)

            self.log(f"✅ 已生成比对结果文件：{file1_copy_path} 和 {file2_copy_path}")
        except Exception as e:
            self.log(f"❌ 生成比对结果文件时发生错误：{str(e)}")

    def _modify_original_file(self, file_path, sheet_name, is_first_file):
        """直接修改原始Excel文件 - 性能优化版本"""
        try:
            # 加载工作簿
            wb = load_workbook(file_path)
            ws = wb[sheet_name]

            # 获取主键
            primary_keys = [field for field, rule in self.rules.items() if rule["is_primary"]]

            # 获取需要比对的列
            compare_columns = list(self.rules.keys())

            # 创建红色填充样式
            red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

            # 预处理数据 - 构建主键到差异数据的映射字典
            diff_dict = {}
            missing_in_file2_keys = set()  # 表一有但表二没有的主键
            missing_in_file1_keys = set()  # 表二有但表一没有的主键

            if hasattr(self.worker, 'diff_full_rows'):
                for item in self.worker.diff_full_rows:
                    # 构建主键 - 需要与对比部分使用相同的逻辑
                    if is_first_file:
                        # 表一文件使用source数据构建主键
                        key_parts = [str(item['source'].get(pk, '')) for pk in primary_keys]
                    else:
                        # 表二文件使用target数据构建主键（与对比逻辑一致）
                        key_parts = [str(item['target'].get(pk, '')) for pk in primary_keys]

                    # 处理多主键拼接（与对比部分一致）
                    key = ' + '.join(key_parts) if len(key_parts) > 1 else (key_parts[0] if key_parts else "")
                    diff_dict[key] = item

            # 处理缺失数据的主键
            if hasattr(self.worker, 'missing_rows'):
                for row in self.worker.missing_rows:
                    # 表一中存在但表二中缺失的数据，使用表一的主键
                    key_parts = [str(row.get(pk, '')) for pk in primary_keys]
                    key = ' + '.join(key_parts) if len(key_parts) > 1 else (key_parts[0] if key_parts else "")
                    missing_in_file2_keys.add(key)

            # 处理多余数据的主键
            if hasattr(self.worker, 'extra_in_file2'):
                for row in self.worker.extra_in_file2:
                    # 表二中存在但表一中缺失的数据，使用表二的主键
                    key_parts = [str(row.get(pk, '')) for pk in primary_keys]
                    key = ' + '.join(key_parts) if len(key_parts) > 1 else (key_parts[0] if key_parts else "")
                    missing_in_file1_keys.add(key)

            # 创建列名到列索引的映射（一次性处理）
            col_name_to_index = {}
            for col_idx in range(1, ws.max_column + 1):
                col_name = ws.cell(row=1, column=col_idx).value
                if col_name:
                    # 清理列名（去除*和空格）
                    cleaned_col_name = str(col_name).replace('*', '').strip()
                    col_name_to_index[cleaned_col_name] = col_idx

            # 在第一行添加新列标题
            max_col = ws.max_column
            ws.cell(row=1, column=max_col + 1, value="对比结果")
            for i, col in enumerate(compare_columns):
                ws.cell(row=1, column=max_col + 2 + i, value=f"{col}")

            # 创建一个辅助函数来计算主键值（与对比逻辑保持一致）
            def calculate_composite_key(row_data, is_table2=False):
                """根据规则计算复合主键值"""
                key_parts = []

                for pk in primary_keys:
                    # 获取主键对应的规则
                    pk_rule = self.rules.get(pk)

                    # 如果是表二且主键有计算规则，则按规则计算
                    if is_table2 and pk_rule and pk_rule.get("calc_rule"):
                        # 对于表二，如果主键需要计算，则使用计算规则
                        calc_rule = pk_rule["calc_rule"]
                        data_type = pk_rule["data_type"]

                        try:
                            # 模拟计算过程（简化版）
                            # 实际应该使用与CompareWorker中相同的calculate_field方法
                            if '+' in calc_rule and data_type == "文本":
                                # 字符串拼接情况，如"公司代码+资产编码"
                                fields = [f.strip() for f in calc_rule.split('+')]
                                concatenated_value = ""
                                for field in fields:
                                    field_col_idx = col_name_to_index.get(field)
                                    if field_col_idx and row_data.get(field_col_idx):
                                        concatenated_value += str(row_data[field_col_idx])
                                key_parts.append(concatenated_value)
                            else:
                                # 其他情况使用直接获取的值
                                pk_col_idx = col_name_to_index.get(pk)
                                if pk_col_idx and row_data.get(pk_col_idx):
                                    key_parts.append(str(row_data[pk_col_idx]))
                                else:
                                    key_parts.append("")
                        except:
                            # 出错时使用直接获取的值
                            pk_col_idx = col_name_to_index.get(pk)
                            if pk_col_idx and row_data.get(pk_col_idx):
                                key_parts.append(str(row_data[pk_col_idx]))
                            else:
                                key_parts.append("")
                    else:
                        # 表一或其他不需要计算的情况，直接使用值
                        pk_col_idx = col_name_to_index.get(pk)
                        if pk_col_idx and row_data.get(pk_col_idx):
                            key_parts.append(str(row_data[pk_col_idx]))
                        else:
                            key_parts.append("")

                return ' + '.join(key_parts) if len(key_parts) > 1 else (key_parts[0] if key_parts else "")

            # 批量处理所有数据行，减少重复计算
            row_updates = []  # 收集所有需要更新的行信息
            fill_operations = []  # 收集所有需要标红的操作

            # 先收集所有行的信息
            for row_idx in range(2, ws.max_row + 1):
                # 读取当前行的所有数据
                row_data = {}
                for col_idx in range(1, ws.max_column + 1):
                    row_data[col_idx] = ws.cell(row=row_idx, column=col_idx).value

                # 构建当前行的主键（与对比部分保持一致）
                if not is_first_file:  # 表二文件
                    # 使用表二的主键计算逻辑
                    key = calculate_composite_key(row_data, is_table2=True)
                else:  # 表一文件
                    # 表一使用直接获取的主键值
                    key_parts = []
                    for pk in primary_keys:
                        pk_col_idx = col_name_to_index.get(pk)
                        if pk_col_idx and row_data.get(pk_col_idx):
                            key_parts.append(str(row_data[pk_col_idx]))
                        else:
                            key_parts.append("")
                    key = ' + '.join(key_parts) if len(key_parts) > 1 else (key_parts[0] if key_parts else "")

                # 确定对比结果
                comparison_result = ""
                if key in missing_in_file2_keys:
                    comparison_result = "此数据不存在于SAP" if is_first_file else "此数据不存在于平台"
                elif key in missing_in_file1_keys:
                    comparison_result = "此数据不存在于平台" if is_first_file else "此数据不存在于SAP"
                elif key in diff_dict:
                    comparison_result = "不一致"
                else:
                    comparison_result = "一致"

                # 收集该行需要的更新信息
                row_updates.append({
                    'row_idx': row_idx,
                    'key': key,
                    'comparison_result': comparison_result,
                    'row_data': row_data
                })

            # 批量执行更新操作，减少与Excel文件的交互次数
            for update_info in row_updates:
                row_idx = update_info['row_idx']
                key = update_info['key']
                comparison_result = update_info['comparison_result']
                row_data = update_info['row_data']

                # 填入对比结果
                ws.cell(row=row_idx, column=max_col + 1, value=comparison_result)

                # 填入各列的差异详情
                if key in diff_dict:
                    diff_data = diff_dict[key]
                    # 根据是表一还是表二来获取正确的数据源
                    if is_first_file:
                        source_data = diff_data.get('source', {})
                        target_data = diff_data.get('target', {})
                    else:
                        source_data = diff_data.get('source', {})
                        target_data = diff_data.get('target', {})

                    for i, col in enumerate(compare_columns):
                        if col in source_data and col in target_data:
                            val1 = source_data[col]
                            val2 = target_data[col]

                            # 获取该列的规则
                            rule = self.rules.get(col, {})
                            data_type = rule.get("data_type", "文本")  # 默认为文本类型
                            tail_diff = rule.get("tail_diff")

                            # 使用规则判断值是否相等
                            are_equal = self.worker.values_equal_by_rule(val1, val2, data_type, tail_diff, col)
                            if not are_equal:
                                # 如果是资产分类且有映射，使用原始值
                                if col == "资产分类" and hasattr(self.worker, 'asset_code_to_original'):
                                    original_val1 = self.worker.asset_code_to_original.get(val1, val1)
                                    original_val2 = self.worker.asset_code_to_original.get(val2, val2)
                                    diff_detail = f"不一致：表一={original_val1}, 表二={original_val2}"
                                else:
                                    diff_detail = f"不一致：表一={val1}, 表二={val2}"

                                ws.cell(row=row_idx, column=max_col + 2 + i, value=diff_detail)

                                # 记录需要标红的单元格
                                if comparison_result == "不一致":
                                    fill_operations.append((row_idx, max_col + 2 + i))

                    # 记录需要标红的对比结果单元格
                    if comparison_result in ["不一致", "此数据不存在于SAP", "此数据不存在于平台"]:
                        fill_operations.append((row_idx, max_col + 1))

            # 批量执行所有标红操作
            for row_idx, col_idx in fill_operations:
                ws.cell(row=row_idx, column=col_idx).fill = red_fill

            # 保存修改后的文件
            wb.save(file_path)
            wb.close()

        except Exception as e:
            self.log(f"修改文件 {file_path} 时出错: {str(e)}")
            raise e

    def log(self, message):
        """日志输出"""
        self.log_area.appendPlainText(message)

def exception_hook(exc_type, exc_value, exc_traceback):
    """全局异常钩子，防止崩溃"""
    try:
        ex = QApplication.instance().topLevelWidgets()[0]
        if hasattr(ex, "log"):
            error_message = "".join(traceback.format_exception(exc_type, exc_value, exc_traceback))
            logging.error(error_message)
            ex.log(f"❌ 发生异常：{exc_value}")
        else:
            sys.__excepthook__(exc_type, exc_value, exc_traceback)
    except:
        sys.__excepthook__(exc_type, exc_value, exc_traceback)


if __name__ == "__main__":
    sys.excepthook = exception_hook
    app = QApplication(sys.argv)
    icon_path = resource_path('icon.ico')
    app.setWindowIcon(QIcon(icon_path))
    ex = ExcelComparer()
    ex.show()
    exit_code = app.exec_()
    sys.exit(exit_code)