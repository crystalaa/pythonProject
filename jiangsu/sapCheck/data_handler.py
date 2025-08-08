# data_handler.py
import os
import re

import pandas as pd
from concurrent.futures import ThreadPoolExecutor
from PyQt5.QtCore import QThread, pyqtSignal
from openpyxl import load_workbook
import xlrd

class LoadColumnWorker(QThread):
    """用于在独立线程中读取列名"""
    columns_loaded = pyqtSignal(str, list)  # 参数为文件路径和列名列表
    error_occurred = pyqtSignal(str)
    sheet_names_loaded = pyqtSignal(str, list)

    def __init__(self, file_path, sheet_name=None):
        super().__init__()
        self.file_path = file_path
        self.sheet_name = sheet_name

    def run(self):
        try:
            # 读取页签名称
            sheet_names = get_sheet_names(self.file_path)
            self.sheet_names_loaded.emit(self.file_path, sheet_names)
            if not self.sheet_name:
                return
            # 读取列名
            columns = read_excel_columns(self.file_path, self.sheet_name)
            if columns is None:
                return
            self.columns_loaded.emit(self.file_path, columns)
        except Exception as e:
            self.error_occurred.emit(str(e))

def read_excel_columns(file_path, sheet_name):
    """快速读取Excel文件的列名"""
    try:
        if not sheet_name:  # 空字符串、None 都视为未选择
            return

        if file_path.lower().endswith('.xls'):
            # 处理 .xls 文件
            import xlrd
            try:
                workbook = xlrd.open_workbook(file_path)
                worksheet = workbook.sheet_by_name(sheet_name)
                if worksheet.nrows > 0:
                    columns = [cell.value for cell in worksheet.row(0)]
                    cleaned_columns = [col.replace('*', '').strip() if isinstance(col, str) else col for col in columns]
                    return cleaned_columns
            except xlrd.biffh.XLRDError as e:
                # 如果是版本问题，尝试用 openpyxl 处理（可能是实际为 .xlsx 格式的文件）
                if "xlsx file" in str(e).lower():
                    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
                    ws = wb[sheet_name]
                    columns = [cell.value for cell in next(ws.iter_rows())]
                    cleaned_columns = [col.replace('*', '').strip() if isinstance(col, str) else col for col in columns]
                    wb.close()
                    return cleaned_columns
                else:
                    raise e
        else:
            # 处理 .xlsx 文件
            wb = load_workbook(filename=file_path, read_only=True, data_only=True)
            ws = wb[sheet_name]
            columns = [cell.value for cell in next(ws.iter_rows())]
            cleaned_columns = [col.replace('*', '').strip() if isinstance(col, str) else col for col in columns]
            wb.close()
            return cleaned_columns
    except Exception as e:
        raise Exception(f"读取Excel列名时发生错误: {str(e)}")



def read_excel_with_header(file_path, sheet_name, skip_rows=0, is_file1=True):
    """
    读取带一级/二级表头（合并单元格已自动填充）的 Excel
    is_file1=True  -> 平台文件：一级-二级 拼接
    is_file1=False -> ERP文件：跳过 skip_rows 行后取表头
    返回列名与规则完全对齐的 DataFrame
    """
    if file_path.lower().endswith('.xlsx'):
        # ---------- 1. 普通模式读表头 ----------
        wb = load_workbook(file_path, data_only=True, read_only=False)
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if is_file1:
            if ws.merged_cells.ranges:
                # 平台文件：一级 + 二级
                level1 = [str(v or '') for v in rows[0]]
                level2 = [str(v or '') for v in rows[1]]
                # 把一级合并单元格向右填充
                for merged in ws.merged_cells.ranges:
                    if merged.bounds[1] == 1:  # 第1行
                        min_col, max_col = merged.bounds[0], merged.bounds[2]
                        fill_val = level1[min_col - 1]
                        for c in range(min_col, max_col + 1):
                            level1[c - 1] = fill_val
                cols = [f"{a}-{b}".strip('-') for a, b in zip(level1, level2)]
                data_start = 2
            else:
                header_row = 1
                cols = [str(v or '') for v in rows[header_row - 1]]
                data_start = header_row
        else:
            if ws.merged_cells.ranges:
                # ERP 文件：跳过 skip_rows
                header_row = 1 if is_file1 else (skip_rows + 1)
                cols = [str(v or '') for v in rows[header_row - 1]]
                data_start = header_row
            else:
                header_row = 1
                cols = [str(v or '') for v in rows[header_row - 1]]
                data_start = header_row
        cols = [re.sub(r'[\*\s]+', '', c) for c in cols]
        data_rows = list(ws.iter_rows(min_row=data_start + 1, values_only=True))

        df = pd.DataFrame(data_rows, columns=cols)
        wb.close()

        # ---------- 2. read_only 读数据 ----------
        # wb_data = load_workbook(file_path, data_only=True, read_only=True)
        # ws_data = wb_data[sheet_name]
        # data_rows = list(ws_data.iter_rows(min_row=data_start + 1, values_only=True))
        # df = pd.DataFrame(data_rows, columns=cols)
        # wb_data.close()
        return df

    else:
        # ---------- .xls 处理 ----------
        bk = xlrd.open_workbook(file_path)
        sh = bk.sheet_by_name(sheet_name)
        if is_file1 and sh.nrows >= 2:
            level1 = [str(sh.cell_value(0, c)) for c in range(sh.ncols)]
            level2 = [str(sh.cell_value(1, c)) for c in range(sh.ncols)]
            # xlrd merged_cells 的格式：(row_low, row_high, col_low, col_high)
            for crange in sh.merged_cells:
                if crange[0] == 0:  # 第1行
                    min_col, max_col = crange[2], crange[3]
                    fill_val = level1[min_col]
                    for c in range(min_col, max_col):
                        level1[c] = fill_val
            cols = [f"{a}-{b}".strip('-') for a, b in zip(level1, level2)]
            data_start = 2
        else:
            cols = [str(sh.cell_value(skip_rows, c)) for c in range(sh.ncols)]
            data_start = skip_rows + 1
    cols = [re.sub(r'[\*\s]+', '', c) for c in cols]
    data = [sh.row_values(r) for r in range(data_start, sh.nrows)]
    df = pd.DataFrame(data, columns=cols)
    return df

def _read_xls_with_header(file_path, sheet_name, skip_rows, is_file1):
    import xlrd
    bk = xlrd.open_workbook(file_path)
    sh = bk.sheet_by_name(sheet_name)
    if is_file1 and sh.nrows >= 2:
        level1 = [str(sh.cell_value(0, c)) for c in range(sh.ncols)]
        level2 = [str(sh.cell_value(1, c)) for c in range(sh.ncols)]
        cols = [f"{a}-{b}".strip('-') for a, b in zip(level1, level2)]
        data_start = 2
    else:
        cols = [str(sh.cell_value(skip_rows, c)) for c in range(sh.ncols)]
        data_start = skip_rows + 1
    data = [sh.row_values(r) for r in range(data_start, sh.nrows)]
    df = pd.DataFrame(data, columns=cols)
    df.columns = [re.sub(r'[\*\s]+', '', c) for c in df.columns]
    return df


def read_excel_fast(file_path, sheet_name,is_file1=True):
    return read_excel_with_header(file_path, sheet_name, skip_rows=1, is_file1=is_file1)
    # """快速读取Excel文件"""
    # try:
    #     if file_path.lower().endswith('.xls'):
    #         # 处理 .xls 文件
    #
    #         workbook = xlrd.open_workbook(file_path)
    #         worksheet = workbook.sheet_by_name(sheet_name)
    #
    #         data = []
    #         columns = None
    #         for i in range(worksheet.nrows):
    #             row = worksheet.row(i)
    #             if i == 0:
    #                 columns = [cell.value for cell in row]
    #             else:
    #                 data.append([cell.value for cell in row])
    #
    #         return pd.DataFrame(data, columns=columns)
    #     else:
    #         wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    #         ws = wb[sheet_name]
    #         data = []
    #         columns = None
    #         for i, row in enumerate(ws.rows):
    #             if i == 0:
    #                 columns = [cell.value for cell in row]
    #             else:
    #                 data.append([cell.value for cell in row])
    #         wb.close()
    #         return pd.DataFrame(data, columns=columns)
    # except Exception as e:
    #     raise Exception(f"读取Excel文件时发生错误: {str(e)}")

def get_sheet_names(file_path):
    """获取 Excel 文件的所有页签名称"""
    try:

        if file_path.lower().endswith('.xls'):
            # 处理 .xls 文件
            # 处理 .xls 文件
            import xlrd
            try:
                workbook = xlrd.open_workbook(file_path)
                sheetnames = workbook.sheet_names()
                return sheetnames
            except xlrd.biffh.XLRDError as e:
                # 如果是版本问题，尝试用 openpyxl 处理（可能是实际为 .xlsx 格式的文件）
                if "xlsx file" in str(e).lower():
                    wb = load_workbook(filename=file_path, read_only=True)
                    sheetnames = wb.sheetnames
                    wb.close()
                    return sheetnames
                else:
                    raise e
        else:
            # 处理 .xlsx 文件
            wb = load_workbook(filename=file_path, read_only=True)
            sheetnames = wb.sheetnames
            wb.close()
            return sheetnames
    except Exception as e:
        raise Exception(f"读取页签名称时发生错误: {str(e)}")

def read_mapping_table(file_path):
    """读取资产分类映射表，返回 DataFrame"""
    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        if '资产分类映射表' not in wb.sheetnames:
            raise Exception("未找到'资产分类映射表'页签")

        ws = wb['资产分类映射表']

        # 只读取第二行作为表头（第二级表头）
        headers = [cell.value if cell.value is not None else "" for cell in next(ws.iter_rows(min_row=2, max_row=2))]

        # 读取数据行（从第3行开始）
        data = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            data.append(row)

        df = pd.DataFrame(data, columns=headers)
        wb.close()
        return df
    except Exception as e:
        raise Exception(f"读取资产分类映射表时发生错误: {str(e)}")
