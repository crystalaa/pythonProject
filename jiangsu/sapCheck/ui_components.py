# ui_components.py
import sys
import traceback
import logging
import os
import shutil
from PyQt5.QtWidgets import QWidget, QPushButton, QFileDialog, QLabel, QVBoxLayout, QHBoxLayout, \
    QPlainTextEdit, QTabWidget, QComboBox, QProgressDialog, QApplication
from PyQt5.QtCore import Qt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from data_handler import LoadColumnWorker
from rule_handler import read_rules
from comparator import CompareWorker


class ExcelComparer(QWidget):
    """主窗口类"""

    def __init__(self):
        super().__init__()
        self.file1 = ""
        self.file2 = ""
        self.sheet_name1 = ""
        self.sheet_name2 = ""
        self.initUI()
        self.worker = None
        self.summary_data = {}
        self.columns1 = []
        self.columns2 = []
        self.rules = {}  # 存储解析后的规则
        self.rule_file = ""
        # 初始化 worker 变量
        self.worker_sheet1 = None
        self.worker_sheet2 = None
        self.worker_load1 = None
        self.worker_load2 = None
        self.loading_dialog = None
        # 读取规则文件
        self.load_rules_file()

    def load_rules_file(self):
        """加载规则文件"""
        try:
            # 获取exe文件所在目录
            if hasattr(sys, '_MEIPASS'):
                # 打包后的exe环境
                exe_dir = os.path.dirname(sys.executable)
            else:
                # 开发环境
                exe_dir = os.path.dirname(os.path.abspath(__file__))

            rule_file_path = os.path.join(exe_dir, "rule.xlsx")
            self.rule_file = rule_file_path
            if os.path.exists(rule_file_path):
                self.rules = read_rules(rule_file_path)
                self.log(f"✅ 成功加载规则文件: {rule_file_path}")
            else:
                self.log(f"❌ 未找到规则文件: {rule_file_path}")
                # 可以选择是否继续运行或者退出
        except Exception as e:
            self.log(f"❌ 读取规则文件失败: {str(e)}")

    def initUI(self):
        """初始化用户界面"""
        self.setWindowTitle("Excel文件比较工具V2.4")
        self.resize(1000, 700)

        main_layout = QVBoxLayout()

        # 文件选择区域
        file_layout = QHBoxLayout()

        left_layout = QVBoxLayout()
        self.label1 = QLabel("未选择表一")
        self.btn1 = QPushButton("选择表一")
        self.btn1.clicked.connect(self.select_file1)

        self.sheet_label1 = QLabel("选择表一页签：")
        self.sheet_combo1 = QComboBox()
        self.sheet_combo1.currentTextChanged.connect(self.on_sheet_selection_changed)

        left_layout.addWidget(self.label1)
        left_layout.addWidget(self.btn1)
        left_layout.addWidget(self.sheet_label1)
        left_layout.addWidget(self.sheet_combo1)

        right_layout = QVBoxLayout()
        self.label2 = QLabel("未选择表二")
        self.btn2 = QPushButton("选择表二")
        self.btn2.clicked.connect(self.select_file2)

        self.sheet_label2 = QLabel("选择表二页签：")
        self.sheet_combo2 = QComboBox()
        self.sheet_combo2.currentTextChanged.connect(self.on_sheet_selection_changed)

        right_layout.addWidget(self.label2)
        right_layout.addWidget(self.btn2)
        right_layout.addWidget(self.sheet_label2)
        right_layout.addWidget(self.sheet_combo2)
        file_layout.addLayout(left_layout)
        file_layout.addLayout(right_layout)
        # 按钮区域
        button_layout = QHBoxLayout()
        self.compare_btn = QPushButton("比较文件")
        self.compare_btn.setFixedWidth(150)
        self.compare_btn.clicked.connect(self.compare_files)
        self.compare_btn.setEnabled(False)
        self.export_btn = QPushButton("导出报告")
        self.export_btn.setFixedWidth(150)
        self.export_btn.setEnabled(False)
        self.export_btn.clicked.connect(self.export_report)
        button_layout.addStretch()
        button_layout.addWidget(self.compare_btn)
        button_layout.addWidget(self.export_btn)
        # 日志和报告区域
        self.tab_widget = QTabWidget()
        self.log_area = QPlainTextEdit()
        self.log_area.setReadOnly(True)
        self.log_area.setStyleSheet("background-color: #f0f0f0;")
        self.summary_area = QPlainTextEdit()
        self.summary_area.setReadOnly(True)
        self.summary_area.setStyleSheet("background-color: #f0f0f0;")
        self.tab_widget.addTab(self.log_area, "比对日志")
        self.tab_widget.addTab(self.summary_area, "汇总报告")
        # 主布局组合
        main_layout.addLayout(file_layout)
        main_layout.addLayout(button_layout)
        main_layout.addWidget(self.tab_widget)

        self.setLayout(main_layout)

    def closeEvent(self, event):
        """窗口关闭时确保线程安全退出"""
        if hasattr(self, 'worker') and self.worker is not None and self.worker.isRunning():
            self.worker.quit()
            self.worker.wait()
        if hasattr(self, 'worker_load1') and self.worker_load1 is not None and self.worker_load1.isRunning():
            self.worker_load1.quit()
            self.worker_load1.wait()
        if hasattr(self, 'worker_load2') and self.worker_load2 is not None and self.worker_load2.isRunning():
            self.worker_load2.quit()
            self.worker_load2.wait()
        if hasattr(self, 'worker_sheet1') and self.worker_sheet1 is not None and self.worker_sheet1.isRunning():
            self.worker_sheet1.quit()
            self.worker_sheet1.wait()
        if hasattr(self, 'worker_sheet2') and self.worker_sheet2 is not None and self.worker_sheet2.isRunning():
            self.worker_sheet2.quit()
            self.worker_sheet2.wait()
        super().closeEvent(event)

    def reset_file_state(self, is_file1=True, is_file2=False):
        if is_file1:
            self.columns1 = []
            self.sheet_combo1.clear()
            self.sheet_combo1.setEnabled(True)
            self.sheet_label1.setText("选择表一页签：")
            if hasattr(self, 'worker_sheet1'):
                self.worker_sheet1 = None
        if is_file2:
            self.columns2 = []
            self.sheet_combo2.clear()
            self.sheet_combo2.setEnabled(True)
            self.sheet_label2.setText("选择表二页签：")
            if hasattr(self, 'worker_sheet2'):
                self.worker_sheet2 = None
        self.compare_btn.setEnabled(False)
        self.log_area.clear()
        self.summary_area.clear()
        self.export_btn.setEnabled(False)

    def select_file1(self):
        self.reset_file_state(is_file1=True, is_file2=False)
        file, _ = QFileDialog.getOpenFileName(self, "选择 Excel 文件", "", "Excel 文件 (*.xlsx *.xls)")
        if file:
            self.file1 = file
            filename = os.path.basename(file)
            self.label1.setText(f"表一: {filename}")
            # 显示加载对话框
            self.show_loading_dialog("正在加载表一页签...")
            self.load_sheet_and_columns(file, is_file1=True)

    def select_file2(self):
        self.reset_file_state(is_file1=False, is_file2=True)
        file, _ = QFileDialog.getOpenFileName(self, "选择 Excel 文件", "", "Excel 文件 (*.xlsx *.xls)")
        if file:
            self.file2 = file
            filename = os.path.basename(file)

            self.label2.setText(f"表二: {filename}")
            self.show_loading_dialog("正在加载表二页签...")
            self.load_sheet_and_columns(file, is_file2=True)

    def show_loading_dialog(self, message="正在加载，请稍候..."):
        """显示加载对话框"""
        if not self.loading_dialog:
            self.loading_dialog = QProgressDialog(message, None, 0, 0, self)
            self.loading_dialog.setWindowModality(Qt.WindowModal)
            self.loading_dialog.setWindowTitle("加载中")
            self.loading_dialog.setCancelButton(None)
            self.loading_dialog.show()

    def load_sheet_and_columns(self, file_path, is_file1=False, is_file2=False):

        worker = LoadColumnWorker(file_path)
        worker.sheet_names_loaded.connect(self.on_sheet_names_loaded)
        worker.sheet_names_loaded.connect(self.close_loading_dialog)
        # worker.columns_loaded.connect(self.on_columns_loaded)
        # worker.error_occurred.connect(self.on_column_error)
        if is_file1:
            self.worker_load1 = worker
        elif is_file2:
            self.worker_load2 = worker
        worker.start()

    def on_sheet_names_loaded(self, file_path, sheet_names):
        if file_path == self.file1:
            self.sheet_combo1.clear()
            self.sheet_combo1.addItems(sheet_names)
            self.sheet_combo1.setCurrentIndex(0)
        elif file_path == self.file2:
            self.sheet_combo2.clear()
            self.sheet_combo2.addItems(sheet_names)
            self.sheet_combo2.setCurrentIndex(0)

    def on_sheet_selection_changed(self):
        """页签选择变化时的处理函数"""
        # 简单更新比较按钮状态
        self.update_compare_button_state()

    def update_compare_button_state(self):
        sheet_selected = self.sheet_combo1.currentText() and self.sheet_combo2.currentText()
        if not sheet_selected:
            self.compare_btn.setEnabled(False)
            return

        self.compare_btn.setEnabled(True)

    def compare_files(self):
        if not self.file1 or not self.file2:
            self.log("请先选择两个 Excel 文件！")
            return
        sheet_name1 = self.sheet_combo1.currentText()
        sheet_name2 = self.sheet_combo2.currentText()
        if not sheet_name1 or not sheet_name2:
            self.log("请选择两个文件的页签！")
            return

        self.log_area.clear()
        self.summary_area.clear()
        self.export_btn.setEnabled(False)

        # 获取主键字段
        primary_keys = [field for field, rule in self.rules.items() if rule["is_primary"]]
        if not primary_keys:
            self.log("规则文件中未定义主键字段，请检查规则文件！")
            return
        self.loading_dialog = QProgressDialog("正在比较文件，请稍候...", None, 0, 0, self)
        self.loading_dialog.setWindowModality(Qt.WindowModal)
        self.loading_dialog.setWindowTitle("比较中")
        self.loading_dialog.setCancelButton(None)
        self.loading_dialog.show()

        self.worker = CompareWorker(self.file1, self.file2, self.rule_file, sheet_name1, sheet_name2,
                                    primary_keys=primary_keys,
                                    rules=self.rules)
        self.worker.log_signal.connect(self.log)
        # 连接信号以在比较完成时关闭对话框
        self.worker.finished.connect(self.close_loading_dialog)
        self.worker.finished.connect(lambda: self.export_btn.setEnabled(True))
        self.worker.finished.connect(self.on_compare_finished)
        self.worker.start()

    def close_loading_dialog(self):
        """关闭加载对话框"""
        if self.loading_dialog:
            self.loading_dialog.close()
            self.loading_dialog = None

    def on_compare_finished(self):
        try:
            if hasattr(self.worker, 'summary'):
                self.summary_data = self.worker.summary
                primary_key = self.summary_data.get("primary_key", "主键")
                total_file1 = self.summary_data['total_file1']
                total_file2 = self.summary_data['total_file2']
                missing_count = self.summary_data['missing_count']
                extra_count = self.summary_data.get('extra_count', 0)
                common_count = self.summary_data['common_count']
                diff_count = self.summary_data['diff_count']
                equal_count = self.summary_data['equal_count']
                diff_ratio = self.summary_data['diff_ratio']
                missing_columns = self.summary_data.get("missing_columns", [])
                missing_columns_str = ", ".join(missing_columns) if missing_columns else "无"

                summary_text = (
                    f"📊 比对汇总报告\n"
                    f"--------------------------------\n"
                    f"• 总{primary_key}数量（表一）：{total_file1}\n"
                    f"• 总{primary_key}数量（表二）：{total_file2}\n"
                    f"• 表二中缺失的{primary_key}：{missing_count}\n"
                    f"• 表二中多出的{primary_key}：{extra_count}\n"
                    f"• 共同{primary_key}数量：{common_count}\n"
                    f"• 列不一致的{primary_key}数量：{diff_count}\n"
                    f"• 列一致的{primary_key}数量：{equal_count}\n"
                    f"• 表二中缺失的列：{missing_columns_str}\n"
                    f"--------------------------------\n"
                    f"• 差异数据占比：{diff_ratio:.2%}\n"
                )
                self.summary_area.setPlainText(summary_text)
                self.export_btn.setEnabled(True)
        except Exception as e:
            self.summary_area.setPlainText(f"❌ 显示汇总报告时发生错误：{str(e)}\n请查看比对日志了解详细信息。")
            self.export_btn.setEnabled(False)

    def export_report(self):
        """复制原始文件并修改副本，添加对比结果和差异详情"""
        if not hasattr(self, 'worker') or not hasattr(self.worker, 'missing_rows') or not hasattr(self.worker,
                                                                                                  'diff_full_rows'):
            self.log("没有可导出的数据，请先执行比对！")
            return

        try:
            # 获取保存路径
            directory = QFileDialog.getExistingDirectory(self, "选择保存路径")
            if not directory:
                self.log("导出已取消。")
                return

            # 复制并修改表一文件
            file1_name = os.path.splitext(os.path.basename(self.file1))[0]
            file1_copy_path = f"{directory}/{file1_name}_比对结果.xlsx"
            import shutil
            shutil.copy2(self.file1, file1_copy_path)
            self._modify_original_file(file1_copy_path, self.sheet_combo1.currentText(), is_first_file=True)

            # 复制并修改表二文件
            file2_name = os.path.splitext(os.path.basename(self.file2))[0]
            file2_copy_path = f"{directory}/{file2_name}_比对结果.xlsx"
            shutil.copy2(self.file2, file2_copy_path)
            self._modify_original_file(file2_copy_path, self.sheet_combo2.currentText(), is_first_file=False)

            self.log(f"✅ 已生成比对结果文件：{file1_copy_path} 和 {file2_copy_path}")
        except Exception as e:
            self.log(f"❌ 生成比对结果文件时发生错误：{str(e)}")

    def _modify_original_file(self, file_path, sheet_name, is_first_file):
        """直接修改原始Excel文件 - 性能优化版本"""
        try:
            # 加载工作簿
            wb = load_workbook(file_path)
            ws = wb[sheet_name]

            # 获取主键
            primary_keys = [field for field, rule in self.rules.items() if rule["is_primary"]]

            # 获取需要比对的列
            compare_columns = list(self.rules.keys())

            # 创建红色填充样式
            red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

            # 预处理数据 - 构建主键到差异数据的映射字典
            diff_dict = {}
            missing_in_file2_keys = set()  # 表一有但表二没有的主键
            missing_in_file1_keys = set()  # 表二有但表一没有的主键

            if hasattr(self.worker, 'diff_full_rows'):
                for item in self.worker.diff_full_rows:
                    # 构建主键 - 需要与对比部分使用相同的逻辑
                    if is_first_file:
                        # 表一文件使用source数据构建主键
                        key_parts = [str(item['source'].get(pk, '')) for pk in primary_keys]
                    else:
                        # 表二文件使用target数据构建主键（与对比逻辑一致）
                        key_parts = [str(item['target'].get(pk, '')) for pk in primary_keys]

                    # 处理多主键拼接（与对比部分一致）
                    key = ' + '.join(key_parts) if len(key_parts) > 1 else (key_parts[0] if key_parts else "")
                    diff_dict[key] = item

            # 处理缺失数据的主键
            if hasattr(self.worker, 'missing_rows'):
                for row in self.worker.missing_rows:
                    # 表一中存在但表二中缺失的数据，使用表一的主键
                    key_parts = [str(row.get(pk, '')) for pk in primary_keys]
                    key = ' + '.join(key_parts) if len(key_parts) > 1 else (key_parts[0] if key_parts else "")
                    missing_in_file2_keys.add(key)

            # 处理多余数据的主键
            if hasattr(self.worker, 'extra_in_file2'):
                for row in self.worker.extra_in_file2:
                    # 表二中存在但表一中缺失的数据，使用表二的主键
                    key_parts = [str(row.get(pk, '')) for pk in primary_keys]
                    key = ' + '.join(key_parts) if len(key_parts) > 1 else (key_parts[0] if key_parts else "")
                    missing_in_file1_keys.add(key)

            # 创建列名到列索引的映射（一次性处理）
            col_name_to_index = {}
            for col_idx in range(1, ws.max_column + 1):
                col_name = ws.cell(row=1, column=col_idx).value
                if col_name:
                    # 清理列名（去除*和空格）
                    cleaned_col_name = str(col_name).replace('*', '').strip()
                    col_name_to_index[cleaned_col_name] = col_idx

            # 在第一行添加新列标题
            max_col = ws.max_column
            ws.cell(row=1, column=max_col + 1, value="对比结果")
            for i, col in enumerate(compare_columns):
                ws.cell(row=1, column=max_col + 2 + i, value=f"{col}")

            # 创建一个辅助函数来计算主键值（与对比逻辑保持一致）
            def calculate_composite_key(row_data, is_table2=False):
                """根据规则计算复合主键值"""
                key_parts = []

                for pk in primary_keys:
                    # 获取主键对应的规则
                    pk_rule = self.rules.get(pk)

                    # 如果是表二且主键有计算规则，则按规则计算
                    if is_table2 and pk_rule and pk_rule.get("calc_rule"):
                        # 对于表二，如果主键需要计算，则使用计算规则
                        calc_rule = pk_rule["calc_rule"]
                        data_type = pk_rule["data_type"]

                        try:
                            # 模拟计算过程（简化版）
                            # 实际应该使用与CompareWorker中相同的calculate_field方法
                            if '+' in calc_rule and data_type == "文本":
                                # 字符串拼接情况，如"公司代码+资产编码"
                                fields = [f.strip() for f in calc_rule.split('+')]
                                concatenated_value = ""
                                for field in fields:
                                    field_col_idx = col_name_to_index.get(field)
                                    if field_col_idx and row_data.get(field_col_idx):
                                        concatenated_value += str(row_data[field_col_idx])
                                key_parts.append(concatenated_value)
                            else:
                                # 其他情况使用直接获取的值
                                pk_col_idx = col_name_to_index.get(pk)
                                if pk_col_idx and row_data.get(pk_col_idx):
                                    key_parts.append(str(row_data[pk_col_idx]))
                                else:
                                    key_parts.append("")
                        except:
                            # 出错时使用直接获取的值
                            pk_col_idx = col_name_to_index.get(pk)
                            if pk_col_idx and row_data.get(pk_col_idx):
                                key_parts.append(str(row_data[pk_col_idx]))
                            else:
                                key_parts.append("")
                    else:
                        # 表一或其他不需要计算的情况，直接使用值
                        pk_col_idx = col_name_to_index.get(pk)
                        if pk_col_idx and row_data.get(pk_col_idx):
                            key_parts.append(str(row_data[pk_col_idx]))
                        else:
                            key_parts.append("")

                return ' + '.join(key_parts) if len(key_parts) > 1 else (key_parts[0] if key_parts else "")

            # 批量处理所有数据行，减少重复计算
            row_updates = []  # 收集所有需要更新的行信息
            fill_operations = []  # 收集所有需要标红的操作

            # 先收集所有行的信息
            for row_idx in range(2, ws.max_row + 1):
                # 读取当前行的所有数据
                row_data = {}
                for col_idx in range(1, ws.max_column + 1):
                    row_data[col_idx] = ws.cell(row=row_idx, column=col_idx).value

                # 构建当前行的主键（与对比部分保持一致）
                if not is_first_file:  # 表二文件
                    # 使用表二的主键计算逻辑
                    key = calculate_composite_key(row_data, is_table2=True)
                else:  # 表一文件
                    # 表一使用直接获取的主键值
                    key_parts = []
                    for pk in primary_keys:
                        pk_col_idx = col_name_to_index.get(pk)
                        if pk_col_idx and row_data.get(pk_col_idx):
                            key_parts.append(str(row_data[pk_col_idx]))
                        else:
                            key_parts.append("")
                    key = ' + '.join(key_parts) if len(key_parts) > 1 else (key_parts[0] if key_parts else "")

                # 确定对比结果
                comparison_result = ""
                if key in missing_in_file2_keys:
                    comparison_result = "此数据不存在于SAP" if is_first_file else "此数据不存在于平台"
                elif key in missing_in_file1_keys:
                    comparison_result = "此数据不存在于平台" if is_first_file else "此数据不存在于SAP"
                elif key in diff_dict:
                    comparison_result = "不一致"
                else:
                    comparison_result = "一致"

                # 收集该行需要的更新信息
                row_updates.append({
                    'row_idx': row_idx,
                    'key': key,
                    'comparison_result': comparison_result,
                    'row_data': row_data
                })

            # 批量执行更新操作，减少与Excel文件的交互次数
            for update_info in row_updates:
                row_idx = update_info['row_idx']
                key = update_info['key']
                comparison_result = update_info['comparison_result']
                row_data = update_info['row_data']

                # 填入对比结果
                ws.cell(row=row_idx, column=max_col + 1, value=comparison_result)

                # 填入各列的差异详情
                if key in diff_dict:
                    diff_data = diff_dict[key]
                    # 根据是表一还是表二来获取正确的数据源
                    if is_first_file:
                        source_data = diff_data.get('source', {})
                        target_data = diff_data.get('target', {})
                    else:
                        source_data = diff_data.get('source', {})
                        target_data = diff_data.get('target', {})

                    for i, col in enumerate(compare_columns):
                        if col in source_data and col in target_data:
                            val1 = source_data[col]
                            val2 = target_data[col]

                            # 获取该列的规则
                            rule = self.rules.get(col, {})
                            data_type = rule.get("data_type", "文本")  # 默认为文本类型
                            tail_diff = rule.get("tail_diff")

                            # 使用规则判断值是否相等
                            are_equal = self.worker.values_equal_by_rule(val1, val2, data_type, tail_diff, col)
                            if not are_equal:
                                # 如果是资产分类且有映射，使用原始值
                                if col == "资产分类" and hasattr(self.worker, 'asset_code_to_original'):
                                    original_val1 = self.worker.asset_code_to_original.get(val1, val1)
                                    original_val2 = self.worker.asset_code_to_original.get(val2, val2)
                                    diff_detail = f"不一致：表一={original_val1}, 表二={original_val2}"
                                else:
                                    diff_detail = f"不一致：表一={val1}, 表二={val2}"

                                ws.cell(row=row_idx, column=max_col + 2 + i, value=diff_detail)

                                # 记录需要标红的单元格
                                if comparison_result == "不一致":
                                    fill_operations.append((row_idx, max_col + 2 + i))

                    # 记录需要标红的对比结果单元格
                    if comparison_result in ["不一致", "此数据不存在于SAP", "此数据不存在于平台"]:
                        fill_operations.append((row_idx, max_col + 1))

            # 批量执行所有标红操作
            for row_idx, col_idx in fill_operations:
                ws.cell(row=row_idx, column=col_idx).fill = red_fill

            # 保存修改后的文件
            wb.save(file_path)
            wb.close()

        except Exception as e:
            self.log(f"修改文件 {file_path} 时出错: {str(e)}")
            raise e

    def log(self, message):
        """日志输出"""
        self.log_area.appendPlainText(message)


def exception_hook(exc_type, exc_value, exc_traceback):
    """全局异常钩子，防止崩溃"""
    try:
        ex = QApplication.instance().topLevelWidgets()[0]
        if hasattr(ex, "log"):
            error_message = "".join(traceback.format_exception(exc_type, exc_value, exc_traceback))
            logging.error(error_message)
            ex.log(f"❌ 发生异常：{exc_value}")
        else:
            sys.__excepthook__(exc_type, exc_value, exc_traceback)
    except:
        sys.__excepthook__(exc_type, exc_value, exc_traceback)
