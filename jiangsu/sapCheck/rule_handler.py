# rule_handler.py
import pandas as pd
from openpyxl import load_workbook

def read_rules(file_path):
    """读取规则文件，返回规则字典"""
    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        ws = wb['比对规则']
        rules = {}

        for row in ws.iter_rows(min_row=2, values_only=True):  # 假设第一行是标题
            table1_field, table2_field, data_type, tail_diff, is_primary, calc_rule = row[:6]
            if table1_field is None or table2_field is None:
                continue  # 跳过空行
            rules[table1_field] = {
                "table2_field": table2_field,
                "data_type": data_type.lower(),
                "tail_diff": tail_diff,
                "is_primary": is_primary == "是",
                "calc_rule": calc_rule  # 新增：存储计算规则
            }
        wb.close()
        return rules
    except Exception as e:
        raise Exception(f"读取规则文件时发生错误: {str(e)}")

def read_enum_mapping(rule_file):
    """
    读取规则文件中的'枚举值-线站电压等级'页签
    返回 dict: 名称 -> 编码
    """
    try:
        df = pd.read_excel(rule_file, sheet_name='枚举值-线站电压等级', dtype=str)
        # 假设列名就是“编码”和“名称”
        df = df[['编码', '名称']].dropna()
        return dict(zip(df['名称'].astype(str).str.strip(),
                        df['编码'].astype(str).str.strip()))
    except Exception as e:
        raise Exception(f"读取枚举值映射失败: {e}")