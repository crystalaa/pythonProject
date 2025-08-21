# rule_handler.py
import pandas as pd
from openpyxl import load_workbook

def read_rules(file_path):
    """读取规则文件，返回规则字典"""
    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        ws = wb['比对规则']
        rules = {}

        for row in ws.iter_rows(min_row=2, values_only=True):  # 假设第一行是标题
            table1_field, table2_field, data_type, tail_diff, is_primary, calc_rule = row[:6]
            if table1_field is None or table2_field is None:
                continue  # 跳过空行
            rules[table1_field] = {
                "table2_field": table2_field,
                "data_type": data_type.lower(),
                "tail_diff": tail_diff,
                "is_primary": is_primary == "是",
                "calc_rule": calc_rule  # 新增：存储计算规则
            }
        wb.close()
        return rules
    except Exception as e:
        raise Exception(f"读取规则文件时发生错误: {str(e)}")

def read_enum_mapping(rule_file):
    """
    读取规则文件中的'枚举值-线站电压等级'页签
    返回 dict: 名称 -> 编码
    """
    try:
        df = pd.read_excel(rule_file, sheet_name='枚举值-线站电压等级', dtype=str)
        # 假设列名就是“编码”和“名称”
        df = df[['编码', '名称']].dropna()
        return dict(zip(df['名称'].astype(str).str.strip(),
                        df['编码'].astype(str).str.strip()))
    except Exception as e:
        raise Exception(f"读取枚举值映射失败: {e}")

# 新增函数
def read_erp_combo_map(rule_file):
    """
    返回 dict：
        key   = 平台单值（如 'A'）
        value = 允许的组合字符串列表（如 ['A', 'A|B', 'A|B|C']）
    """
    df = pd.read_excel(rule_file, sheet_name='枚举值-关联实物管理系统代码及名称', dtype=str)
    # 获取第一列和第三列的数据
    first_col = df.iloc[:, 0]  # 第一列（索引为0）
    third_col = df.iloc[:, 2]  # 第三列（索引为2）

    # 合并这两列成为一个DataFrame
    data_df = pd.DataFrame({
        'platform_code': first_col,
        'erp_code': third_col
    }).dropna()  # 删除空值行

    # 跳过可能的标题行（如果有的话）
    if len(data_df) > 0:
        first_row = data_df.iloc[0]
        # 简单判断是否为标题行（通过检查是否包含常见的标题关键词）
        if any(keyword in str(first_row['platform_code']).lower() for keyword in ['平台', '代码', '标识']) or \
                any(keyword in str(first_row['erp_code']).lower() for keyword in ['erp', '卡片', '标识']):
            data_df = data_df.iloc[1:]  # 跳过标题行

    # 创建映射字典
    combo_map = {}
    for _, row in data_df.iterrows():
        platform_val = str(row['platform_code']).strip()
        erp_val = str(row['erp_code']).strip()

        # 跳过空值行
        if not platform_val or not erp_val or platform_val.lower() == 'nan' or erp_val.lower() == 'nan':
            continue

        # 如果ERP值包含|，则拆分为多个值
        if '|' in erp_val:
            erp_values = set(erp_val.split('|'))
        else:
            erp_values = {erp_val}

        # 添加到映射字典中
        if platform_val in combo_map:
            combo_map[platform_val].update(erp_values)
        else:
            combo_map[platform_val] = erp_values

    return combo_map