import sys
from concurrent.futures import ThreadPoolExecutor

import pandas as pd
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QFileDialog, QLabel, QVBoxLayout, QHBoxLayout, \
    QPlainTextEdit, QProgressBar, QTabWidget, QListWidget, QAbstractItemView
from PyQt5.QtCore import QThread, pyqtSignal, Qt
from openpyxl import load_workbook



def read_excel_columns(file_path, sheet_name="附表1资产卡片期初数据收集模板"):
    """快速读取Excel文件的列名"""
    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        ws = wb[sheet_name]
        columns = [cell.value for cell in next(ws.iter_rows())]
        cleaned_columns = [col.replace('*', '').strip() if isinstance(col, str) else col for col in columns]
        return cleaned_columns
    except Exception as e:
        raise Exception(f"读取Excel列名时发生错误: {str(e)}")
def read_excel_fast(file_path, sheet_name):
    """快速读取Excel文件"""
    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        ws = wb[sheet_name]


        data = []
        columns = None

        for i, row in enumerate(ws.rows):
            if i == 0:
                columns = [cell.value for cell in row]
            else:
                data.append([cell.value for cell in row])

        return pd.DataFrame(data, columns=columns)
    except Exception as e:
        raise Exception(f"读取Excel文件时发生错误: {str(e)}")

class LoadColumnWorker(QThread):
    """用于在独立线程中读取列名"""
    columns_loaded = pyqtSignal(str, list)  # 参数为文件路径和列名列表
    error_occurred = pyqtSignal(str)

    def __init__(self, file_path):
        super().__init__()
        self.file_path = file_path

    def run(self):
        try:
            columns = read_excel_columns(self.file_path)
            self.columns_loaded.emit(self.file_path, columns)
        except Exception as e:
            self.error_occurred.emit(str(e))
            self.summary = {
                "total_file1": 0,
                "total_file2": 0,
                "missing_count": 0,
                "common_count": 0,
                "diff_count": 0,
                "diff_ratio": 0.0
            }
            return


class CompareWorker(QThread):
    """用于在独立线程中执行比较操作"""
    log_signal = pyqtSignal(str)
    progress_signal = pyqtSignal(int)  # 用于更新进度条
    columns_signal = pyqtSignal(list)  # 用于通知主界面更新列名

    def __init__(self, file1, file2, primary_keys=None):
        super().__init__()
        self.file1 = file1
        self.file2 = file2
        self.primary_keys = primary_keys or ['资产编码']  # 默认为资产编码
        self.missing_assets = []
        self.diff_records = []
        self.summary = {}
        self.missing_rows = []  # 存储文件2中缺失的主键对应的文件1整行数据
        self.diff_full_rows = []  # 存储列不一致的文件1和文件2整行数据

    def run(self):
        """线程执行的主函数"""
        try:
            # 创建线程池读取文件
            self.log_signal.emit("正在并行读取Excel文件...")

            with ThreadPoolExecutor(max_workers=2) as executor:
                # 提交两个读取任务
                future1 = executor.submit(read_excel_fast, self.file1, "附表1资产卡片期初数据收集模板")
                future2 = executor.submit(read_excel_fast, self.file2, "附表1资产卡片期初数据收集模板")

                try:
                    df1 = future1.result()
                    df2 = future2.result()
                except Exception as e:
                    raise Exception(f"读取文件时发生错误: {str(e)}")

            self.log_signal.emit("✅ Excel文件读取完成，开始比较数据...")

            # 发送列名
            # self.columns_signal.emit(df1.columns.tolist())

        except KeyError:
            self.log_signal.emit("发生错误：指定的页签不存在，请确认页签名是否正确！")
            return
        except Exception as e:
            self.log_signal.emit(f"发生未知错误：{str(e)}")
            return

        # 清理列名（去除星号和空白）
        df1.columns = df1.columns.str.replace('[*\\s]', '', regex=True)
        df2.columns = df2.columns.str.replace('[*\\s]', '', regex=True)

        # 检查列是否一致
        if not df1.columns.equals(df2.columns):
            self.log_signal.emit("错误：两个文件的列不一致，请检查列名或顺序是否相同！")
            return

        # 检查是否存在主键列
        missing_keys = [key for key in self.primary_keys if key not in df1.columns]
        if missing_keys:
            self.log_signal.emit(f"错误：列中缺少以下主键：{', '.join(missing_keys)}，请检查文件结构！")
            return

        # 设置多主键为索引
        df1.set_index(self.primary_keys, inplace=True)
        df2.set_index(self.primary_keys, inplace=True)

        # ✅ 新增：确保索引是字符串类型
        df1.index = df1.index.map(lambda x: tuple(str(i) for i in x) if isinstance(x, tuple) else str(x))
        df2.index = df2.index.map(lambda x: tuple(str(i) for i in x) if isinstance(x, tuple) else str(x))

        # 提示行数不一致
        if len(df1) != len(df2):
            self.log_signal.emit(f"提示：两个文件的行数不一致（源文件有 {len(df1)} 行，目标文件有 {len(df2)} 行）")

        # 找出缺失的主键
        missing_in_file2 = df1.index.difference(df2.index)
        if not missing_in_file2.empty:
            missing_df = df1.loc[missing_in_file2].copy()
            original_codes = missing_in_file2.astype(str)  # 原始主键列表
            missing_df.reset_index(drop=True, inplace=True)

            # 获取原始列顺序
            columns_order = df1.columns.tolist()

            # 插入主键列为第2列（索引为1）
            for key in self.primary_keys:
                missing_df.insert(1, key, original_codes.map(lambda x: x.split(' ')[0] if ' ' in x else x))

            # 重新排列列顺序
            ordered_columns = []
            for col in columns_order:
                ordered_columns.append(col)
                if col == columns_order[0]:
                    for key in self.primary_keys:
                        ordered_columns.append(key)

            # 确保所有列都存在
            final_columns = [col for col in ordered_columns if col in missing_df.columns]
            missing_df = missing_df[final_columns]

            # 转换为有序字典列表
            self.missing_rows = missing_df.to_dict(orient='records')
            missing_list = "\n".join([f" - {code}" for code in missing_in_file2])
            self.log_signal.emit(f"【目标文件中缺失的主键】（共 {len(missing_in_file2)} 条）：\n{missing_list}")

        # 找出共同主键
        common_codes = df1.index.intersection(df2.index)
        if common_codes.empty:
            self.log_signal.emit("警告：两个文件中没有共同的主键！")
            return

        # 获取共同主键的数据
        df1_common = df1.loc[common_codes]
        df2_common = df2.loc[common_codes]

        # 将数据转换为字符串并替换NaN值
        df1_compare = df1_common.astype(str).replace('nan', '')
        df2_compare = df2_common.astype(str).replace('nan', '')

        # 计算差异
        comparison = df1_compare.compare(df2_compare, align_axis=0)

        # 提取差异数据
        diff_count = 0
        diff_log_messages = []

        mask = (df1_compare == df2_compare).all(axis=1)
        diff_rows = df1_compare[~mask]

        for code, row in diff_rows.iterrows():
            diff_details = []
            for col in df1_compare.columns:
                val1 = df1_compare.loc[code, col]
                val2 = df2_compare.loc[code, col]
                if val1 != val2:
                    diff_details.append(f" - 列 [{col}] 不一致：源文件={val1}, 目标文件={val2}")
                    self.diff_records.append({
                        "主键": str(code),
                        "列名": col,
                        "源文件值": val1,
                        "目标文件值": val2
                    })

            if diff_details:
                diff_log_messages.append(f"\n主键：{code}")
                diff_log_messages.extend(diff_details)

                source_dict = df1_common.loc[code].to_dict()
                target_dict = df2_common.loc[code].to_dict()

                # 保存为有序字典
                self.diff_full_rows.append({
                    "source": source_dict,
                    "target": target_dict
                })

                diff_count += 1
                if diff_count % 1000 == 0 or diff_count == len(df1_common):
                    self.progress_signal.emit(int(diff_count / len(df1_common) * 100))

        # 计算统计信息
        equal_count = len(common_codes) - diff_count
        self.summary = {
            "total_file1": len(df1),
            "total_file2": len(df2),
            "missing_count": len(missing_in_file2),
            "common_count": len(common_codes),
            "diff_count": diff_count,
            "equal_count": equal_count,
            "diff_ratio": diff_count / len(common_codes) if len(common_codes) > 0 else 0
        }

        # 输出比较结果
        if diff_count == 0:
            self.log_signal.emit("【共同主键的数据完全一致】，没有差异。")
        else:
            self.log_signal.emit(f"【存在差异的列】（共 {diff_count} 行）：")
            if diff_log_messages:
                self.log_signal.emit('\n'.join(diff_log_messages))
            else:
                self.log_signal.emit("⚠️ 未找到具体差异列，请检查数据是否一致。")


class ExcelComparer(QWidget):
    """主窗口类"""

    def __init__(self):
        super().__init__()
        self.file1 = ""
        self.file2 = ""
        self.initUI()
        self.worker = None
        self.summary_data = {}
        self.columns1 = []
        self.columns2 = []

    def initUI(self):
        """初始化用户界面"""
        self.setWindowTitle("Excel文件比较工具")
        self.resize(1000, 700)

        main_layout = QVBoxLayout()

        top_layout = QHBoxLayout()

        file_select_layout = QVBoxLayout()
        self.label1 = QLabel("未选择源文件")
        self.btn1 = QPushButton("选择源文件")
        self.btn1.clicked.connect(self.select_file1)

        self.label2 = QLabel("未选择目标文件")
        self.btn2 = QPushButton("选择目标文件")
        self.btn2.clicked.connect(self.select_file2)

        file_select_layout.addWidget(self.label1)
        file_select_layout.addWidget(self.btn1)
        file_select_layout.addWidget(self.label2)
        file_select_layout.addWidget(self.btn2)

        # 主键选择控件
        self.primary_key_label = QLabel("选择主键列（可多选）：")
        self.primary_key_list = QListWidget()
        self.primary_key_list.setSelectionMode(QAbstractItemView.MultiSelection)
        self.primary_key_list.setEnabled(False)
        self.primary_key_list.itemSelectionChanged.connect(self.on_primary_key_selection_changed)

        file_select_layout.addWidget(self.primary_key_label)
        file_select_layout.addWidget(self.primary_key_list)

        button_layout = QVBoxLayout()
        self.compare_btn = QPushButton("比较文件")
        self.compare_btn.setFixedWidth(150)
        self.compare_btn.clicked.connect(self.compare_files)
        self.compare_btn.setEnabled(False)

        self.export_btn = QPushButton("导出报告")
        self.export_btn.setFixedWidth(150)
        self.export_btn.setEnabled(False)
        self.export_btn.clicked.connect(self.export_report)

        button_layout.addStretch()
        button_layout.addWidget(self.compare_btn)
        button_layout.addWidget(self.export_btn)
        button_layout.addStretch()

        top_layout.addLayout(file_select_layout, 2)
        top_layout.addLayout(button_layout, 1)

        self.tab_widget = QTabWidget()
        self.log_area = QPlainTextEdit()
        self.log_area.setReadOnly(True)
        self.log_area.setStyleSheet("background-color: #f0f0f0;")

        self.summary_area = QPlainTextEdit()
        self.summary_area.setReadOnly(True)
        self.summary_area.setStyleSheet("background-color: #f0f0f0;")

        self.tab_widget.addTab(self.log_area, "比对日志")
        self.tab_widget.addTab(self.summary_area, "汇总报告")

        # 添加进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        self.progress_bar.setFixedHeight(20)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setAlignment(Qt.AlignCenter)

        main_layout.addLayout(top_layout)
        main_layout.addWidget(self.tab_widget)
        main_layout.addWidget(self.progress_bar)

        self.setLayout(main_layout)

    def on_primary_key_selection_changed(self):
        """主键选择变化时更新比较按钮状态"""
        self.update_compare_button_state()

    def select_file1(self):
        """选择源文件"""
        file, _ = QFileDialog.getOpenFileName(self, "选择 Excel 文件", "", "Excel 文件 (*.xlsx *.xls)")
        if file:
            self.file1 = file
            self.label1.setText(f"源文件: {file}")
            self.load_columns(file, is_file1=True)

    def select_file2(self):
        """选择目标文件"""
        file, _ = QFileDialog.getOpenFileName(self, "选择 Excel 文件", "", "Excel 文件 (*.xlsx *.xls)")
        if file:
            self.file2 = file
            self.label2.setText(f"目标文件: {file}")
            self.load_columns(file, is_file2=True)

    def update_compare_button_state(self):
        """根据主键是否加载和是否选择来更新比较按钮状态"""
        if not self.columns1 or not self.columns2:
            self.compare_btn.setEnabled(False)
            return

        selected_items = self.primary_key_list.selectedItems()
        primary_keys = [item.text() for item in selected_items]

        # 如果主键列存在且用户选择了主键，启用按钮
        if primary_keys:
            self.compare_btn.setEnabled(True)
        else:
            self.compare_btn.setEnabled(False)


    def load_columns(self, file_path, is_file1=False, is_file2=False):
        """加载列名，使用线程避免卡顿"""
        worker = LoadColumnWorker(file_path)
        worker.columns_loaded.connect(self.on_columns_loaded)
        worker.error_occurred.connect(self.on_column_error)
        if is_file1:
            self.worker_load1 = worker
        elif is_file2:
            self.worker_load2 = worker
        worker.start()

    def on_columns_loaded(self, file_path, columns):
        """处理列名加载完成"""
        if file_path == self.file1:
            self.columns1 = columns
        elif file_path == self.file2:
            self.columns2 = columns

        # 如果两个文件都已加载列名
        if self.columns1 and self.columns2:
            if self.columns1 != self.columns2:
                self.log("警告：两个文件的列名不一致，请确认列名是否一致。")
            else:
                self.update_primary_key_list(self.columns1)

            # 检查主键是否存在
            selected_items = self.primary_key_list.selectedItems()
            primary_keys = [item.text() for item in selected_items] or ['资产编码']
            missing_in_file1 = [key for key in primary_keys if key not in self.columns1]
            missing_in_file2 = [key for key in primary_keys if key not in self.columns2]
            if missing_in_file1:
                self.log(f"错误：主键列 {', '.join(missing_in_file1)} 在源文件中不存在！")
            if missing_in_file2:
                self.log(f"错误：主键列 {', '.join(missing_in_file2)} 在目标文件中不存在！")
        self.update_compare_button_state()
    def on_column_error(self, error_msg):
        """处理列名读取错误"""
        self.log(f"列名读取错误：{error_msg}")

    def update_primary_key_list(self, columns):
        """更新主键列表"""
        self.primary_key_list.clear()
        self.primary_key_list.addItems(columns)
        self.primary_key_list.setEnabled(True)

        # 默认选中“资产编码”
        if '资产编码' in columns:
            self.primary_key_list.findItems('资产编码', Qt.MatchExactly)[0].setSelected(True)

    def compare_files(self):
        """开始比较文件"""
        if not self.file1 or not self.file2:
            self.log("请先选择两个 Excel 文件！")
            return

        self.log_area.clear()
        self.summary_area.clear()
        self.export_btn.setEnabled(False)

        # 获取用户选择的主键
        selected_items = self.primary_key_list.selectedItems()
        primary_keys = [item.text() for item in selected_items] or ['资产编码']

        # 校验主键是否在两个文件中都存在
        missing_in_file1 = [key for key in primary_keys if key not in self.columns1]
        missing_in_file2 = [key for key in primary_keys if key not in self.columns2]
        if missing_in_file1 or missing_in_file2:
            error_msg = ""
            if missing_in_file1:
                error_msg += f"主键列 {', '.join(missing_in_file1)} 在源文件中不存在！\n"
            if missing_in_file2:
                error_msg += f"主键列 {', '.join(missing_in_file2)} 在目标文件中不存在！"
            self.log(error_msg)
            return

        # 创建并启动比较线程
        self.worker = CompareWorker(self.file1, self.file2, primary_keys=primary_keys)
        self.worker.log_signal.connect(self.log)
        self.worker.progress_signal.connect(self.update_progress)
        self.worker.columns_signal.connect(self.update_primary_key_list)  # 新增信号连接
        self.worker.finished.connect(lambda: self.progress_bar.setValue(100))
        self.worker.finished.connect(lambda: self.export_btn.setEnabled(True))
        self.worker.finished.connect(self.on_compare_finished)
        self.worker.start()

    def update_progress(self, value):
        """更新进度条"""
        self.progress_bar.setValue(value)

    def on_compare_finished(self):
        """比较完成后的处理"""
        if hasattr(self.worker, 'summary'):
            self.summary_data = self.worker.summary
            summary_text = (
                f"📊 比对汇总报告\n"
                f"--------------------------------\n"
                f"• 总资产编码数量（源文件）：{self.summary_data['total_file1']}\n"
                f"• 总资产编码数量（目标文件）：{self.summary_data['total_file2']}\n"
                f"• 目标文件中缺失的资产编码：{self.summary_data['missing_count']}\n"
                f"• 共同资产编码数量：{self.summary_data['common_count']}\n"
                f"• 列不一致的资产编码数量：{self.summary_data['diff_count']}\n"
                f"• 列一致的资产编码数量：{self.summary_data['equal_count']}\n"
                f"--------------------------------\n"
                f"• 差异数据占比：{self.summary_data['diff_ratio']:.2%}\n"
            )
            self.summary_area.setPlainText(summary_text)
            self.export_btn.setEnabled(True)

    def export_report(self):
        """导出报告到一个Excel文件，包含两个sheet"""
        if not hasattr(self, 'worker') or not hasattr(self.worker, 'missing_assets') or not hasattr(self.worker,
                                                                                                    'diff_records'):
            self.log("没有可导出的数据，请先执行比对！")
            return

        directory = QFileDialog.getExistingDirectory(self, "选择保存路径")
        if not directory:
            self.log("导出已取消。")
            return

        output_file = f"{directory}/资产比对结果报告.xlsx"

        # 使用 pandas 的 ExcelWriter 导出多个sheet
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # 导出缺失资产编码
            if self.worker.missing_rows:
                missing_df = pd.DataFrame(self.worker.missing_rows)
                missing_df.to_excel(writer, sheet_name='缺失数据', index=False)

            # 导出列不一致的数据并高亮差异列
            if self.worker.diff_full_rows:
                self._export_diff_data_with_highlight_to_sheet(writer, '列不一致数据', self.worker.diff_full_rows)

        self.log(f"✅ 已导出：{output_file}")

    def _export_diff_data_with_highlight_to_sheet(self, writer, sheet_name, diff_full_rows):
        """将差异数据导出到指定的 sheet，并高亮不一致的列"""
        from openpyxl.styles import PatternFill

        wb = writer.book
        ws = wb.create_sheet(sheet_name)

        # 获取列顺序（以第一个目标行为准）
        first_target = diff_full_rows[0]["target"]
        headers = list(first_target.keys())

        # 写入表头
        ws.append(headers)

        red_fill = PatternFill(start_color="FFEE1111", end_color="FFEE1111", fill_type="solid")

        for row_data in diff_full_rows:
            target_data = row_data["target"]
            source_data = row_data["source"]

            target_row = [target_data.get(k, '') for k in headers]
            target_row_idx = ws.max_row + 1
            ws.append(target_row)

            # 高亮不一致的列
            for col_idx, key in enumerate(headers, start=1):
                val1 = source_data.get(key, '')
                val2 = target_data.get(key, '')
                if val1 != val2:
                    ws.cell(row=target_row_idx, column=col_idx).fill = red_fill

    def log(self, message):
        """日志输出"""
        self.log_area.appendPlainText(message)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    ex = ExcelComparer()
    ex.show()
    sys.exit(app.exec_())
