import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from utils import normalize_value


def read_rules(file_path):
    """读取规则文件，返回规则字典"""
    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        ws = wb['比对规则']
        rules = {}

        for row in ws.iter_rows(min_row=2, values_only=True):  # 假设第一行是标题
            table1_field, table2_field, data_type, tail_diff, is_primary = row
            if table1_field is None or table2_field is None:
                continue  # 跳过空行
            rules[table1_field] = {
                "table2_field": table2_field,
                "data_type": data_type.lower(),
                "tail_diff": tail_diff,
                "is_primary": is_primary == "是"
            }
        wb.close()
        return rules
    except Exception as e:
        raise Exception(f"读取规则文件时发生错误: {str(e)}")


def read_excel_columns(file_path, sheet_name):
    """快速读取Excel文件的列名"""
    try:
        if not sheet_name:  # 空字符串、None 都视为未选择
            return
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        ws = wb[sheet_name]
        columns = [cell.value for cell in next(ws.iter_rows())]
        cleaned_columns = [col.replace('*', '').strip() if isinstance(col, str) else col for col in columns]
        wb.close()
        return cleaned_columns
    except Exception as e:
        raise Exception(f"读取Excel列名时发生错误: {str(e)}")


def read_excel_fast(file_path, sheet_name):
    """快速读取Excel文件"""
    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        ws = wb[sheet_name]
        data = []
        columns = None
        for i, row in enumerate(ws.rows):
            if i == 0:
                columns = [cell.value for cell in row]
            else:
                data.append([cell.value for cell in row])
        wb.close()
        return pd.DataFrame(data, columns=columns)
    except Exception as e:
        raise Exception(f"读取Excel文件时发生错误: {str(e)}")


def get_sheet_names(file_path):
    """获取 Excel 文件的所有页签名称"""
    try:
        wb = load_workbook(filename=file_path, read_only=True)
        sheetnames = wb.sheetnames
        wb.close()
        return sheetnames
    except Exception as e:
        raise Exception(f"读取页签名称时发生错误: {str(e)}")


def export_report(output_file, missing_rows, extra_in_file2, diff_full_rows):
    """导出报告到一个Excel文件，包含多个sheet"""
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        if missing_rows:
            missing_df = pd.DataFrame(missing_rows)
            missing_df.to_excel(writer, sheet_name='表二缺失数据', index=False)
        if extra_in_file2:
            extra_df = pd.DataFrame(extra_in_file2)
            extra_df.to_excel(writer, sheet_name='表二多出数据', index=False)
        if diff_full_rows:
            _export_diff_data_with_highlight_to_sheet(writer, '列不一致数据', diff_full_rows)


def _export_diff_data_with_highlight_to_sheet(writer, sheet_name, diff_full_rows):
    """将差异数据导出到指定的 sheet，并高亮不一致的列"""
    if not diff_full_rows:
        return
    wb = writer.book
    ws = wb.create_sheet(sheet_name)

    first_target = diff_full_rows[0]["target"]
    headers = list(first_target.keys())

    ws.append(headers)

    red_fill = PatternFill(start_color="FFEE1111", end_color="FFEE1111", fill_type="solid")

    for row_data in diff_full_rows:
        target_data = row_data["target"]
        source_data = row_data["source"]

        target_row = [target_data.get(k, '') for k in headers]
        target_row_idx = ws.max_row + 1
        ws.append(target_row)

        for col_idx, key in enumerate(headers, start=1):
            val1 = source_data.get(key, '')
            val2 = target_data.get(key, '')
            val1 = normalize_value(val1)
            val2 = normalize_value(val2)

            if val1 != val2 and not (val1 == '' and val2 == ''):
                ws.cell(row=target_row_idx, column=col_idx).fill = red_fill