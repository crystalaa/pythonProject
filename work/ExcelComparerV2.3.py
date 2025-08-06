import sys
import traceback
import logging
import os
import pandas as pd
from concurrent.futures import ThreadPoolExecutor
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QFileDialog, QLabel, QVBoxLayout, QHBoxLayout, \
    QPlainTextEdit, QProgressBar, QTabWidget, QComboBox, QProgressDialog
from PyQt5.QtCore import QThread, pyqtSignal, Qt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# 配置日志记录器
logging.basicConfig(
    filename="error_log.txt",
    level=logging.ERROR,
    format="%(asctime)s - %(levelname)s - %(message)s"
)


def read_rules(file_path):
    """读取规则文件，返回规则字典"""
    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        ws = wb['比对规则']
        rules = {}

        for row in ws.iter_rows(min_row=2, values_only=True):  # 假设第一行是标题
            table1_field, table2_field, data_type, tail_diff, is_primary = row
            if table1_field is None or table2_field is None:
                continue  # 跳过空行
            rules[table1_field] = {
                "table2_field": table2_field,
                "data_type": data_type.lower(),
                "tail_diff": tail_diff,
                "is_primary": is_primary == "是"
            }
        wb.close()
        return rules
    except Exception as e:
        raise Exception(f"读取规则文件时发生错误: {str(e)}")


def resource_path(relative_path):
    """获取打包后资源的绝对路径"""
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)


def read_excel_columns(file_path, sheet_name):
    """快速读取Excel文件的列名"""
    try:
        if not sheet_name:  # 空字符串、None 都视为未选择
            return
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        ws = wb[sheet_name]
        columns = [cell.value for cell in next(ws.iter_rows())]
        cleaned_columns = [col.replace('*', '').strip() if isinstance(col, str) else col for col in columns]
        wb.close()
        return cleaned_columns
    except Exception as e:
        raise Exception(f"读取Excel列名时发生错误: {str(e)}")


def read_excel_fast(file_path, sheet_name):
    """快速读取Excel文件"""
    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
        ws = wb[sheet_name]
        data = []
        columns = None
        for i, row in enumerate(ws.rows):
            if i == 0:
                columns = [cell.value for cell in row]
            else:
                data.append([cell.value for cell in row])
        wb.close()
        return pd.DataFrame(data, columns=columns)
    except Exception as e:
        raise Exception(f"读取Excel文件时发生错误: {str(e)}")


def get_sheet_names(file_path):
    """获取 Excel 文件的所有页签名称"""
    try:
        wb = load_workbook(filename=file_path, read_only=True)
        sheetnames = wb.sheetnames
        wb.close()
        return sheetnames
    except Exception as e:
        raise Exception(f"读取页签名称时发生错误: {str(e)}")


class LoadColumnWorker(QThread):
    """用于在独立线程中读取列名"""
    columns_loaded = pyqtSignal(str, list)  # 参数为文件路径和列名列表
    error_occurred = pyqtSignal(str)
    sheet_names_loaded = pyqtSignal(str, list)

    def __init__(self, file_path, sheet_name=None):
        super().__init__()
        self.file_path = file_path
        self.sheet_name = sheet_name

    def run(self):
        try:
            # 读取页签名称
            sheet_names = get_sheet_names(self.file_path)
            self.sheet_names_loaded.emit(self.file_path, sheet_names)
            if not self.sheet_name:
                return
            # 读取列名
            columns = read_excel_columns(self.file_path, self.sheet_name)
            if columns is None:
                return
            self.columns_loaded.emit(self.file_path, columns)
        except Exception as e:
            self.error_occurred.emit(str(e))


class CompareWorker(QThread):
    """用于在独立线程中执行比较操作"""
    log_signal = pyqtSignal(str)
    progress_signal = pyqtSignal(int)  # 用于更新进度条

    def __init__(self, file1, file2, sheet_name1, sheet_name2, primary_keys=None, rules=None):
        super().__init__()
        self.file1 = file1
        self.file2 = file2
        self.sheet_name1 = sheet_name1
        self.sheet_name2 = sheet_name2
        self.primary_keys = primary_keys if primary_keys else []
        self.rules = rules if rules else {}
        self.missing_assets = []
        self.diff_records = []
        self.summary = {}
        self.missing_rows = []
        self.extra_in_file2 = []
        self.diff_full_rows = []

    @staticmethod
    def normalize_value(val):
        """统一空值表示"""
        if pd.isna(val) or val is None or (isinstance(val, str) and str(val).strip() == ''):
            return ''
        return str(val).strip()

    def run(self):
        try:
            self.log_signal.emit("正在并行读取Excel文件...")

            with ThreadPoolExecutor(max_workers=2) as executor:
                future1 = executor.submit(read_excel_fast, self.file1, self.sheet_name1)
                future2 = executor.submit(read_excel_fast, self.file2, self.sheet_name2)
                try:
                    df1 = future1.result()
                    df2 = future2.result()
                except Exception as e:
                    raise Exception(f"读取文件时发生错误: {str(e)}")

            self.log_signal.emit("✅ Excel文件读取完成，开始比较数据...")
            # 检查数据行是否存在
            if df1.empty:
                self.log_signal.emit("❌ 错误：表一除了表头外没有数据行，请检查文件内容！")
                return

            if df2.empty:
                self.log_signal.emit("❌ 错误：表二除了表头外没有数据行，请检查文件内容！")
                return

            df1.columns = df1.columns.str.replace('[*\\s]', '', regex=True)
            df2.columns = df2.columns.str.replace('[*\\s]', '', regex=True)

            # 检查规则中的列是否在表一和表二表二中都存在
            table1_columns_to_compare = list(self.rules.keys())  # 表一字段名
            table2_columns_to_compare = [rule["table2_field"] for rule in self.rules.values()]  # 表二字段名
            columns_to_compare = list(self.rules.keys())

            missing_in_file1 = [col for col in table1_columns_to_compare if col not in df1.columns]
            missing_in_file2 = [col for col in table2_columns_to_compare if col not in df2.columns]

            if missing_in_file1 or missing_in_file2:
                error_msg = ""
                if missing_in_file1:
                    error_msg += f"表一缺失以下规则定义的列：{', '.join(missing_in_file1)}\n"
                if missing_in_file2:
                    error_msg += f"表二缺失以下规则定义的列：{', '.join(missing_in_file2)}\n"
                self.log_signal.emit(f"❌ 比对失败：{error_msg}")
                return

            # 映射 df2 的列名为 df1 的列名
            mapped_columns = {}
            for field1, rule in self.rules.items():
                field2 = rule["table2_field"]
                if field2 in df2.columns:
                    mapped_columns[field2] = field1
            df2.rename(columns=mapped_columns, inplace=True)

            # 保留需要比对的列
            all_needed_columns = list(set(columns_to_compare + self.primary_keys))
            df1 = df1[all_needed_columns]
            df2 = df2[all_needed_columns]

            #
            # # 确保 df2 包含 df1 的所有列，缺失列视为空值处理
            # missing_columns = [col for col in df1.columns if col not in df2.columns]
            # if missing_columns:
            #     self.log_signal.emit(f"提示：表二缺失以下列：{', '.join(missing_columns)}，将视为空值处理。")
            #     for col in missing_columns:
            #         df2[col] = ''
            # 检查主键列是否为空
            if not self.primary_keys:
                self.log_signal.emit("❌ 错误：规则文件中未定义主键字段，请检查规则文件！")
                return

                # 检查主键列在数据中是否存在
            for pk in self.primary_keys:
                if pk not in df1.columns:
                    self.log_signal.emit(f"❌ 错误：表一中不存在主键列 '{pk}'")
                    return
                if pk not in df2.columns:
                    self.log_signal.emit(f"❌ 错误：表二中不存在主键列 '{pk}'")
                    return

                # 检查主键是否有重复值
            df1_duplicates = df1[df1.duplicated(subset=self.primary_keys, keep=False)]
            if not df1_duplicates.empty:
                duplicate_count = df1_duplicates.shape[0]
                self.log_signal.emit(f"❌ 错误：表一中存在 {duplicate_count} 条重复的主键记录")
                # 显示前几个重复的主键示例
                duplicate_examples = df1_duplicates[self.primary_keys].head(5)
                example_lines = []
                for _, row in duplicate_examples.iterrows():
                    keys = [str(row[pk]) for pk in self.primary_keys]
                    example_lines.append(" + ".join(keys))
                examples = "\n".join([f" - {example}" for example in example_lines])
                self.log_signal.emit(f"重复主键示例（前5个）：\n{examples}")
                return

            df2_duplicates = df2[df2.duplicated(subset=self.primary_keys, keep=False)]
            if not df2_duplicates.empty:
                duplicate_count = df2_duplicates.shape[0]
                self.log_signal.emit(f"❌ 错误：表二中存在 {duplicate_count} 条重复的主键记录")
                # 显示前几个重复的主键示例
                duplicate_examples = df2_duplicates[self.primary_keys].head(5)
                example_lines = []
                for _, row in duplicate_examples.iterrows():
                    keys = [str(row[pk]) for pk in self.primary_keys]
                    example_lines.append(" + ".join(keys))
                examples = "\n".join([f" - {example}" for example in example_lines])
                self.log_signal.emit(f"重复主键示例（前5个）：\n{examples}")
                return

            # 检查主键列是否有空值
            for pk in self.primary_keys:
                df1_empty_keys = df1[pd.isna(df1[pk]) | (df1[pk].astype(str).str.strip() == '')]
                df2_empty_keys = df2[pd.isna(df2[pk]) | (df2[pk].astype(str).str.strip() == '')]

                if len(df1_empty_keys) > 0:
                    self.log_signal.emit(f"⚠️ 警告：表一中主键列 '{pk}' 存在 {len(df1_empty_keys)} 条空值记录")

                if len(df2_empty_keys) > 0:
                    self.log_signal.emit(f"⚠️ 警告：表二中主键列 '{pk}' 存在 {len(df2_empty_keys)} 条空值记录")

            # 保存原始数据帧用于导出（包含主键列）
            df1_original = df1.copy()
            df2_original = df2.copy()

            # 设置主键索引
            df1.set_index(self.primary_keys, inplace=True)
            df2.set_index(self.primary_keys, inplace=True)

            # 规范化索引格式
            df1.index = df1.index.map(lambda x: tuple(str(i) for i in x) if isinstance(x, tuple) else (str(x),))
            df2.index = df2.index.map(lambda x: tuple(str(i) for i in x) if isinstance(x, tuple) else (str(x),))
            # 检查索引中是否有空值
            df1_empty_index = df1.index.map(
                lambda x: any(pd.isna(i) or str(i).strip() == '' for i in (x if isinstance(x, tuple) else (x,))))
            df2_empty_index = df2.index.map(
                lambda x: any(pd.isna(i) or str(i).strip() == '' for i in (x if isinstance(x, tuple) else (x,))))

            df1_empty_count = sum(df1_empty_index)
            df2_empty_count = sum(df2_empty_index)

            if df1_empty_count > 0:
                self.log_signal.emit(f"⚠️ 警告：表一中有 {df1_empty_count} 条记录的主键为空")

            if df2_empty_count > 0:
                self.log_signal.emit(f"⚠️ 警告：表二中有 {df2_empty_count} 条记录的主键为空")

            if len(df1) != len(df2):
                self.log_signal.emit(f"提示：两个文件的行数不一致（表一有 {len(df1)} 行，表二有 {len(df2)} 行）")

            # 查找表二中缺失的主键
            missing_in_file2 = df1.index.difference(df2.index)
            if not missing_in_file2.empty:
                missing_df = df1.loc[missing_in_file2].copy()
                original_codes = missing_in_file2.map(lambda x: ' + '.join(map(str, x)))
                missing_df.reset_index(drop=True, inplace=True)

                for idx, key in enumerate(self.primary_keys):
                    missing_df.insert(1 + idx, key, original_codes.map(lambda x: x.split(' + ')[idx]))

                self.missing_rows = missing_df.to_dict(orient='records')
                missing_list = "\n".join([f" - {code}" for code in missing_in_file2])
                self.log_signal.emit(f"【表二中缺失的主键】（共 {len(missing_in_file2)} 条）：\n{missing_list}")

            # 查找表二中多出的主键
            missing_in_file1 = df2.index.difference(df1.index)
            if not missing_in_file1.empty:
                missing_df_file1 = df2.loc[missing_in_file1].copy()
                original_codes_file1 = missing_in_file1.map(lambda x: ' + '.join(map(str, x)))
                missing_df_file1.reset_index(drop=True, inplace=True)

                for idx, key in enumerate(self.primary_keys):
                    missing_df_file1.insert(1 + idx, key, original_codes_file1.map(lambda x: x.split(' + ')[idx]))

                self.extra_in_file2 = missing_df_file1.to_dict(orient='records')
                missing_list_file1 = "\n".join([f" - {code}" for code in missing_in_file1])
                self.log_signal.emit(
                    f"【表二中多出的主键】（表一中没有，共 {len(missing_in_file1)} 条）：\n{missing_list_file1}")

            # 找出共同的主键
            common_codes = df1.index.intersection(df2.index)
            if common_codes.empty:
                self.log_signal.emit("警告：两个文件中没有共同的主键！")
                return

            # 替换原有的数据比较部分为以下代码：
            try:
                self.log_signal.emit("开始进行向量化数据比较...")
                df1_common = df1.loc[common_codes]
                df2_common = df2.loc[common_codes]

                # 格式化索引
                df1_common.index = df1_common.index.map(lambda x: ' + '.join(x) if isinstance(x, tuple) else str(x))
                df2_common.index = df2_common.index.map(lambda x: ' + '.join(x) if isinstance(x, tuple) else str(x))
                # 使用向量化操作进行批量比较
                diff_dict = {}

                # 预处理索引以便后续使用
                index_mapping = pd.Series(df1_common.index, index=df1_common.index)

                for field1, rule in self.rules.items():
                    # 只比对规则文件中定义的列
                    if field1 not in df1_common.columns or field1 not in df2_common.columns:
                        continue

                    data_type = rule["data_type"]
                    tail_diff = rule.get("tail_diff")

                    # 向量化获取两列数据
                    series1 = df1_common[field1]
                    series2 = df2_common[field1]

                    if data_type == "数值":
                        # 数值型比较
                        series1_num = pd.to_numeric(series1, errors='coerce')
                        series2_num = pd.to_numeric(series2, errors='coerce')

                        if tail_diff is None:
                            diff_mask = (series1_num != series2_num) & \
                                        ~(pd.isna(series1_num) & pd.isna(series2_num))
                        else:
                            diff_mask = (abs(series1_num - series2_num) > float(tail_diff)) & \
                                        ~(pd.isna(series1_num) & pd.isna(series2_num))

                    elif data_type == "日期":
                        # 日期型比较
                        both_empty = (series1.fillna('').astype(str).str.strip() == '') & \
                                     (series2.fillna('').astype(str).str.strip() == '')

                        s1_str = series1.astype(str)
                        s2_str = series2.astype(str)

                        if tail_diff == "月":
                            series1_cmp = s1_str.str[:7]
                            series2_cmp = s2_str.str[:7]
                        elif tail_diff == "日":
                            series1_cmp = s1_str.str[:10]
                            series2_cmp = s2_str.str[:10]
                        elif tail_diff == "时":
                            series1_cmp = s1_str.str[:13]
                            series2_cmp = s2_str.str[:13]
                        elif tail_diff == "分":
                            series1_cmp = s1_str.str[:16]
                            series2_cmp = s2_str.str[:16]
                        elif tail_diff == "秒":
                            series1_cmp = s1_str.str[:19]
                            series2_cmp = s2_str.str[:19]
                        else:
                            series1_cmp = s1_str.str[:4]
                            series2_cmp = s2_str.str[:4]

                        diff_mask = (series1_cmp != series2_cmp) & ~both_empty

                    elif data_type == "文本":
                        # 文本型比较
                        series1_norm = series1.fillna('').astype(str).str.strip()
                        series2_norm = series2.fillna('').astype(str).str.strip()
                        diff_mask = series1_norm != series2_norm

                    # 找出有差异的行索引
                    diff_indices = df1_common[diff_mask].index

                    # 批量添加差异记录
                    for idx in diff_indices:
                        if idx not in diff_dict:
                            diff_dict[idx] = []

                        val1 = CompareWorker.normalize_value(series1.loc[idx])
                        val2 = CompareWorker.normalize_value(series2.loc[idx])
                        diff_dict[idx].append((field1, val1, val2))

                self.log_signal.emit(f"向量化比较完成，共发现 {len(diff_dict)} 条差异记录")

            except Exception as e:
                self.log_signal.emit(f"向量化比较出错，使用传统方法: {str(e)}")
                # 如果向量化方法出错，回退到原来的逐行比较方法

            # 生成日志和汇总信息
            diff_log_messages = []
            self.diff_full_rows = []

            # 创建主键到原始值的映射，用于恢复导出数据中的主键列
            pk_mapping = {}
            for code in common_codes:
                code_str = ' + '.join(code) if isinstance(code, tuple) else str(code)
                pk_mapping[code_str] = code

            for code, diffs in diff_dict.items():
                code_str = ' + '.join(code) if isinstance(code, tuple) else str(code)
                diff_details = [f" - 列 [{col}] 不一致：表一={val1}, 表二={val2}" for col, val1, val2 in diffs]
                diff_log_messages.append(f"\n主键：{code}")
                diff_log_messages.extend(diff_details)

                # 使用原始数据帧查找完整行数据（包含主键列）
                try:
                    # 获取原始主键值
                    original_pk_values = pk_mapping.get(code_str, code)

                    # 构建筛选条件
                    if isinstance(original_pk_values, tuple):
                        # 多主键情况
                        condition1 = True
                        condition2 = True
                        for i, pk in enumerate(self.primary_keys):
                            condition1 = condition1 & (df1_original[pk].astype(str) == original_pk_values[i])
                            condition2 = condition2 & (df2_original[pk].astype(str) == original_pk_values[i])
                    else:
                        # 单主键情况
                        pk = self.primary_keys[0]
                        condition1 = (df1_original[pk].astype(str) == original_pk_values)
                        condition2 = (df2_original[pk].astype(str) == original_pk_values)

                    # 获取完整行数据
                    source_dict = df1_original[condition1].iloc[0].to_dict()
                    target_dict = df2_original[condition2].iloc[0].to_dict()

                    self.diff_full_rows.append({
                        "source": source_dict,
                        "target": target_dict
                    })
                except (IndexError, KeyError, Exception) as e:
                    # 出现异常时使用原来的方法作为备选
                    source_dict = df1_common.loc[code_str].to_dict()
                    target_dict = df2_common.loc[code_str].to_dict()

                    # 手动添加主键信息
                    original_pk_values = pk_mapping.get(code_str, code)
                    if isinstance(original_pk_values, tuple):
                        for i, pk in enumerate(self.primary_keys):
                            source_dict[pk] = original_pk_values[i]
                            target_dict[pk] = original_pk_values[i]
                    else:
                        if self.primary_keys:
                            source_dict[self.primary_keys[0]] = original_pk_values
                            target_dict[self.primary_keys[0]] = original_pk_values

                    self.diff_full_rows.append({
                        "source": source_dict,
                        "target": target_dict
                    })

            diff_count = len(diff_dict)
            equal_count = len(common_codes) - diff_count
            primary_key_str = " + ".join(self.primary_keys)

            self.summary = {
                "primary_key": primary_key_str,
                "total_file1": len(df1),
                "total_file2": len(df2),
                "missing_count": len(missing_in_file2),
                "extra_count": len(missing_in_file1),
                "common_count": len(common_codes),
                "diff_count": diff_count,
                "equal_count": equal_count,
                "diff_ratio": diff_count / len(common_codes) if len(common_codes) > 0 else 0.0,
            }

            if diff_count == 0:
                self.log_signal.emit("【共同主键的数据完全一致】，没有差异。")
            else:
                self.log_signal.emit(f"【存在差异的列】（共 {diff_count} 行）：")
                if diff_log_messages:
                    self.log_signal.emit('\n'.join(diff_log_messages))
                else:
                    self.log_signal.emit("⚠️ 未找到具体差异列，请检查数据是否一致。")

        except Exception as e:
            logging.error(traceback.format_exc())
            self.log_signal.emit(f"发生错误：{str(e)}")
        finally:
            self.quit()
            self.wait()


class ExcelComparer(QWidget):
    """主窗口类"""

    def __init__(self):
        super().__init__()
        self.file1 = ""
        self.file2 = ""
        self.sheet_name1 = ""
        self.sheet_name2 = ""
        self.initUI()
        self.worker = None
        self.summary_data = {}
        self.columns1 = []
        self.columns2 = []
        self.rules = {}  # 存储解析后的规则
        # 初始化 worker 变量
        self.worker_sheet1 = None
        self.worker_sheet2 = None
        self.worker_load1 = None
        self.worker_load2 = None
        self.loading_dialog = None
        # 读取规则文件
        self.load_rules_file()

    def load_rules_file(self):
        """加载规则文件"""
        try:
            # 获取exe文件所在目录
            if hasattr(sys, '_MEIPASS'):
                # 打包后的exe环境
                exe_dir = os.path.dirname(sys.executable)
            else:
                # 开发环境
                exe_dir = os.path.dirname(os.path.abspath(__file__))

            rule_file_path = os.path.join(exe_dir, "rule.xlsx")

            if os.path.exists(rule_file_path):
                self.rules = read_rules(rule_file_path)
                self.log(f"✅ 成功加载规则文件: {rule_file_path}")
            else:
                self.log(f"❌ 未找到规则文件: {rule_file_path}")
                # 可以选择是否继续运行或者退出
        except Exception as e:
            self.log(f"❌ 读取规则文件失败: {str(e)}")
    def initUI(self):
        """初始化用户界面"""
        self.setWindowTitle("Excel文件比较工具V2.4")
        self.resize(1000, 700)

        main_layout = QVBoxLayout()

        # 文件选择区域
        file_layout = QHBoxLayout()

        left_layout = QVBoxLayout()
        self.label1 = QLabel("未选择表一")
        self.btn1 = QPushButton("选择表一")
        self.btn1.clicked.connect(self.select_file1)


        self.sheet_label1 = QLabel("选择表一页签：")
        self.sheet_combo1 = QComboBox()
        self.sheet_combo1.currentTextChanged.connect(self.on_sheet_selection_changed)

        left_layout.addWidget(self.label1)
        left_layout.addWidget(self.btn1)
        left_layout.addWidget(self.sheet_label1)
        left_layout.addWidget(self.sheet_combo1)

        right_layout = QVBoxLayout()
        self.label2 = QLabel("未选择表二")
        self.btn2 = QPushButton("选择表二")
        self.btn2.clicked.connect(self.select_file2)

        self.sheet_label2 = QLabel("选择表二页签：")
        self.sheet_combo2 = QComboBox()
        self.sheet_combo2.currentTextChanged.connect(self.on_sheet_selection_changed)

        right_layout.addWidget(self.label2)
        right_layout.addWidget(self.btn2)
        right_layout.addWidget(self.sheet_label2)
        right_layout.addWidget(self.sheet_combo2)
        file_layout.addLayout(left_layout)
        file_layout.addLayout(right_layout)
        # 按钮区域
        button_layout = QHBoxLayout()
        self.compare_btn = QPushButton("比较文件")
        self.compare_btn.setFixedWidth(150)
        self.compare_btn.clicked.connect(self.compare_files)
        self.compare_btn.setEnabled(False)
        self.export_btn = QPushButton("导出报告")
        self.export_btn.setFixedWidth(150)
        self.export_btn.setEnabled(False)
        self.export_btn.clicked.connect(self.export_report)
        button_layout.addStretch()
        button_layout.addWidget(self.compare_btn)
        button_layout.addWidget(self.export_btn)
        # 日志和报告区域
        self.tab_widget = QTabWidget()
        self.log_area = QPlainTextEdit()
        self.log_area.setReadOnly(True)
        self.log_area.setStyleSheet("background-color: #f0f0f0;")
        self.summary_area = QPlainTextEdit()
        self.summary_area.setReadOnly(True)
        self.summary_area.setStyleSheet("background-color: #f0f0f0;")
        self.tab_widget.addTab(self.log_area, "比对日志")
        self.tab_widget.addTab(self.summary_area, "汇总报告")
        # 主布局组合
        main_layout.addLayout(file_layout)
        main_layout.addLayout(button_layout)
        main_layout.addWidget(self.tab_widget)

        self.setLayout(main_layout)

    def closeEvent(self, event):
        """窗口关闭时确保线程安全退出"""
        if hasattr(self, 'worker') and self.worker is not None and self.worker.isRunning():
            self.worker.quit()
            self.worker.wait()
        if hasattr(self, 'worker_load1') and self.worker_load1 is not None and self.worker_load1.isRunning():
            self.worker_load1.quit()
            self.worker_load1.wait()
        if hasattr(self, 'worker_load2') and self.worker_load2 is not None and self.worker_load2.isRunning():
            self.worker_load2.quit()
            self.worker_load2.wait()
        if hasattr(self, 'worker_sheet1') and self.worker_sheet1 is not None and self.worker_sheet1.isRunning():
            self.worker_sheet1.quit()
            self.worker_sheet1.wait()
        if hasattr(self, 'worker_sheet2') and self.worker_sheet2 is not None and self.worker_sheet2.isRunning():
            self.worker_sheet2.quit()
            self.worker_sheet2.wait()
        super().closeEvent(event)

    def reset_file_state(self, is_file1=True, is_file2=False):
        if is_file1:
            self.columns1 = []
            self.sheet_combo1.clear()
            self.sheet_combo1.setEnabled(True)
            self.sheet_label1.setText("选择表一页签：")
            if hasattr(self, 'worker_sheet1'):
                self.worker_sheet1 = None
        if is_file2:
            self.columns2 = []
            self.sheet_combo2.clear()
            self.sheet_combo2.setEnabled(True)
            self.sheet_label2.setText("选择表二页签：")
            if hasattr(self, 'worker_sheet2'):
                self.worker_sheet2 = None
        self.compare_btn.setEnabled(False)
        self.log_area.clear()
        self.summary_area.clear()
        self.export_btn.setEnabled(False)

    def select_file1(self):
        self.reset_file_state(is_file1=True, is_file2=False)
        file, _ = QFileDialog.getOpenFileName(self, "选择 Excel 文件", "", "Excel 文件 (*.xlsx *.xls)")
        if file:
            self.file1 = file
            filename = os.path.basename(file)
            self.label1.setText(f"表一: {filename}")
            # 显示加载对话框
            self.show_loading_dialog("正在加载表一页签...")
            self.load_sheet_and_columns(file, is_file1=True)

    def select_file2(self):
        self.reset_file_state(is_file1=False, is_file2=True)
        file, _ = QFileDialog.getOpenFileName(self, "选择 Excel 文件", "", "Excel 文件 (*.xlsx *.xls)")
        if file:
            self.file2 = file
            filename = os.path.basename(file)

            self.label2.setText(f"表二: {filename}")
            self.show_loading_dialog("正在加载表二页签...")
            self.load_sheet_and_columns(file, is_file2=True)

    def show_loading_dialog(self, message="正在加载，请稍候..."):
        """显示加载对话框"""
        if not self.loading_dialog:
            self.loading_dialog = QProgressDialog(message, None, 0, 0, self)
            self.loading_dialog.setWindowModality(Qt.WindowModal)
            self.loading_dialog.setWindowTitle("加载中")
            self.loading_dialog.setCancelButton(None)
            self.loading_dialog.show()
    def load_sheet_and_columns(self, file_path, is_file1=False, is_file2=False):

        worker = LoadColumnWorker(file_path)
        worker.sheet_names_loaded.connect(self.on_sheet_names_loaded)
        worker.sheet_names_loaded.connect(self.close_loading_dialog)
        # worker.columns_loaded.connect(self.on_columns_loaded)
        # worker.error_occurred.connect(self.on_column_error)
        if is_file1:
            self.worker_load1 = worker
        elif is_file2:
            self.worker_load2 = worker
        worker.start()

    def on_sheet_names_loaded(self, file_path, sheet_names):
        if file_path == self.file1:
            self.sheet_combo1.clear()
            self.sheet_combo1.addItems(sheet_names)
            self.sheet_combo1.setCurrentIndex(0)
        elif file_path == self.file2:
            self.sheet_combo2.clear()
            self.sheet_combo2.addItems(sheet_names)
            self.sheet_combo2.setCurrentIndex(0)

    def on_sheet_selection_changed(self):
        """页签选择变化时的处理函数"""
        # 简单更新比较按钮状态
        self.update_compare_button_state()

    # def on_sheet_selected(self, is_file1=False, is_file2=False):
    #     self.loading_dialog = QProgressDialog("正在加载列名...", "取消", 0, 0, self)
    #     self.loading_dialog.setModal(True)
    #     self.loading_dialog.setWindowTitle("请稍候")
    #     self.loading_dialog.setCancelButton(None)  # 禁止取消
    #     self.loading_dialog.show()
    #     if is_file1 and self.file1:
    #         sheet_name = self.sheet_combo1.currentText()
    #         if not sheet_name:
    #             return
    #
    #         if hasattr(self, 'worker_sheet1') and self.worker_sheet1 and self.worker_sheet1.isRunning():
    #             self.worker_sheet1.quit()
    #             self.worker_sheet1.wait()
    #
    #         worker = LoadColumnWorker(self.file1, sheet_name=sheet_name)
    #         try:
    #             worker.columns_loaded.disconnect()
    #         except:
    #             pass
    #         try:
    #             worker.error_occurred.disconnect()
    #         except:
    #             pass
    #         worker.columns_loaded.connect(self.handle_columns_loaded)
    #         worker.error_occurred.connect(self.handle_column_error)
    #         self.worker_sheet1 = worker
    #         worker.start()
    #     elif is_file2 and self.file2:
    #         sheet_name = self.sheet_combo2.currentText()
    #         if not sheet_name:
    #             return
    #
    #         if hasattr(self, 'worker_sheet2') and self.worker_sheet2 and self.worker_sheet2.isRunning():
    #             self.worker_sheet2.quit()
    #             self.worker_sheet2.wait()
    #
    #         worker = LoadColumnWorker(self.file2, sheet_name=sheet_name)
    #         try:
    #             worker.columns_loaded.disconnect()
    #         except:
    #             pass
    #         try:
    #             worker.error_occurred.disconnect()
    #         except:
    #             pass
    #         worker.columns_loaded.connect(self.handle_columns_loaded)
    #         worker.error_occurred.connect(self.handle_column_error)
    #         self.worker_sheet2 = worker
    #         worker.start()
    #     self.update_compare_button_state()

    # def handle_columns_loaded(self, file_path, columns):
    #     self.loading_dialog.close()
    #     self.on_columns_loaded(file_path, columns)

    # def handle_column_error(self, error_msg):
    #     self.loading_dialog.close()
    #     self.on_column_error(error_msg)

    # def on_columns_loaded(self, file_path, columns):
    #     if file_path == self.file1:
    #         self.columns1 = columns
    #     elif file_path == self.file2:
    #         self.columns2 = columns
    #     else:
    #         return
    #
    #     if not self.sheet_combo1.currentText() or not self.sheet_combo2.currentText():
    #         return
    #
    #     source_cols = set(self.columns1)
    #     target_cols = set(self.columns2)
    #     common_cols = sorted(source_cols & target_cols)
    #     self.log_area.clear()
    #     if common_cols:
    #         self.log("已加载共同列。")
    #     else:
    #         self.log("两个文件没有共同的列，无法进行比较。")
    #
    #     self.update_compare_button_state()

    # def on_column_error(self, error_msg):
    #     self.log(f"列名读取错误：{error_msg}")

    def update_compare_button_state(self):
        sheet_selected = self.sheet_combo1.currentText() and self.sheet_combo2.currentText()
        if not sheet_selected:
            self.compare_btn.setEnabled(False)
            return

        self.compare_btn.setEnabled(True)

    def compare_files(self):
        if not self.file1 or not self.file2:
            self.log("请先选择两个 Excel 文件！")
            return
        sheet_name1 = self.sheet_combo1.currentText()
        sheet_name2 = self.sheet_combo2.currentText()
        if not sheet_name1 or not sheet_name2:
            self.log("请选择两个文件的页签！")
            return

        self.log_area.clear()
        self.summary_area.clear()
        self.export_btn.setEnabled(False)

        # 获取主键字段
        primary_keys = [field for field, rule in self.rules.items() if rule["is_primary"]]
        if not primary_keys:
            self.log("规则文件中未定义主键字段，请检查规则文件！")
            return
        self.loading_dialog = QProgressDialog("正在比较文件，请稍候...", None, 0, 0, self)
        self.loading_dialog.setWindowModality(Qt.WindowModal)
        self.loading_dialog.setWindowTitle("比较中")
        self.loading_dialog.setCancelButton(None)
        self.loading_dialog.show()

        self.worker = CompareWorker(self.file1, self.file2, sheet_name1, sheet_name2, primary_keys=primary_keys,
                                    rules=self.rules)
        self.worker.log_signal.connect(self.log)
        # 连接信号以在比较完成时关闭对话框
        self.worker.finished.connect(self.close_loading_dialog)
        self.worker.finished.connect(lambda: self.export_btn.setEnabled(True))
        self.worker.finished.connect(self.on_compare_finished)
        self.worker.start()

    def close_loading_dialog(self):
        """关闭加载对话框"""
        if self.loading_dialog:
            self.loading_dialog.close()
            self.loading_dialog = None

    def on_compare_finished(self):
        try:
            if hasattr(self.worker, 'summary'):
                self.summary_data = self.worker.summary
                primary_key = self.summary_data.get("primary_key", "主键")
                total_file1 = self.summary_data['total_file1']
                total_file2 = self.summary_data['total_file2']
                missing_count = self.summary_data['missing_count']
                extra_count = self.summary_data.get('extra_count', 0)
                common_count = self.summary_data['common_count']
                diff_count = self.summary_data['diff_count']
                equal_count = self.summary_data['equal_count']
                diff_ratio = self.summary_data['diff_ratio']
                missing_columns = self.summary_data.get("missing_columns", [])
                missing_columns_str = ", ".join(missing_columns) if missing_columns else "无"

                summary_text = (
                    f"📊 比对汇总报告\n"
                    f"--------------------------------\n"
                    f"• 总{primary_key}数量（表一）：{total_file1}\n"
                    f"• 总{primary_key}数量（表二）：{total_file2}\n"
                    f"• 表二中缺失的{primary_key}：{missing_count}\n"
                    f"• 表二中多出的{primary_key}：{extra_count}\n"
                    f"• 共同{primary_key}数量：{common_count}\n"
                    f"• 列不一致的{primary_key}数量：{diff_count}\n"
                    f"• 列一致的{primary_key}数量：{equal_count}\n"
                    f"• 表二中缺失的列：{missing_columns_str}\n"
                    f"--------------------------------\n"
                    f"• 差异数据占比：{diff_ratio:.2%}\n"
                )
                self.summary_area.setPlainText(summary_text)
                self.export_btn.setEnabled(True)
        except Exception as e:
            self.summary_area.setPlainText(f"❌ 显示汇总报告时发生错误：{str(e)}\n请查看比对日志了解详细信息。")
            self.export_btn.setEnabled(False)

    def export_report(self):
        """导出报告到一个Excel文件，包含两个sheet"""
        if not hasattr(self, 'worker') or not hasattr(self.worker, 'missing_rows') or not hasattr(self.worker,
                                                                                                  'diff_full_rows'):
            self.log("没有可导出的数据，请先执行比对！")
            return

        directory = QFileDialog.getExistingDirectory(self, "选择保存路径")
        if not directory:
            self.log("导出已取消。")
            return

        output_file = f"{directory}/资产比对结果报告.xlsx"

        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            if self.worker.missing_rows:
                missing_df = pd.DataFrame(self.worker.missing_rows)
                missing_df.to_excel(writer, sheet_name='表二缺失数据', index=False)
            if getattr(self.worker, 'extra_in_file2', None):
                extra_df = pd.DataFrame(self.worker.extra_in_file2)
                extra_df.to_excel(writer, sheet_name='表二多出数据', index=False)
            if self.worker.diff_full_rows:
                self._export_diff_data_with_highlight_to_sheet(writer, '列不一致数据', self.worker.diff_full_rows)

        self.log(f"✅ 已导出：{output_file}")

    def _export_diff_data_with_highlight_to_sheet(self, writer, sheet_name, diff_full_rows):
        """将差异数据导出到指定的 sheet，并高亮不一致的列"""
        if not diff_full_rows:
            return
        wb = writer.book
        ws = wb.create_sheet(sheet_name)

        first_target = diff_full_rows[0]["target"]
        headers = list(first_target.keys())

        ws.append(headers)

        red_fill = PatternFill(start_color="FFEE1111", end_color="FFEE1111", fill_type="solid")

        for row_data in diff_full_rows:
            target_data = row_data["target"]
            source_data = row_data["source"]

            target_row = [target_data.get(k, '') for k in headers]
            target_row_idx = ws.max_row + 1
            ws.append(target_row)

            for col_idx, key in enumerate(headers, start=1):
                val1 = source_data.get(key, '')
                val2 = target_data.get(key, '')
                val1 = CompareWorker.normalize_value(val1)
                val2 = CompareWorker.normalize_value(val2)

                if val1 != val2 and not (val1 == '' and val2 == ''):
                    ws.cell(row=target_row_idx, column=col_idx).fill = red_fill

    def log(self, message):
        """日志输出"""
        self.log_area.appendPlainText(message)


def exception_hook(exc_type, exc_value, exc_traceback):
    """全局异常钩子，防止崩溃"""
    try:
        ex = QApplication.instance().topLevelWidgets()[0]
        if hasattr(ex, "log"):
            error_message = "".join(traceback.format_exception(exc_type, exc_value, exc_traceback))
            logging.error(error_message)
            ex.log(f"❌ 发生异常：{exc_value}")
        else:
            sys.__excepthook__(exc_type, exc_value, exc_traceback)
    except:
        sys.__excepthook__(exc_type, exc_value, exc_traceback)


if __name__ == "__main__":
    sys.excepthook = exception_hook

    app = QApplication(sys.argv)
    icon_path = resource_path('icon.ico')
    app.setWindowIcon(QIcon(icon_path))
    ex = ExcelComparer()
    ex.show()
    exit_code = app.exec_()

    sys.exit(exit_code)
