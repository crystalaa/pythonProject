import sys
import pandas as pd
from PyQt5.QtGui import QIcon
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QFileDialog, QLabel, QVBoxLayout, QHBoxLayout, \
    QPlainTextEdit, QProgressBar, QTabWidget
from PyQt5.QtCore import QThread, pyqtSignal, Qt
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor


def read_excel_chunked(file_path, sheet_name, chunk_size=10000):
    """分块读取 Excel 文件，每次读取 chunk_size 行"""
    wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    columns = [cell.value for cell in next(ws.rows)]  # 获取列名
    chunk = []

    # 逐行读取文件
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3:  # 跳过前3行
            continue
        chunk.append(row)

        if len(chunk) == chunk_size:
            yield pd.DataFrame(chunk, columns=columns)  # 返回当前块
            chunk = []  # 清空当前块

    if chunk:  # 如果还有剩余行
        yield pd.DataFrame(chunk, columns=columns)


class CompareWorker(QThread):
    """用于在独立线程中执行比较操作"""
    log_signal = pyqtSignal(str)
    progress_signal = pyqtSignal(int)

    def __init__(self, file1, file2):
        super().__init__()
        self.file1 = file1
        self.file2 = file2
        self.missing_rows = []
        self.diff_full_rows = []
        self.summary = {}

    def run(self):
        """线程执行的主函数"""
        try:
            self.log_signal.emit("正在比较文件...")
            with ThreadPoolExecutor(max_workers=2) as executor:
                future1 = executor.submit(read_excel_chunked, self.file1, "附表1资产卡片期初数据收集模板")
                future2 = executor.submit(read_excel_chunked, self.file2, "附表1资产卡片期初数据收集模板")
                chunk_iter1 = future1.result()
                chunk_iter2 = future2.result()

            # 遍历所有块并逐行比对
            chunk_count = 0
            diff_count = 0
            for df1, df2 in zip(chunk_iter1, chunk_iter2):
                chunk_count += 1
                # 对齐资产编码为索引
                df1.set_index('资产编码', inplace=True)
                df2.set_index('资产编码', inplace=True)

                # 比对数据
                diff_df, missing_df = self.compare_chunks(df1, df2)

                # 记录差异
                self.diff_full_rows.extend(diff_df)
                self.missing_rows.extend(missing_df)

                diff_count += len(diff_df)

                # 更新进度条
                self.progress_signal.emit(int((chunk_count * 100) / (len(chunk_iter1) + len(chunk_iter2))))

            self.summary = {
                "diff_count": diff_count,
                "missing_count": len(self.missing_rows)
            }
            self.log_signal.emit("比对完成！")

        except Exception as e:
            self.log_signal.emit(f"发生错误：{str(e)}")

    def compare_chunks(self, df1, df2):
        """比较两个 DataFrame 中的差异"""
        diff_rows = []
        missing_rows = []

        # 获取共同的资产编码
        common_codes = df1.index.intersection(df2.index)

        # 找出缺失的资产编码
        missing_in_file2 = df1.index.difference(df2.index)
        missing_rows.extend(df1.loc[missing_in_file2].to_dict(orient='records'))

        # 找出差异行
        df1_common = df1.loc[common_codes]
        df2_common = df2.loc[common_codes]

        df1_compare = df1_common.astype(str).replace('nan', '')
        df2_compare = df2_common.astype(str).replace('nan', '')

        comparison = df1_compare.compare(df2_compare, align_axis=0)

        # 找出不一致的行
        mask = (df1_compare == df2_compare).all(axis=1)
        diff_df = df1_compare[~mask]

        for asset_code, row in diff_df.iterrows():
            diff_details = []
            for col in df1_compare.columns:
                val1 = df1_compare.loc[asset_code, col]
                val2 = df2_compare.loc[asset_code, col]
                if val1 != val2:
                    diff_details.append(f" - 列 [{col}] 不一致：源文件={val1}, 目标文件={val2}")
            if diff_details:
                diff_rows.append({
                    "资产编码": asset_code,
                    "diff_details": diff_details
                })
        return diff_rows, missing_rows


class ExcelComparer(QWidget):
    """主窗口类"""

    def __init__(self):
        super().__init__()
        self.file1 = ""
        self.file2 = ""
        self.initUI()
        self.worker = None
        self.summary_data = {}

    def initUI(self):
        """初始化用户界面"""
        self.setWindowTitle("Excel文件比较工具")
        self.resize(1000, 700)

        main_layout = QVBoxLayout()

        top_layout = QHBoxLayout()

        file_select_layout = QVBoxLayout()
        self.label1 = QLabel("未选择源文件")
        self.btn1 = QPushButton("选择源文件")
        self.btn1.clicked.connect(self.select_file1)

        self.label2 = QLabel("未选择目标文件")
        self.btn2 = QPushButton("选择目标文件")
        self.btn2.clicked.connect(self.select_file2)

        file_select_layout.addWidget(self.label1)
        file_select_layout.addWidget(self.btn1)
        file_select_layout.addWidget(self.label2)
        file_select_layout.addWidget(self.btn2)

        button_layout = QVBoxLayout()
        self.compare_btn = QPushButton("比较文件")
        self.compare_btn.setFixedWidth(150)
        self.compare_btn.clicked.connect(self.compare_files)

        self.export_btn = QPushButton("导出报告")
        self.export_btn.setFixedWidth(150)
        self.export_btn.setEnabled(False)
        self.export_btn.clicked.connect(self.export_report)

        button_layout.addStretch()
        button_layout.addWidget(self.compare_btn)
        button_layout.addWidget(self.export_btn)
        button_layout.addStretch()

        top_layout.addLayout(file_select_layout, 2)
        top_layout.addLayout(button_layout, 1)

        self.tab_widget = QTabWidget()
        self.log_area = QPlainTextEdit()
        self.log_area.setReadOnly(True)
        self.log_area.setStyleSheet("background-color: #f0f0f0;")

        self.summary_area = QPlainTextEdit()
        self.summary_area.setReadOnly(True)
        self.summary_area.setStyleSheet("background-color: #f0f0f0;")

        self.tab_widget.addTab(self.log_area, "比对日志")
        self.tab_widget.addTab(self.summary_area, "汇总报告")

        # 添加进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        self.progress_bar.setFixedHeight(20)
        self.progress_bar.setTextVisible(True)
        self.progress_bar.setAlignment(Qt.AlignCenter)

        main_layout.addLayout(top_layout)
        main_layout.addWidget(self.tab_widget)
        main_layout.addWidget(self.progress_bar)

        self.setLayout(main_layout)

    def select_file1(self):
        """选择源文件"""
        file, _ = QFileDialog.getOpenFileName(self, "选择 Excel 文件", "", "Excel 文件 (*.xlsx *.xls)")
        if file:
            self.file1 = file
            self.label1.setText(f"源文件: {file}")

    def select_file2(self):
        """选择目标文件"""
        file, _ = QFileDialog.getOpenFileName(self, "选择 Excel 文件", "", "Excel 文件 (*.xlsx *.xls)")
        if file:
            self.file2 = file
            self.label2.setText(f"目标文件: {file}")

    def compare_files(self):
        """开始比较文件"""
        if not self.file1 or not self.file2:
            self.log("请先选择两个 Excel 文件！")
            return

        self.log_area.clear()
        self.summary_area.clear()
        self.export_btn.setEnabled(False)

        # 创建并启动比较线程
        self.worker = CompareWorker(self.file1, self.file2)
        self.worker.log_signal.connect(self.log)
        self.worker.progress_signal.connect(self.update_progress)
        self.worker.finished.connect(lambda: self.progress_bar.setValue(100))
        self.worker.finished.connect(lambda: self.export_btn.setEnabled(True))
        self.worker.finished.connect(self.on_compare_finished)
        self.worker.start()

    def update_progress(self, value):
        """更新进度条"""
        self.progress_bar.setValue(value)

    def on_compare_finished(self):
        """比较完成后的处理"""
        if hasattr(self.worker, 'summary'):
            self.summary_data = self.worker.summary
            summary_text = (
                f"📊 比对汇总报告\n"
                f"--------------------------------\n"
                f"• 差异数据总数：{self.summary_data['diff_count']}\n"
                f"• 目标文件中缺失的资产编码：{self.summary_data['missing_count']}\n"
                f"--------------------------------\n"
            )
            self.summary_area.setPlainText(summary_text)
            self.export_btn.setEnabled(True)

    def export_report(self):
        """导出报告"""
        if not hasattr(self, 'worker') or not hasattr(self.worker, 'missing_rows') or not hasattr(self.worker,
                                                                                                  'diff_full_rows'):
            self.log("没有可导出的数据，请先执行比对！")
            return

        directory = QFileDialog.getExistingDirectory(self, "选择保存路径")
        if not directory:
            self.log("导出已取消。")
            return

        # 导出缺失资产编码
        if self.worker.missing_rows:
            missing_df = pd.DataFrame(self.worker.missing_rows)
            missing_df.to_excel(f"{directory}/目标文件中缺失的资产数据.xlsx", index=False)
            self.log("✅ 已导出：目标文件中缺失的资产数据.xlsx")

        # 导出列不一致数据
        if self.worker.diff_full_rows:
            self._export_diff_data_with_highlight(f"{directory}/目标文件_列不一致的整行数据.xlsx",
                                                  self.worker.diff_full_rows)
            self.log("✅ 已导出：目标文件_列不一致的整行数据.xlsx")

        self.log("📊 比对报告导出完成！")

    def _export_diff_data_with_highlight(self, file_path, diff_full_rows):
        """仅导出目标文件中不一致的行，并高亮不一致的列"""
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill

        wb = Workbook()
        ws = wb.active

        # 获取列顺序（以第一个目标行为准）
        first_target = diff_full_rows[0]["资产编码"]
        headers = list(first_target.keys())  # 保持原始列顺序

        if '资产编码' in headers:
            headers.remove('资产编码')
            headers.insert(1, '资产编码')  # 插入到第2列位置

        # 写入表头
        ws.append([headers[i] for i in range(len(headers))])

        red_fill = PatternFill(start_color="FFEE1111", end_color="FFEE1111", fill_type="solid")

        for row_data in diff_full_rows:
            target_data = row_data["target"]

            # 按照 headers 顺序构造目标行数据
            target_row = [target_data.get(k, '') for k in headers]
            target_row_idx = ws.max_row + 1
            ws.append(target_row)

            # 比较并高亮不一致的列（跳过资产编码列）
            source_data = row_data["source"]
            for col_idx, key in enumerate(headers, start=1):
                if key == '资产编码':
                    continue
                val1 = source_data.get(key, '')
                val2 = target_data.get(key, '')
                if val1 != val2:
                    ws.cell(row=target_row_idx, column=col_idx).fill = red_fill

        wb.save(file_path)

    def log(self, message):
        """日志输出"""
        self.log_area.appendPlainText(message)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    ex = ExcelComparer()
    ex.show()
    sys.exit(app.exec_())